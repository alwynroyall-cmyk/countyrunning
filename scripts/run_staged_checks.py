"""Run staged validation checks for the WRRL operational chain.

Stages:
1. Raw ingest validation
2. Raw -> audited consolidation validation
3. Audit generation validation
4. Main scoring regression validation
"""

from __future__ import annotations

import argparse
import json
import hashlib
import sys
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

from openpyxl import load_workbook

PROJECT_ROOT = Path(__file__).resolve().parents[1]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from league_scorer.audit import LeagueAuditor
from league_scorer.common_files import race_discovery_exclusions
from league_scorer.input_layout import build_input_paths
from league_scorer.main import LeagueScorer
from league_scorer.output_layout import ensure_output_subdirs
from league_scorer.race_processor import extract_race_number
from league_scorer.race_validation import validate_race_schema
from league_scorer.source_loader import discover_race_files, load_race_dataframe
from scripts.analyse_data_quality import analyse_season


@dataclass
class StageResult:
    stage: int
    name: str
    status: str
    message: str
    details: dict[str, Any] = field(default_factory=dict)


def _quality_success_pct(audited_summary: dict[str, Any]) -> float:
    completeness_values = [
        100.0 - float(audited_summary.get("blank_category_pct", 0.0)),
        100.0 - float(audited_summary.get("blank_name_pct", 0.0)),
        100.0 - float(audited_summary.get("blank_gender_pct", 0.0)),
        100.0 - float(audited_summary.get("invalid_time_pct", 0.0)),
    ]
    return round(sum(completeness_values) / len(completeness_values), 2)


def _resolve_data_root(explicit: Path | None) -> Path | None:
    if explicit is not None:
        return explicit

    prefs = Path.home() / ".wrrl_prefs.json"
    if prefs.exists():
        try:
            data = json.loads(prefs.read_text(encoding="utf-8"))
            value = data.get("data_root")
            if value:
                return Path(value)
        except Exception:
            return None
    return None


def _find_raw_dir(input_dir: Path) -> Path | None:
    candidate = build_input_paths(input_dir).raw_data_dir
    return candidate if candidate.exists() and candidate.is_dir() else None


def _discover_raw_race_files(raw_dir: Path) -> dict[int, list[Path]]:
    by_race: dict[int, list[Path]] = {}
    for pattern in ("*.xlsx", "*.xlsm", "*.xls", "*.csv"):
        for file in sorted(raw_dir.rglob(pattern)):
            race_num = extract_race_number(file.stem)
            if race_num is None:
                continue
            by_race.setdefault(race_num, []).append(file)
    return by_race


def _workbook_fingerprint(path: Path) -> dict[str, Any]:
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        sheets: dict[str, Any] = {}
        for ws in wb.worksheets:
            digest = hashlib.sha256()
            row_count = 0
            col_count = 0
            for row in ws.iter_rows(values_only=True):
                row_count += 1
                col_count = max(col_count, len(row))
                joined = "|".join("" if v is None else str(v) for v in row)
                digest.update(joined.encode("utf-8", errors="ignore"))
                digest.update(b"\n")
            sheets[ws.title] = {
                "rows": row_count,
                "cols": col_count,
                "sha256": digest.hexdigest(),
            }
        return {
            "workbook_name": path.name,
            "sheets": sheets,
        }
    finally:
        wb.close()


def run_checks(
    args: argparse.Namespace,
    progress_cb: Callable[[str], None] | None = None,
) -> tuple[bool, list[StageResult]]:
    def _progress(message: str) -> None:
        if progress_cb is None:
            return
        try:
            progress_cb(message)
        except Exception:
            return

    data_root = _resolve_data_root(args.data_root)
    if data_root is None:
        message = "No data root configured. Set --data-root or configure WRRL data root in app settings."
        if args.allow_missing_data:
            return True, [StageResult(0, "Preflight", "skipped", message)]
        return False, [StageResult(0, "Preflight", "failed", message)]

    input_dir = data_root / str(args.year) / "inputs"
    output_dir = data_root / str(args.year) / "outputs"
    ensure_output_subdirs(output_dir)

    if not input_dir.exists():
        message = f"Input directory not found: {input_dir}"
        if args.allow_missing_data:
            return True, [StageResult(0, "Preflight", "skipped", message)]
        return False, [StageResult(0, "Preflight", "failed", message)]

    results: list[StageResult] = []

    # Stage 1
    _progress("Stage 1/4: Validating raw ingest workbooks")
    raw_dir = _find_raw_dir(input_dir)
    if raw_dir is None:
        results.append(
            StageResult(
                1,
                "Raw Ingest Validation",
                "failed",
                "Raw data folder not found under inputs.",
                {"input_dir": str(input_dir)},
            )
        )
        return False, results

    raw_races = _discover_raw_race_files(raw_dir)
    schema_ok = 0
    schema_fail = 0
    schema_warnings = 0
    stage1_errors: list[str] = []

    for race_num, files in sorted(raw_races.items()):
        for file in files:
            try:
                df = load_race_dataframe(file)
                validation = validate_race_schema(df, file)
                schema_ok += 1
                schema_warnings += len(validation.issues)
            except Exception as exc:
                schema_fail += 1
                stage1_errors.append(f"Race {race_num} {file.name}: {exc}")

    status = "passed" if schema_fail == 0 else "failed"
    results.append(
        StageResult(
            1,
            "Raw Ingest Validation",
            status,
            f"Validated {schema_ok + schema_fail} raw workbook(s); failures={schema_fail}, warnings={schema_warnings}.",
            {
                "raw_dir": str(raw_dir),
                "races_detected": sorted(raw_races),
                "errors": stage1_errors[:50],
            },
        )
    )
    if status == "failed":
        return False, results

    # Stage 2
    _progress("Stage 2/4: Consolidating raw to audited and running quality gate")
    audited_dir = build_input_paths(input_dir).audited_dir
    audited_race_files = discover_race_files(audited_dir, excluded_names=race_discovery_exclusions())
    raw_race_nums = set(raw_races)
    audited_race_nums = set(audited_race_files)

    missing_from_audited = sorted(raw_race_nums - audited_race_nums)
    stage2_details: dict[str, Any] = {
        "missing_from_audited": missing_from_audited,
        "audited_files": {k: str(v) for k, v in audited_race_files.items()},
    }

    quality_output_dir = getattr(args, "data_quality_output_dir", Path("output") / "quality" / "data-quality")
    quality_threshold = float(getattr(args, "quality_gate_threshold", 80.0))
    quality_gate_passed = True
    try:
        quality_payload, quality_json, quality_md = analyse_season(args.year, data_root, quality_output_dir)
        audited_summary = quality_payload.get("audited_summary", {})
        hotspots = quality_payload.get("hotspots", {}).get("audited_blank_category", [])
        quality_success_pct = _quality_success_pct(audited_summary)
        quality_gate_passed = quality_success_pct >= quality_threshold
        stage2_details["data_quality"] = {
            "report_json": str(quality_json),
            "report_markdown": str(quality_md),
            "quality_gate_threshold_pct": quality_threshold,
            "quality_success_pct": quality_success_pct,
            "quality_gate_passed": quality_gate_passed,
            "audited_blank_category_pct": audited_summary.get("blank_category_pct", 0.0),
            "audited_invalid_time_pct": audited_summary.get("invalid_time_pct", 0.0),
            "top_blank_category_hotspots": hotspots[:3],
        }
    except Exception as exc:
        quality_gate_passed = False
        stage2_details["data_quality_error"] = str(exc)

    status = "passed" if (not missing_from_audited and quality_gate_passed) else "failed"
    gate_summary = f"quality gate={quality_threshold:.1f}%"
    if "data_quality" in stage2_details:
        gate_summary = (
            f"quality gate={quality_threshold:.1f}% "
            f"(observed={stage2_details['data_quality']['quality_success_pct']:.2f}%)."
        )
    if "data_quality_error" in stage2_details:
        gate_summary = f"quality gate unavailable ({stage2_details['data_quality_error']})."
    results.append(
        StageResult(
            2,
            "Raw To Audited Consolidation",
            status,
            f"Raw races={len(raw_race_nums)}, audited races={len(audited_race_nums)}; {gate_summary}",
            stage2_details,
        )
    )
    if status == "failed":
        return False, results

    # Stage 3
    _progress("Stage 3/4: Generating audit workbook")
    auditor = LeagueAuditor(input_dir=input_dir, output_dir=output_dir, year=args.year)
    audit_path = auditor.run(race_files=audited_race_files)

    total_issue_count = sum(len(v) for v in auditor.all_race_issues.values())
    actionable_issue_count = int(getattr(auditor, "latest_actionable_issue_count", total_issue_count))
    results.append(
        StageResult(
            3,
            "Audit Generation Validation",
            "passed",
            f"Audit workbook generated with {actionable_issue_count} actionable issue row(s).",
            {
                "audit_workbook": str(audit_path),
                "issue_count": actionable_issue_count,
                "total_issue_count": total_issue_count,
            },
        )
    )

    # Stage 4
    _progress("Stage 4/4: Running scoring regression")
    scorer = LeagueScorer(input_dir=input_dir, output_dir=output_dir, year=args.year)
    warnings = scorer.run(race_files=audited_race_files)

    from league_scorer.graphical.results_workbook import find_latest_results_workbook

    latest_results = find_latest_results_workbook(output_dir)
    if latest_results is None:
        results.append(
            StageResult(4, "Main Scoring Regression", "failed", "No results workbook was generated.")
        )
        return False, results

    _progress("Stage 4/4: Fingerprinting results workbook")
    fingerprint = _workbook_fingerprint(latest_results)
    baseline_path = Path(args.baseline_file)
    baseline_exists = baseline_path.exists()

    if args.write_baseline or not baseline_exists:
        baseline_path.parent.mkdir(parents=True, exist_ok=True)
        baseline_path.write_text(json.dumps(fingerprint, indent=2), encoding="utf-8")
        results.append(
            StageResult(
                4,
                "Main Scoring Regression",
                "passed",
                "Baseline written from current results workbook.",
                {
                    "baseline": str(baseline_path),
                    "warnings": warnings,
                    "results_workbook": str(latest_results),
                },
            )
        )
        return True, results

    expected = json.loads(baseline_path.read_text(encoding="utf-8"))
    _progress("Stage 4/4: Comparing results against baseline")
    status = "passed" if expected.get("sheets") == fingerprint.get("sheets") else "failed"
    message = "Results fingerprint matches baseline." if status == "passed" else "Results fingerprint differs from baseline."

    details = {
        "baseline": str(baseline_path),
        "results_workbook": str(latest_results),
        "warnings": warnings,
    }

    if status == "failed":
        expected_sheets = expected.get("sheets", {})
        current_sheets = fingerprint.get("sheets", {})
        changed = sorted(
            set(expected_sheets) | set(current_sheets)
        )
        diffs: list[dict[str, Any]] = []
        for sheet in changed:
            if expected_sheets.get(sheet) != current_sheets.get(sheet):
                diffs.append(
                    {
                        "sheet": sheet,
                        "expected": expected_sheets.get(sheet),
                        "current": current_sheets.get(sheet),
                    }
                )
        details["diffs"] = diffs[:50]

    results.append(StageResult(4, "Main Scoring Regression", status, message, details))
    return status == "passed", results


def write_report(results: list[StageResult], report_dir: Path, success: bool) -> None:
    report_dir.mkdir(parents=True, exist_ok=True)
    payload = {
        "generated_at": datetime.utcnow().isoformat() + "Z",
        "success": success,
        "stages": [asdict(item) for item in results],
    }
    (report_dir / "staged_checks_report.json").write_text(json.dumps(payload, indent=2), encoding="utf-8")

    lines = [
        "# WRRL Staged Checks Report",
        "",
        f"Generated: {payload['generated_at']}",
        f"Success: {success}",
        "",
    ]
    for item in results:
        lines.append(f"## Stage {item.stage} - {item.name}")
        lines.append(f"Status: {item.status}")
        lines.append("")
        lines.append(item.message)
        lines.append("")

        if item.stage == 2:
            quality = item.details.get("data_quality") if isinstance(item.details, dict) else None
            if isinstance(quality, dict):
                lines.append("### Stage 2 Quality Gate Details")
                lines.append("")
                lines.append(
                    f"- Threshold: {quality.get('quality_gate_threshold_pct', 80.0)}%"
                )
                lines.append(
                    f"- Observed Success: {quality.get('quality_success_pct', 0.0)}%"
                )
                lines.append(
                    f"- Gate Passed: {quality.get('quality_gate_passed', False)}"
                )
                lines.append(
                    f"- Audited Blank Category %: {quality.get('audited_blank_category_pct', 0.0)}"
                )
                lines.append(
                    f"- Audited Invalid Time %: {quality.get('audited_invalid_time_pct', 0.0)}"
                )
                lines.append(
                    f"- Data Quality Report: {quality.get('report_markdown', '')}"
                )
                lines.append("")

                hotspots = quality.get("top_blank_category_hotspots", [])
                if hotspots:
                    lines.append("### Top Blank Category Hotspots")
                    lines.append("")
                    lines.append("| File | Blank Category % | Blank Count |")
                    lines.append("| --- | ---: | ---: |")
                    for hotspot in hotspots:
                        file_name = Path(str(hotspot.get("file", ""))).name
                        lines.append(
                            f"| {file_name} | {hotspot.get('blank_category_pct', 0.0)} | {hotspot.get('blank_category', 0)} |"
                        )
                    lines.append("")

                suggestions: list[str] = []
                blank_pct = float(quality.get("audited_blank_category_pct", 0.0))
                invalid_time_pct = float(quality.get("audited_invalid_time_pct", 0.0))

                if blank_pct > 0:
                    suggestions.append(
                        "Add or tighten category mapping rules in audit normalization for common missing patterns (for example age-band aliases and empty placeholders)."
                    )
                    suggestions.append(
                        "Apply targeted source cleanup to the top hotspot files first, then re-run staged checks to confirm reduction."
                    )

                if invalid_time_pct > 0:
                    suggestions.append(
                        "Normalize common time formatting variants in source files before parsing (for example stray spaces, punctuation, and hh:mm:ss variants)."
                    )

                if hotspots:
                    suggestions.append(
                        "For each hotspot race, add a small race-specific pre-clean step in consolidation to infer category from known metadata where safe."
                    )

                if suggestions:
                    lines.append("### Suggested Fixes")
                    lines.append("")
                    for suggestion in suggestions:
                        lines.append(f"- {suggestion}")
                    lines.append("")

            if isinstance(item.details, dict) and item.details.get("data_quality_error"):
                lines.append("### Stage 2 Quality Gate Details")
                lines.append("")
                lines.append(f"- Data Quality Error: {item.details['data_quality_error']}")
                lines.append("")
    (report_dir / "staged_checks_report.md").write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Run WRRL staged operational checks.")
    parser.add_argument("--year", type=int, default=1999)
    parser.add_argument("--data-root", type=Path, default=None)
    parser.add_argument(
        "--report-dir",
        type=Path,
        default=Path("output") / "quality" / "staged-checks",
    )
    parser.add_argument(
        "--baseline-file",
        type=Path,
        default=Path("tests") / "baselines" / "season_1999_results_baseline.json",
    )
    parser.add_argument(
        "--data-quality-output-dir",
        type=Path,
        default=Path("output") / "quality" / "data-quality",
    )
    parser.add_argument(
        "--quality-gate-threshold",
        type=float,
        default=80.0,
        help="Minimum quality success percentage required to proceed past Stage 2.",
    )
    parser.add_argument("--write-baseline", action="store_true")
    parser.add_argument("--allow-missing-data", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    success, results = run_checks(args)
    write_report(results, args.report_dir, success)

    for item in results:
        print(f"[stage {item.stage}] {item.name}: {item.status} - {item.message}")

    return 0 if success else 1


if __name__ == "__main__":
    raise SystemExit(main())
