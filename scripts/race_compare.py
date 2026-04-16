"""Standalone Excel workbook comparison utility.

Usage:
    python scripts/race_compare.py file1.xlsx file2.xlsx [--output Race_Compare.xlsx]

This tool matches race sheets between two workbooks and writes points differences into a new workbook.
"""

from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Iterable

import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

_POINT_KEYWORDS = ("point", "pts")
_SEX_KEYWORDS = ("sex", "gender")
_RUNNER_KEYWORDS = ("runner name", "runner", "name", "athlete")
_CLUB_KEYWORDS = ("club", "team")


def parse_args(argv: Iterable[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Match race sheets across two workbooks and export point differences into a new workbook.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("file1", type=Path, nargs="?", help="Path to the first workbook")
    parser.add_argument("file2", type=Path, nargs="?", help="Path to the second workbook")
    parser.add_argument("--output", type=Path, default=Path("Race_Compare.xlsx"), help="Output workbook path")
    parser.add_argument("--gui", action="store_true", help="Launch a GUI for selecting files and output path")
    return parser.parse_args(argv)


def validate_file(path: Path) -> None:
    if not path.exists():
        raise FileNotFoundError(f"File not found: {path}")
    if not path.is_file():
        raise FileNotFoundError(f"Not a file: {path}")
    if path.suffix.lower() not in {".xls", ".xlsx"}:
        raise ValueError(f"Unsupported file type: {path}")


def run_comparison(file1: Path, file2: Path, output_path: Path) -> dict[str, object]:
    validate_file(file1)
    validate_file(file2)

    sheets1 = load_sheet_names(file1)
    sheets2 = load_sheet_names(file2)
    pairs = match_sheets(sheets1, sheets2)
    if not pairs:
        raise ValueError("No matching sheets found between the two workbooks.")

    comparisons: list[tuple[str, str, list[dict[str, object]]]] = []
    diagnostics: list[dict[str, object]] = []
    for sheet1, sheet2 in pairs:
        headers1, rows1 = load_sheet_data(file1, sheet1)
        headers2, rows2 = load_sheet_data(file2, sheet2)
        points_spec1 = find_points_columns(headers1)
        points_spec2 = find_points_columns(headers2)
        if points_spec1 is None or points_spec2 is None:
            diagnostics.append(
                {
                    "sheet1": sheet1,
                    "sheet2": sheet2,
                    "skipped": True,
                    "reason": "missing points columns",
                }
            )
            continue
        runner_col1 = find_column(headers1, _RUNNER_KEYWORDS)
        club_col1 = find_column(headers1, _CLUB_KEYWORDS)
        runner_col2 = find_column(headers2, _RUNNER_KEYWORDS)
        club_col2 = find_column(headers2, _CLUB_KEYWORDS)
        sex_col1 = find_sex_column(headers1)
        sex_col2 = find_sex_column(headers2)
        scored1 = build_scored_rows(rows1, runner_col1, club_col1, points_spec1, sex_col1)
        scored2 = build_scored_rows(rows2, runner_col2, club_col2, points_spec2, sex_col2)
        diff_rows = compare_sheet_rows(scored1, scored2, points_spec1, points_spec2, sex_col1, sex_col2)
        comparisons.append((sheet1, sheet2, diff_rows))
        diagnostics.append(
            {
                "sheet1": sheet1,
                "sheet2": sheet2,
                "runner_col1": runner_col1,
                "club_col1": club_col1,
                "runner_col2": runner_col2,
                "club_col2": club_col2,
                "points_spec1": points_spec1,
                "points_spec2": points_spec2,
                "scored1": len(scored1),
                "scored2": len(scored2),
                "diff_rows": len(diff_rows),
            }
        )

    if not comparisons:
        raise ValueError("No comparable point-scored race sheets were found.")

    write_output_workbook(output_path, file1, file2, comparisons)
    return {"pairs": pairs, "diagnostics": diagnostics, "output": str(output_path.resolve())}


def launch_gui() -> int:
    try:
        from PySide6.QtCore import Qt
        from PySide6.QtWidgets import (
            QApplication,
            QComboBox,
            QGridLayout,
            QHBoxLayout,
            QLabel,
            QLineEdit,
            QMainWindow,
            QPushButton,
            QTextEdit,
            QVBoxLayout,
            QWidget,
            QFileDialog,
            QMessageBox,
        )
    except ImportError:
        print("Error: PySide6 is required for GUI mode.", file=sys.stderr)
        return 1

    class CompareWindow(QMainWindow):
        def __init__(self):
            super().__init__()
            self.setWindowTitle("Race Comparison GUI")
            self._build_ui()

        def _build_ui(self) -> None:
            central = QWidget(self)
            self.setCentralWidget(central)
            layout = QVBoxLayout(central)

            grid = QGridLayout()
            layout.addLayout(grid)

            self.path1_edit = QLineEdit(self)
            self.path2_edit = QLineEdit(self)
            self.output_edit = QLineEdit(self)
            self.output_edit.setText(str(Path("Race_Compare.xlsx")))

            grid.addWidget(QLabel("File 1:"), 0, 0)
            grid.addWidget(self.path1_edit, 0, 1)
            browse1 = QPushButton("Browse…", self)
            grid.addWidget(browse1, 0, 2)

            grid.addWidget(QLabel("File 2:"), 1, 0)
            grid.addWidget(self.path2_edit, 1, 1)
            browse2 = QPushButton("Browse…", self)
            grid.addWidget(browse2, 1, 2)

            grid.addWidget(QLabel("Output:"), 2, 0)
            grid.addWidget(self.output_edit, 2, 1)
            browse_out = QPushButton("Browse…", self)
            grid.addWidget(browse_out, 2, 2)

            button_layout = QHBoxLayout()
            layout.addLayout(button_layout)
            self.compare_button = QPushButton("Generate Output", self)
            self.clear_button = QPushButton("Clear", self)
            button_layout.addWidget(self.compare_button)
            button_layout.addWidget(self.clear_button)

            self.result_view = QTextEdit(self)
            self.result_view.setReadOnly(True)
            layout.addWidget(self.result_view)

            browse1.clicked.connect(self._choose_file1)
            browse2.clicked.connect(self._choose_file2)
            browse_out.clicked.connect(self._choose_output)
            self.compare_button.clicked.connect(self._on_compare)
            self.clear_button.clicked.connect(self._on_clear)
            self.enable_drop(self.path1_edit)
            self.enable_drop(self.path2_edit)
            self.enable_drop(self.output_edit)

            self.resize(900, 400)

        def enable_drop(self, widget):
            widget.setAcceptDrops(True)
            widget.dragEnterEvent = lambda event: event.accept() if event.mimeData().hasUrls() else event.ignore()
            widget.dropEvent = self._make_drop_handler(widget)

        def _make_drop_handler(self, widget):
            def handler(event):
                urls = [url.toLocalFile() for url in event.mimeData().urls() if url.isLocalFile()]
                if urls:
                    widget.setText(urls[0])
                    if hasattr(widget, "textChanged"):
                        widget.textChanged.emit(widget.text())
            return handler

        def _choose_file1(self) -> None:
            path, _ = QFileDialog.getOpenFileName(self, "Choose first workbook", filter="Excel Files (*.xlsx *.xls)")
            if path:
                self.path1_edit.setText(path)

        def _choose_file2(self) -> None:
            path, _ = QFileDialog.getOpenFileName(self, "Choose second workbook", filter="Excel Files (*.xlsx *.xls)")
            if path:
                self.path2_edit.setText(path)

        def _choose_output(self) -> None:
            path, _ = QFileDialog.getSaveFileName(self, "Choose output workbook", filter="Excel Files (*.xlsx)")
            if path:
                self.output_edit.setText(path)

        def _on_compare(self) -> None:
            file1 = Path(self.path1_edit.text())
            file2 = Path(self.path2_edit.text())
            output_file = Path(self.output_edit.text())
            if not file1.exists() or not file2.exists():
                QMessageBox.warning(self, "Missing files", "Please select both input workbooks.")
                return
            try:
                diagnostics = run_comparison(file1, file2, output_file)
                self.result_view.append(f"Wrote comparison workbook: {output_file}")
                self.result_view.append("--- Comparison sanity check ---")
                self.result_view.append(f"Matched sheet pairs: {len(diagnostics.get('pairs', []))}")
                for detail in diagnostics.get("diagnostics", []):
                    if detail.get("skipped"):
                        self.result_view.append(
                            f"Skipped {detail['sheet1']} vs {detail['sheet2']}: {detail['reason']}"
                        )
                    else:
                        self.result_view.append(
                            f"{detail['sheet1']} vs {detail['sheet2']}: "
                            f"runner1={detail['runner_col1'] or 'UNKNOWN'}, club1={detail['club_col1'] or 'UNKNOWN'}, "
                            f"runner2={detail['runner_col2'] or 'UNKNOWN'}, club2={detail['club_col2'] or 'UNKNOWN'}, "
                            f"scored1={detail['scored1']}, scored2={detail['scored2']}, diffs={detail['diff_rows']}"
                        )
                self.result_view.append(f"Output workbook: {diagnostics.get('output')}")
            except Exception as exc:
                self.result_view.append(f"Comparison failed: {exc}")

        def _on_clear(self) -> None:
            self.path1_edit.clear()
            self.path2_edit.clear()
            self.output_edit.setText(str(Path("Race_Compare.xlsx")))
            self.result_view.clear()

    app = QApplication(sys.argv)
    window = CompareWindow()
    window.show()
    return app.exec()


def load_sheet_names(path: Path) -> list[str]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def normalize_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[^a-z0-9]+", " ", name.lower())
    return cleaned.strip()


def extract_sheet_number(name: str) -> str | None:
    match = re.search(r"(\d+)", name)
    return match.group(1) if match else None


def is_race_sheet(name: str) -> bool:
    normalized = normalize_sheet_name(name)
    return bool(re.search(r"\b(race|raceno)\b.*\d+|\d+.*\b(race|raceno)\b", normalized))


def prioritize_raceSheets(sheet_names: list[str]) -> list[str]:
    race_sheet_names = [name for name in sheet_names if is_race_sheet(name)]
    return race_sheet_names if race_sheet_names else sheet_names


def match_sheets(sheet_names1: list[str], sheet_names2: list[str]) -> list[tuple[str, str]]:
    sheet_names1 = prioritize_raceSheets(sheet_names1)
    sheet_names2 = prioritize_raceSheets(sheet_names2)
    matched: list[tuple[str, str]] = []
    remaining2 = sheet_names2.copy()

    for name1 in sheet_names1:
        if name1 in remaining2:
            matched.append((name1, name1))
            remaining2.remove(name1)

    for name1 in sheet_names1:
        if any(name1 == a for a, _ in matched):
            continue
        num1 = extract_sheet_number(name1)
        if num1:
            candidate = next((name2 for name2 in remaining2 if extract_sheet_number(name2) == num1), None)
            if candidate:
                matched.append((name1, candidate))
                remaining2.remove(candidate)

    for name1 in sheet_names1:
        if any(name1 == a for a, _ in matched):
            continue
        normalized1 = normalize_sheet_name(name1)
        candidate = next(
            (
                name2
                for name2 in remaining2
                if normalize_sheet_name(name2) == normalized1
            ),
            None,
        )
        if candidate:
            matched.append((name1, candidate))
            remaining2.remove(candidate)

    return matched


def load_sheet_data(path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, object]]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Sheet not found: {sheet_name}")
        ws = workbook[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return [], []

    header_index = 0
    for idx, row in enumerate(rows):
        headers = ["" if value is None else str(value).strip().lower() for value in row]
        has_runner = any(any(keyword in cell for keyword in _RUNNER_KEYWORDS) for cell in headers)
        has_points = any(any(keyword in cell for keyword in _POINT_KEYWORDS) for cell in headers)
        has_sex_points = "male" in headers or "female" in headers
        if has_runner and (has_points or has_sex_points):
            header_index = idx
            break

    raw_headers = ["" if value is None else str(value).strip() for value in rows[header_index]]
    unique_headers: list[str] = []
    header_counts: dict[str, int] = {}
    for header in raw_headers:
        key = header.lower()
        index = header_counts.get(key, 0)
        header_counts[key] = index + 1
        if index == 0:
            unique_headers.append(header)
        else:
            unique_headers.append(f"{header} ({index + 1})")

    headers = [header.lower() for header in unique_headers]
    data_rows: list[dict[str, object]] = []
    for row in rows[header_index + 1 :]:
        row_dict = {headers[idx]: row[idx] for idx in range(min(len(headers), len(row)))}
        data_rows.append(row_dict)
    return headers, data_rows


def find_column(headers: list[str], keywords: tuple[str, ...]) -> str | None:
    lower_headers = [h.lower() for h in headers]
    for keyword in keywords:
        for header, lower_header in zip(headers, lower_headers):
            if lower_header == keyword:
                return header
    for keyword in keywords:
        for header, lower_header in zip(headers, lower_headers):
            if keyword in lower_header:
                return header
    return None


def find_points_columns(headers: list[str]) -> dict[str, str] | str | None:
    lower_headers = [h.lower() for h in headers]
    male_points = next((header for header, lower_header in zip(headers, lower_headers) if lower_header == "male points"), None)
    female_points = next((header for header, lower_header in zip(headers, lower_headers) if lower_header == "female points"), None)
    if male_points or female_points:
        return {"male": male_points or "male", "female": female_points or "female"}

    male = next((header for header, lower_header in zip(headers, lower_headers) if lower_header == "male"), None)
    female = next((header for header, lower_header in zip(headers, lower_headers) if lower_header == "female"), None)
    if male or female:
        return {"male": male or "male", "female": female or "female"}

    return find_column(headers, _POINT_KEYWORDS)


def find_sex_column(headers: list[str]) -> str | None:
    return find_column(headers, _SEX_KEYWORDS)


def parse_points(value: object) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        digits = re.findall(r"[-+]?[0-9]*\.?[0-9]+", text)
        return float(digits[0]) if digits else 0.0


def normalize_text(value: str) -> str:
    cleaned = re.sub(r"[^a-z0-9 ]+", " ", value.lower())
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    return cleaned


def normalize_club_name(value: str) -> str:
    club = normalize_text(value)
    club = re.sub(r"\b(rc|running club|running clb|athletic club|athletics club|ac)$", "", club).strip()
    club = re.sub(r"\bclub$", "", club).strip()
    return club


def normalize_runner_name(value: str) -> str:
    normalized = re.sub(r"[^a-z0-9 ]+", " ", value.lower())
    normalized = re.sub(r"\s+", " ", normalized).strip()
    return normalized


def row_key(row: dict[str, object], runner_col: str | None, club_col: str | None) -> tuple[str, str]:
    runner = normalize_runner_name(str(row.get(runner_col, "") if runner_col else ""))
    club = normalize_club_name(str(row.get(club_col, "") if club_col else ""))
    return runner, club


def get_row_points(row: dict[str, object], points_spec: dict[str, str] | str | None, sex_col: str | None) -> float:
    if row is None or points_spec is None:
        return 0.0
    if isinstance(points_spec, dict):
        sex = str(row.get(sex_col, "") if sex_col else "").strip().upper()
        if sex.startswith("M") and points_spec.get("male"):
            return parse_points(row.get(points_spec["male"]))
        if sex.startswith("F") and points_spec.get("female"):
            return parse_points(row.get(points_spec["female"]))
        # fallback to whichever column has a value
        if points_spec.get("male") and parse_points(row.get(points_spec["male"])) > 0:
            return parse_points(row.get(points_spec["male"]))
        if points_spec.get("female") and parse_points(row.get(points_spec["female"])) > 0:
            return parse_points(row.get(points_spec["female"]))
        return 0.0
    return parse_points(row.get(points_spec))


def build_scored_rows(
    rows: list[dict[str, object]],
    runner_col: str | None,
    club_col: str | None,
    points_spec: dict[str, str] | str | None,
    sex_col: str | None = None,
) -> dict[tuple[str, str], dict[str, object]]:
    scored: dict[tuple[str, str], dict[str, object]] = {}
    for row in rows:
        points = get_row_points(row, points_spec, sex_col)
        if points <= 0:
            continue
        runner = str(row.get(runner_col, "") if runner_col else "").strip().lower()
        if not runner:
            continue
        key = row_key(row, runner_col, club_col)
        if key in scored:
            continue
        scored[key] = row
    return scored


def compare_sheet_rows(
    rows1: dict[tuple[str, str], dict[str, object]],
    rows2: dict[tuple[str, str], dict[str, object]],
    points_spec1: dict[str, str] | str | None,
    points_spec2: dict[str, str] | str | None,
    sex_col1: str | None,
    sex_col2: str | None,
) -> list[dict[str, object]]:
    diff_rows: list[dict[str, object]] = []
    all_keys = set(rows1) | set(rows2)
    for key in sorted(all_keys):
        row1 = rows1.get(key)
        row2 = rows2.get(key)
        points1 = get_row_points(row1, points_spec1, sex_col1) if row1 else 0.0
        points2 = get_row_points(row2, points_spec2, sex_col2) if row2 else 0.0
        if points1 == points2:
            continue
        runner = key[0].title() if key[0] else ""
        club = key[1].title() if key[1] else ""
        status = "Changed"
        if row1 is None:
            status = "Only in second workbook"
        elif row2 is None:
            status = "Only in first workbook"
        diff_rows.append(
            {
                "Runner": runner,
                "Club": club,
                "Points 1": points1,
                "Points 2": points2,
                "Difference": points2 - points1,
                "Status": status,
            }
        )
    return diff_rows


def safe_sheet_title(name: str, index: int) -> str:
    title = re.sub(r"[^A-Za-z0-9 _-]", "", name)[:25].strip()
    if not title:
        title = f"Race{index}"
    return title[:31]


def write_output_workbook(
    output_path: Path,
    file1: Path,
    file2: Path,
    comparisons: list[tuple[str, str, list[dict[str, object]]]],
) -> None:
    workbook = Workbook()
    summary: Worksheet = workbook.active
    summary.title = "Summary"
    summary.append(["Generated at", datetime.utcnow().replace(microsecond=0).isoformat() + "Z"])
    summary.append(["Output file", str(output_path.resolve())])
    summary.append(["File 1", str(file1)])
    summary.append(["File 2", str(file2)])
    summary.append([])
    summary.append(["Sheet 1", "Sheet 2", "Diff Rows"])

    for sheet1, sheet2, diff_rows in comparisons:
        summary.append([sheet1, sheet2, len(diff_rows)])

    for index, (sheet1, sheet2, diff_rows) in enumerate(comparisons, start=1):
        title = safe_sheet_title(f"{sheet1} vs {sheet2}", index)
        sheet = workbook.create_sheet(title)
        headers = ["Runner", "Club", "Points 1", "Points 2", "Difference", "Status"]
        sheet.append(headers)
        if not diff_rows:
            sheet.append(["No score differences found", "", "", "", "", ""])
            continue
        for row in diff_rows:
            sheet.append([row[h] for h in headers])

    workbook.save(output_path)


def main(argv: Iterable[str] | None = None) -> int:
    args = parse_args(argv)

    if args.gui or (args.file1 is None and args.file2 is None):
        return launch_gui()

    if args.file1 is None or args.file2 is None:
        print("Error: two workbook paths must be provided unless --gui is used.", file=sys.stderr)
        return 1

    try:
        run_comparison(args.file1, args.file2, args.output)
        print(f"Wrote comparison workbook: {args.output}")
        return 0
    except Exception as exc:
        print(f"Error: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
