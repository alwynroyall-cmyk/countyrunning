"""Create a deterministic anonymised test season from a real season input set.

Default usage copies:
  data/2025/inputs -> data/1999/inputs

The same real runner name is mapped to the same anonymised name across all files.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import shutil
from pathlib import Path

from openpyxl import load_workbook

FIRST_NAMES = [
    "Alex", "Avery", "Bailey", "Blake", "Casey", "Charlie", "Cory", "Dale", "Devon", "Drew",
    "Eden", "Elliot", "Emery", "Finley", "Frankie", "Georgie", "Harley", "Hayden", "Indy", "Jaden",
    "Jamie", "Jesse", "Jordan", "Jules", "Kai", "Kendall", "Lane", "Logan", "Marley", "Morgan",
    "Nico", "Noel", "Oakley", "Parker", "Quinn", "Reese", "Riley", "River", "Robin", "Rowan",
    "Sage", "Sam", "Sawyer", "Sky", "Sydney", "Taylor", "Toby", "Winter", "Zion", "Zephyr",
]

LAST_NAMES = [
    "Adair", "Alden", "Arden", "Ashby", "Aster", "Barden", "Blythe", "Briar", "Caelan", "Caldwell",
    "Carey", "Carlisle", "Corwin", "Darby", "Dawson", "Ellery", "Ellis", "Emerson", "Farrow", "Fletcher",
    "Hadley", "Harlow", "Harper", "Hollis", "Ives", "Jensen", "Keaton", "Kendrick", "Kingsley", "Lennox",
    "Marlow", "Merrick", "Monroe", "Nolan", "Oakley", "Palmer", "Peyton", "Quincy", "Ramsey", "Reagan",
    "Ridley", "Sawyer", "Shaw", "Sutton", "Teagan", "Vale", "Vaughn", "Winslow", "Yardley", "Zeller",
]

FULL_NAME_HEADERS = {
    "name", "runner", "runner name", "raw name", "preferred name", "athlete", "athlete name",
}
FIRST_NAME_HEADERS = {"first", "first name", "forename", "given name"}
LAST_NAME_HEADERS = {"last", "last name", "surname", "family name"}


class NameMapper:
    def __init__(self, seed: str) -> None:
        self.seed = seed

    def _digest(self, key: str) -> bytes:
        return hashlib.sha256(f"{self.seed}|{key}".encode("utf-8")).digest()

    def map_name(self, raw_name: str) -> tuple[str, str, str]:
        key = normalize_name_key(raw_name)
        if not key:
            return "", "", ""

        digest = self._digest(key)
        first = FIRST_NAMES[digest[0] % len(FIRST_NAMES)]
        last = LAST_NAMES[digest[1] % len(LAST_NAMES)]
        suffix = f"{int.from_bytes(digest[2:4], 'big') % 1000:03d}"

        anon_last = f"{last}{suffix}"
        full = f"{first} {anon_last}"
        return first, anon_last, full


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return " ".join(str(value).strip().lower().split())


def normalize_name_key(value: object) -> str:
    if value is None:
        return ""
    text = " ".join(str(value).strip().split())
    if not text:
        return ""
    return text.lower()


def anonymise_workbook(source_path: Path, target_path: Path, mapper: NameMapper) -> int:
    workbook = load_workbook(source_path)
    renamed = 0

    try:
        for sheet in workbook.worksheets:
            if sheet.max_row < 2:
                continue

            headers = [normalize_header(cell.value) for cell in sheet[1]]
            full_name_cols = [idx + 1 for idx, h in enumerate(headers) if h in FULL_NAME_HEADERS]
            first_col = next((idx + 1 for idx, h in enumerate(headers) if h in FIRST_NAME_HEADERS), None)
            last_col = next((idx + 1 for idx, h in enumerate(headers) if h in LAST_NAME_HEADERS), None)

            for row_idx in range(2, sheet.max_row + 1):
                if first_col and last_col:
                    first_value = sheet.cell(row=row_idx, column=first_col).value
                    last_value = sheet.cell(row=row_idx, column=last_col).value
                    combined = f"{first_value or ''} {last_value or ''}".strip()
                    if combined:
                        anon_first, anon_last, _ = mapper.map_name(combined)
                        if anon_first and anon_last:
                            sheet.cell(row=row_idx, column=first_col).value = anon_first
                            sheet.cell(row=row_idx, column=last_col).value = anon_last
                            renamed += 1

                for col_idx in full_name_cols:
                    original = sheet.cell(row=row_idx, column=col_idx).value
                    if original is None:
                        continue
                    original_text = str(original).strip()
                    if not original_text:
                        continue
                    _, _, anon_full = mapper.map_name(original_text)
                    if anon_full:
                        sheet.cell(row=row_idx, column=col_idx).value = anon_full
                        renamed += 1

        target_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(target_path)
    finally:
        workbook.close()

    return renamed


def copy_and_anonymise_inputs(source_input_dir: Path, target_input_dir: Path, mapper: NameMapper, clear_target: bool) -> None:
    if not source_input_dir.exists():
        raise FileNotFoundError(f"Source input folder not found: {source_input_dir}")

    if clear_target and target_input_dir.exists():
        shutil.rmtree(target_input_dir)

    target_input_dir.mkdir(parents=True, exist_ok=True)

    total_files = 0
    workbook_files = 0
    total_names = 0

    for source_file in sorted(source_input_dir.rglob("*")):
        if source_file.is_dir():
            continue

        relative = source_file.relative_to(source_input_dir)
        total_files += 1
        target_file = target_input_dir / relative
        target_file.parent.mkdir(parents=True, exist_ok=True)

        if source_file.suffix.lower() == ".xlsx":
            count = anonymise_workbook(source_file, target_file, mapper)
            workbook_files += 1
            total_names += count
        else:
            shutil.copy2(source_file, target_file)

    print(f"Copied {total_files} files ({workbook_files} workbook files anonymised).")
    print(f"Anonymised name-cell updates: {total_names}")
    print(f"Target input folder: {target_input_dir}")


def ensure_target_output_dir(target_output_dir: Path) -> None:
    target_output_dir.mkdir(parents=True, exist_ok=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Create deterministic anonymised season test data.")
    parser.add_argument("--source-year", type=int, default=2025, help="Source season year (default: 2025)")
    parser.add_argument("--target-year", type=int, default=1999, help="Target test season year (default: 1999)")
    parser.add_argument(
        "--data-root",
        type=Path,
        default=None,
        help="Data root containing year folders (default: auto-detect from prefs, then ./data)",
    )
    parser.add_argument(
        "--seed",
        type=str,
        default="wrrl-back-to-the-future",
        help="Deterministic anonymisation seed",
    )
    parser.add_argument(
        "--keep-target",
        action="store_true",
        help="Keep existing target inputs and overwrite files in place",
    )
    return parser.parse_args()


def resolve_data_root(explicit: Path | None) -> Path:
    if explicit is not None:
        return explicit

    prefs_path = Path.home() / ".wrrl_prefs.json"
    if prefs_path.exists():
        try:
            data = json.loads(prefs_path.read_text(encoding="utf-8"))
            data_root = data.get("data_root")
            if data_root:
                return Path(data_root)
        except Exception:
            pass

    return Path(__file__).resolve().parents[1] / "data"


def main() -> None:
    args = parse_args()
    data_root = resolve_data_root(args.data_root)

    source_input_dir = data_root / str(args.source_year) / "inputs"
    target_root = data_root / str(args.target_year)
    target_input_dir = target_root / "inputs"
    target_output_dir = target_root / "outputs"

    mapper = NameMapper(seed=args.seed)

    copy_and_anonymise_inputs(
        source_input_dir=source_input_dir,
        target_input_dir=target_input_dir,
        mapper=mapper,
        clear_target=not args.keep_target,
    )
    ensure_target_output_dir(target_output_dir)

    print(f"Target output folder ensured: {target_output_dir}")
    print(f"Data root used: {data_root}")
    print("Done.")


if __name__ == "__main__":
    main()
