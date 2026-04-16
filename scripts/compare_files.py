"""Standalone file comparison utility.

Usage:
    python scripts/compare_files.py path/to/file1 path/to/file2
    python scripts/compare_files.py --gui

This script is intentionally standalone and does not integrate with the main application.
"""

from __future__ import annotations

import argparse
import hashlib
import os
import sys
from pathlib import Path
from typing import Iterable


def compute_hash(path: Path, algorithm: str = "sha256") -> str:
    h = hashlib.new(algorithm)
    with path.open("rb") as fh:
        for chunk in iter(lambda: fh.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()


def is_text_file(path: Path, blocksize: int = 4096) -> bool:
    try:
        with path.open("rb") as fh:
            chunk = fh.read(blocksize)
            if not chunk:
                return True
            if b"\x00" in chunk:
                return False
            chunk.decode("utf-8")
            return True
    except (UnicodeDecodeError, OSError):
        return False


def read_lines(path: Path) -> list[str]:
    with path.open("r", encoding="utf-8", errors="surrogateescape") as fh:
        return fh.readlines()


def compare_text_files(path1: Path, path2: Path, context: int) -> int:
    import difflib

    lines1 = read_lines(path1)
    lines2 = read_lines(path2)
    diff = list(difflib.unified_diff(lines1, lines2, fromfile=str(path1), tofile=str(path2), lineterm="", n=context))
    if not diff:
        print("Files are identical textually.")
        return 0

    print("Differences found:\n")
    for line in diff:
        print(line)
    return len(diff)


def compare_binary_files(path1: Path, path2: Path) -> int:
    hash1 = compute_hash(path1)
    hash2 = compute_hash(path2)
    size1 = path1.stat().st_size
    size2 = path2.stat().st_size

    print(f"Binary files differ")
    print(f"    {path1}: size={size1}, sha256={hash1}")
    print(f"    {path2}: size={size2}, sha256={hash2}")
    return 1 if hash1 != hash2 else 0


def print_summary(path1: Path, path2: Path) -> None:
    print(f"Comparing files:")
    print(f"  1. {path1}")
    print(f"  2. {path2}")
    print()
    print(f"Sizes: {path1.stat().st_size} bytes, {path2.stat().st_size} bytes")
    print(f"Text mode: {is_text_file(path1)} / {is_text_file(path2)}")
    print()


def get_excel_sheets(path: Path) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return workbook.sheetnames
    finally:
        workbook.close()


def sheet_to_lines(path: Path, sheet_name: str) -> list[str]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        ws = workbook[sheet_name]
        lines: list[str] = []
        for row in ws.iter_rows(values_only=True):
            line = "\t".join("" if value is None else str(value) for value in row)
            lines.append(line)
        return lines
    finally:
        workbook.close()


def compare_excel_sheets(path1: Path, sheet1: str, path2: Path, sheet2: str, context: int) -> int:
    import difflib

    lines1 = sheet_to_lines(path1, sheet1)
    lines2 = sheet_to_lines(path2, sheet2)
    diff = list(difflib.unified_diff(lines1, lines2, fromfile=f"{path1}:{sheet1}", tofile=f"{path2}:{sheet2}", lineterm="", n=context))
    if not diff:
        print("Sheets are identical.")
        return 0

    print("Differences found:\n")
    for line in diff:
        print(line)
    return len(diff)


def parse_args(argv: Iterable[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Compare two files and show a quick summary, unified diff, or Excel sheet comparison.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument("file1", type=Path, nargs="?", help="Path to the first file")
    parser.add_argument("file2", type=Path, nargs="?", help="Path to the second file")
    parser.add_argument("--sheet1", type=str, help="Sheet name for the first workbook")
    parser.add_argument("--sheet2", type=str, help="Sheet name for the second workbook")
    parser.add_argument("--context", type=int, default=3, help="Number of context lines for unified diff")
    parser.add_argument("--no-summary", action="store_true", help="Skip the comparison summary header")
    parser.add_argument("--gui", action="store_true", help="Launch a GUI for file and sheet selection")
    return parser.parse_args(argv)


class CompareWindowMixin:
    def enable_drop(self, widget):
        widget.setAcceptDrops(True)
        widget.dragEnterEvent = lambda event: event.accept() if event.mimeData().hasUrls() else event.ignore()
        widget.dropEvent = self._make_drop_handler(widget)

    def _make_drop_handler(self, widget):
        def handler(event):
            urls = [url.toLocalFile() for url in event.mimeData().urls() if url.isLocalFile()]
            if urls:
                widget.setText(urls[0])
                if hasattr(widget, "textChanged"):
                    widget.textChanged.emit(widget.text())
        return handler


def launch_gui() -> int:
    try:
        from PySide6.QtCore import Qt
        from PySide6.QtWidgets import (
            QApplication,
            QComboBox,
            QGridLayout,
            QHBoxLayout,
            QLabel,
            QLineEdit,
            QMainWindow,
            QPushButton,
            QTextEdit,
            QVBoxLayout,
            QWidget,
            QFileDialog,
            QMessageBox,
        )
    except ImportError:
        print("Error: PySide6 is required for GUI mode.", file=sys.stderr)
        return 1

    class CompareWindow(QMainWindow, CompareWindowMixin):
        def __init__(self):
            super().__init__()
            self.setWindowTitle("Standalone File Comparer")
            self._build_ui()

        def _build_ui(self) -> None:
            central = QWidget(self)
            self.setCentralWidget(central)
            layout = QVBoxLayout(central)

            grid = QGridLayout()
            layout.addLayout(grid)

            self.path1_edit = QLineEdit(self)
            self.path2_edit = QLineEdit(self)
            self.sheet1_combo = QComboBox(self)
            self.sheet2_combo = QComboBox(self)
            self.sheet1_combo.setEnabled(False)
            self.sheet2_combo.setEnabled(False)

            grid.addWidget(QLabel("File 1:"), 0, 0)
            grid.addWidget(self.path1_edit, 0, 1)
            browse1 = QPushButton("Browse…", self)
            grid.addWidget(browse1, 0, 2)

            grid.addWidget(QLabel("Sheet 1:"), 1, 0)
            grid.addWidget(self.sheet1_combo, 1, 1, 1, 2)

            grid.addWidget(QLabel("File 2:"), 2, 0)
            grid.addWidget(self.path2_edit, 2, 1)
            browse2 = QPushButton("Browse…", self)
            grid.addWidget(browse2, 2, 2)

            grid.addWidget(QLabel("Sheet 2:"), 3, 0)
            grid.addWidget(self.sheet2_combo, 3, 1, 1, 2)

            browse1.clicked.connect(self._choose_file1)
            browse2.clicked.connect(self._choose_file2)
            self.path1_edit.textChanged.connect(self._refresh_sheets)
            self.path2_edit.textChanged.connect(self._refresh_sheets)
            self.enable_drop(self.path1_edit)
            self.enable_drop(self.path2_edit)

            button_layout = QHBoxLayout()
            layout.addLayout(button_layout)
            self.compare_button = QPushButton("Compare", self)
            self.clear_button = QPushButton("Clear", self)
            button_layout.addWidget(self.compare_button)
            button_layout.addWidget(self.clear_button)
            self.compare_button.clicked.connect(self._on_compare)
            self.clear_button.clicked.connect(self._on_clear)

            self.result_view = QTextEdit(self)
            self.result_view.setReadOnly(True)
            layout.addWidget(self.result_view)

            self.resize(900, 600)

        def _choose_file1(self) -> None:
            path, _ = QFileDialog.getOpenFileName(self, "Choose first file")
            if path:
                self.path1_edit.setText(path)

        def _choose_file2(self) -> None:
            path, _ = QFileDialog.getOpenFileName(self, "Choose second file")
            if path:
                self.path2_edit.setText(path)

        def _refresh_sheets(self) -> None:
            self._load_sheets(self.path1_edit.text(), self.sheet1_combo)
            self._load_sheets(self.path2_edit.text(), self.sheet2_combo)

        def _load_sheets(self, text: str, combo: QComboBox) -> None:
            combo.clear()
            combo.setEnabled(False)
            if not text:
                return
            path = Path(text)
            if not path.exists() or not path.is_file():
                return
            if path.suffix.lower() not in {".xls", ".xlsx"}:
                return
            try:
                sheets = get_excel_sheets(path)
            except Exception as exc:
                self.result_view.append(f"Failed to read sheets from {path}: {exc}")
                return
            combo.addItems(sheets)
            combo.setEnabled(True)

        def _on_compare(self) -> None:
            file1 = Path(self.path1_edit.text())
            file2 = Path(self.path2_edit.text())
            if not file1.exists() or not file2.exists():
                QMessageBox.warning(self, "Missing files", "Please select both files before comparing.")
                return
            if file1.suffix.lower() not in {".xls", ".xlsx"} or file2.suffix.lower() not in {".xls", ".xlsx"}:
                QMessageBox.warning(self, "Invalid files", "Please select Excel files for sheet comparison.")
                return
            if not self.sheet1_combo.currentText() or not self.sheet2_combo.currentText():
                QMessageBox.warning(self, "Missing sheets", "Please select a sheet for both files.")
                return
            self.compare_button.setEnabled(False)
            self.result_view.clear()
            try:
                diff = compare_excel_sheets(file1, self.sheet1_combo.currentText(), file2, self.sheet2_combo.currentText(), context=3)
                if diff == 0:
                    self.result_view.append("Sheets are identical.")
                else:
                    self.result_view.append(f"{diff} diff lines shown.")
            except Exception as exc:
                self.result_view.append(f"Comparison failed: {exc}")
            finally:
                self.compare_button.setEnabled(True)

        def _on_clear(self) -> None:
            self.path1_edit.clear()
            self.path2_edit.clear()
            self.sheet1_combo.clear()
            self.sheet2_combo.clear()
            self.result_view.clear()

    app = QApplication(sys.argv)
    window = CompareWindow()
    window.show()
    return app.exec()


def main(argv: Iterable[str] | None = None) -> int:
    args = parse_args(argv)

    if args.gui or (args.file1 is None and args.file2 is None):
        return launch_gui()

    if args.file1 is None or args.file2 is None:
        print("Error: two files must be provided unless --gui is used.", file=sys.stderr)
        return 1

    path1 = args.file1
    path2 = args.file2

    for path in (path1, path2):
        if not path.exists():
            print(f"Error: file not found: {path}", file=sys.stderr)
            return 1
        if not path.is_file():
            print(f"Error: not a file: {path}", file=sys.stderr)
            return 1

    if not args.no_summary:
        print_summary(path1, path2)

    if args.sheet1 or args.sheet2:
        if not args.sheet1 or not args.sheet2:
            print("Error: both --sheet1 and --sheet2 must be specified for Excel sheet comparison.", file=sys.stderr)
            return 1
        try:
            diff_lines = compare_excel_sheets(path1, args.sheet1, path2, args.sheet2, args.context)
            return 0 if diff_lines == 0 else 2
        except Exception as exc:
            print(f"Excel comparison failed: {exc}", file=sys.stderr)
            return 1

    if is_text_file(path1) and is_text_file(path2):
        diff_lines = compare_text_files(path1, path2, args.context)
        return 0 if diff_lines == 0 else 2

    return compare_binary_files(path1, path2)


if __name__ == "__main__":
    raise SystemExit(main())
