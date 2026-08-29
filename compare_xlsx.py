import openpyxl
from pathlib import Path

file1 = Path("documents/Season Standings R05 2026.xlsx")
file2 = Path("documents/Season Standings R05 2026 new.xlsx")

print(f"Comparing:\n  {file1.name}\n  {file2.name}\n")

wb1 = openpyxl.load_workbook(file1, data_only=True)
wb2 = openpyxl.load_workbook(file2, data_only=True)

# Compare sheet names
sheets1 = set(wb1.sheetnames)
sheets2 = set(wb2.sheetnames)

if sheets1 != sheets2:
    print(f"⚠️  Sheet names differ:")
    print(f"  Only in original: {sheets1 - sheets2}")
    print(f"  Only in new: {sheets2 - sheets1}\n")
else:
    print(f"✅ Sheet names match: {sheets1}\n")

# Compare sheet contents
differences_found = False

for sheet_name in sorted(sheets1 & sheets2):
    ws1 = wb1[sheet_name]
    ws2 = wb2[sheet_name]
    
    max_row1 = ws1.max_row
    max_col1 = ws1.max_column
    max_row2 = ws2.max_row
    max_col2 = ws2.max_column
    
    if max_row1 != max_row2 or max_col1 != max_col2:
        print(f"📊 Sheet '{sheet_name}': Dimensions differ")
        print(f"  Original: {max_row1} rows × {max_col1} cols")
        print(f"  New:      {max_row2} rows × {max_col2} cols\n")
        differences_found = True
        continue
    
    # Compare cell values
    cell_diffs = []
    for row in range(1, max_row1 + 1):
        for col in range(1, max_col1 + 1):
            val1 = ws1.cell(row, col).value
            val2 = ws2.cell(row, col).value
            if val1 != val2:
                cell_diffs.append((row, col, val1, val2))
    
    if cell_diffs:
        print(f"📊 Sheet '{sheet_name}': {len(cell_diffs)} cell(s) differ")
        for row, col, val1, val2 in cell_diffs[:10]:  # Show first 10
            print(f"  Cell ({row},{col}): '{val1}' → '{val2}'")
        if len(cell_diffs) > 10:
            print(f"  ... and {len(cell_diffs) - 10} more differences")
        print()
        differences_found = True
    else:
        print(f"✅ Sheet '{sheet_name}': No differences")

if not differences_found:
    print("\n✅ Files are identical!")
