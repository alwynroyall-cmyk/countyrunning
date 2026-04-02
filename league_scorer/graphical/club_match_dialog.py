from __future__ import annotations

import tkinter as tk
from dataclasses import dataclass
from pathlib import Path
from tkinter import messagebox
from typing import Dict, List

import pandas as pd
from openpyxl import load_workbook


_HIGHLIGHT_BG = "#fff1c7"
_DEFAULT_BG = "#f5f5f5"


@dataclass
class ClubMatchCandidate:
    current_club: str
    proposed_club: str
    message: str
    confidence: str
    occurrences: str
    races_seen: str


class ClubMatchApprovalDialog(tk.Toplevel):
    def __init__(self, parent: tk.Misc, summary_df: pd.DataFrame, clubs_path: Path) -> None:
        super().__init__(parent)
        self.title("Club Match Approval")
        self.geometry("980x560")
        self.minsize(860, 420)
        self.configure(bg=_DEFAULT_BG)
        self.transient(parent.winfo_toplevel())

        self._clubs_path = clubs_path
        self._candidates = _build_candidates(summary_df)
        self._candidate_vars: Dict[str, tk.BooleanVar] = {}
        self._row_frames: Dict[str, tk.Frame] = {}
        self._conflict_highlights: set[str] = set()

        self._build_ui()
        self.grab_set()

    def _build_ui(self) -> None:
        title = tk.Label(
            self,
            text="Club Match Approval",
            font=("Segoe UI", 15, "bold"),
            bg=_DEFAULT_BG,
            fg="#22313f",
        )
        title.pack(anchor="w", padx=16, pady=(14, 4))

        subtitle = tk.Label(
            self,
            text="Tick approved club conversions, then confirm before writing them into Club Lookup.",
            font=("Segoe UI", 10),
            bg=_DEFAULT_BG,
            fg="#4a5868",
        )
        subtitle.pack(anchor="w", padx=16, pady=(0, 10))

        header = tk.Frame(self, bg="#dbe5ee")
        header.pack(fill="x", padx=16)
        self._make_header_label(header, "Tick", 0, 10)
        self._make_header_label(header, "Current Club", 1, 24)
        self._make_header_label(header, "Proposed Club", 2, 24)
        self._make_header_label(header, "Message", 3, 48)

        outer = tk.Frame(self, bg=_DEFAULT_BG)
        outer.pack(fill="both", expand=True, padx=16, pady=(0, 12))

        canvas = tk.Canvas(outer, bg=_DEFAULT_BG, highlightthickness=0)
        scrollbar = tk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        self._rows_frame = tk.Frame(canvas, bg=_DEFAULT_BG)

        self._rows_frame.bind(
            "<Configure>",
            lambda _event: canvas.configure(scrollregion=canvas.bbox("all")),
        )

        canvas.create_window((0, 0), window=self._rows_frame, anchor="nw")
        canvas.configure(yscrollcommand=scrollbar.set)

        canvas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        for candidate in self._candidates:
            self._add_candidate_row(candidate)

        action_bar = tk.Frame(self, bg=_DEFAULT_BG)
        action_bar.pack(fill="x", padx=16, pady=(0, 16))

        tk.Button(
            action_bar,
            text="Close",
            command=self.destroy,
            font=("Segoe UI", 10),
            bg="#d0d4de",
            fg="#22313f",
            relief="flat",
            padx=12,
            pady=6,
        ).pack(side="left")

        tk.Button(
            action_bar,
            text="Add conversions to Club Lookup",
            command=self._open_confirmation,
            font=("Segoe UI", 10, "bold"),
            bg="#2d7a4a",
            fg="#ffffff",
            relief="flat",
            padx=14,
            pady=7,
        ).pack(side="right")

    def _make_header_label(self, parent: tk.Misc, text: str, column: int, width: int) -> None:
        lbl = tk.Label(
            parent,
            text=text,
            font=("Segoe UI", 10, "bold"),
            bg="#dbe5ee",
            fg="#22313f",
            anchor="w",
            width=width,
            padx=6,
            pady=6,
        )
        lbl.grid(row=0, column=column, sticky="ew")

    def _add_candidate_row(self, candidate: ClubMatchCandidate) -> None:
        var = tk.BooleanVar(value=False)
        self._candidate_vars[candidate.current_club] = var

        frame = tk.Frame(self._rows_frame, bg=_DEFAULT_BG, bd=1, relief="flat")
        frame.pack(fill="x", pady=1)
        self._row_frames[candidate.current_club] = frame

        checkbox = tk.Checkbutton(
            frame,
            variable=var,
            bg=_DEFAULT_BG,
            activebackground=_DEFAULT_BG,
            highlightthickness=0,
        )
        checkbox.grid(row=0, column=0, sticky="w", padx=(6, 0), pady=6)

        self._make_value_label(frame, candidate.current_club, 1, 24)
        self._make_value_label(frame, candidate.proposed_club, 2, 24)
        self._make_value_label(frame, candidate.message, 3, 72)

        frame.grid_columnconfigure(3, weight=1)

    def _make_value_label(self, parent: tk.Misc, text: str, column: int, width: int) -> None:
        lbl = tk.Label(
            parent,
            text=text,
            font=("Segoe UI", 10),
            bg=_DEFAULT_BG,
            fg="#22313f",
            anchor="w",
            justify="left",
            width=width,
            wraplength=520 if column == 3 else 0,
            padx=6,
            pady=6,
        )
        lbl.grid(row=0, column=column, sticky="ew")

    def _open_confirmation(self) -> None:
        selected = [candidate for candidate in self._candidates if self._candidate_vars[candidate.current_club].get()]
        if not selected:
            messagebox.showwarning(
                "No conversions selected",
                "Tick at least one club conversion before continuing.",
                parent=self,
            )
            return

        lookup_state = _read_club_lookup_state(self._clubs_path)
        summary = _summarise_selection(selected, lookup_state)
        ClubMatchConfirmationDialog(self, summary, self._clubs_path)

    def highlight_conflicts(self, conflict_clubs: List[str]) -> None:
        self._conflict_highlights = set(conflict_clubs)
        for current_club, frame in self._row_frames.items():
            bg = _HIGHLIGHT_BG if current_club in self._conflict_highlights else _DEFAULT_BG
            _set_widget_tree_bg(frame, bg)


class ClubMatchConfirmationDialog(tk.Toplevel):
    def __init__(self, approval_dialog: ClubMatchApprovalDialog, summary: dict, clubs_path: Path) -> None:
        super().__init__(approval_dialog)
        self.title("Confirm Club Lookup Update")
        self.geometry("840x460")
        self.configure(bg=_DEFAULT_BG)
        self.transient(approval_dialog)
        self._approval_dialog = approval_dialog
        self._summary = summary
        self._clubs_path = clubs_path

        self._build_ui()
        self.grab_set()

    def _build_ui(self) -> None:
        tk.Label(
            self,
            text="Confirm Club Lookup Update",
            font=("Segoe UI", 15, "bold"),
            bg=_DEFAULT_BG,
            fg="#22313f",
        ).pack(anchor="w", padx=16, pady=(14, 8))

        text = tk.Text(self, wrap="word", font=("Segoe UI", 10), bg="#ffffff", fg="#22313f")
        text.pack(fill="both", expand=True, padx=16, pady=(0, 12))
        text.insert("1.0", _build_confirmation_text(self._summary))
        text.config(state="disabled")

        action_bar = tk.Frame(self, bg=_DEFAULT_BG)
        action_bar.pack(fill="x", padx=16, pady=(0, 16))

        tk.Button(
            action_bar,
            text="Back To Selection",
            command=self._back_to_selection,
            font=("Segoe UI", 10),
            bg="#d0d4de",
            fg="#22313f",
            relief="flat",
            padx=12,
            pady=6,
        ).pack(side="left")

        confirm_state = "disabled" if self._summary["conflicts"] else "normal"
        tk.Button(
            action_bar,
            text="Confirm Write",
            command=self._confirm_write,
            state=confirm_state,
            font=("Segoe UI", 10, "bold"),
            bg="#2d7a4a",
            fg="#ffffff",
            relief="flat",
            padx=14,
            pady=7,
        ).pack(side="right")

    def _back_to_selection(self) -> None:
        self._approval_dialog.highlight_conflicts([item.current_club for item in self._summary["conflicts"]])
        self.destroy()

    def _confirm_write(self) -> None:
        result = _append_club_conversions(self._clubs_path, self._summary["new_rows"])
        self.destroy()
        self._approval_dialog.destroy()
        messagebox.showinfo(
            "Club Lookup Updated",
            _build_write_result_text(result),
            parent=self.master,
        )


def _build_candidates(summary_df: pd.DataFrame) -> List[ClubMatchCandidate]:
    candidates: List[ClubMatchCandidate] = []
    if summary_df.empty:
        return candidates

    df = summary_df.fillna("")
    for _, row in df.iterrows():
        current_club = str(row.get("Raw Club", "")).strip()
        proposed_club = str(
            row.get("Best Match", row.get("Suggested Club", row.get("Preferred Club", "")))
        ).strip()
        if not current_club or not proposed_club:
            continue
        confidence = str(row.get("Confidence", row.get("Suggested Club Confidence", ""))).strip()
        occurrences = str(row.get("Occurrences", "")).strip()
        races_seen = str(row.get("Races Seen", "")).strip()
        message = str(row.get("Message", "")).strip()
        if not message:
            parts = []
            if confidence:
                parts.append(f"Confidence {confidence}%")
            if occurrences:
                parts.append(f"Occurrences {occurrences}")
            if races_seen:
                parts.append(f"Races {races_seen}")
            message = " | ".join(parts)
        candidates.append(
            ClubMatchCandidate(
                current_club=current_club,
                proposed_club=proposed_club,
                message=message,
                confidence=confidence,
                occurrences=occurrences,
                races_seen=races_seen,
            )
        )
    return candidates


def _read_club_lookup_state(clubs_path: Path) -> dict:
    workbook = load_workbook(clubs_path)
    try:
        worksheet = workbook.active
        headers = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[1]]
        header_map = {name: idx for idx, name in enumerate(headers)}

        required_headers = ["Club", "Preferred name", "Team A", "Team B"]
        missing = [h for h in required_headers if h not in header_map]
        if missing:
            raise ValueError(
                "clubs.xlsx is missing required column(s): "
                + ", ".join(missing)
                + ". Found columns: "
                + ", ".join(headers)
            )

        alias_to_preferred: Dict[str, str] = {}
        preferred_to_divisions: Dict[str, tuple[str, str]] = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            raw = str(row[header_map["Club"]]).strip() if row[header_map["Club"]] is not None else ""
            preferred = str(row[header_map["Preferred name"]]).strip() if row[header_map["Preferred name"]] is not None else ""
            if raw:
                alias_to_preferred[raw.lower()] = preferred
            if preferred and preferred not in preferred_to_divisions:
                team_a = str(row[header_map["Team A"]]).strip() if row[header_map["Team A"]] is not None else ""
                team_b = str(row[header_map["Team B"]]).strip() if row[header_map["Team B"]] is not None else ""
                preferred_to_divisions[preferred] = (team_a, team_b)
    finally:
        workbook.close()

    return {
        "alias_to_preferred": alias_to_preferred,
        "preferred_to_divisions": preferred_to_divisions,
    }


def _summarise_selection(selected: List[ClubMatchCandidate], lookup_state: dict) -> dict:
    exact_existing: List[ClubMatchCandidate] = []
    conflicts: List[ClubMatchCandidate] = []
    new_rows: List[ClubMatchCandidate] = []

    alias_to_preferred = lookup_state["alias_to_preferred"]
    preferred_to_divisions = lookup_state["preferred_to_divisions"]

    for item in selected:
        existing = alias_to_preferred.get(item.current_club.lower())
        if existing:
            if existing == item.proposed_club:
                exact_existing.append(item)
            else:
                conflicts.append(item)
            continue
        if item.proposed_club not in preferred_to_divisions:
            conflicts.append(item)
            continue
        new_rows.append(item)

    return {
        "selected": selected,
        "exact_existing": exact_existing,
        "conflicts": conflicts,
        "new_rows": new_rows,
    }


def _build_confirmation_text(summary: dict) -> str:
    lines = []
    lines.append(f"Selected conversions: {len(summary['selected'])}")
    lines.append(f"New conversions ready to write: {len(summary['new_rows'])}")
    lines.append(f"Exact existing matches: {len(summary['exact_existing'])}")
    lines.append(f"Conflicting existing matches: {len(summary['conflicts'])}")
    lines.append("")

    if summary["new_rows"]:
        lines.append("Ready to write:")
        for item in summary["new_rows"]:
            lines.append(f"- {item.current_club} -> {item.proposed_club}")
        lines.append("")

    if summary["exact_existing"]:
        lines.append("Exact existing matches:")
        for item in summary["exact_existing"]:
            lines.append(f"- {item.current_club} already maps to {item.proposed_club}")
        lines.append("")

    if summary["conflicts"]:
        lines.append("Conflicting existing matches:")
        for item in summary["conflicts"]:
            lines.append(f"- {item.current_club} conflicts with proposed {item.proposed_club}")
        lines.append("")
        lines.append("Return to selection and untick conflicting rows before writing.")

    if not summary["conflicts"]:
        lines.append("Confirm Write will append only new conversions. Exact existing matches will be skipped.")
    return "\n".join(lines)


def _append_club_conversions(clubs_path: Path, selected: List[ClubMatchCandidate]) -> dict:
    workbook = load_workbook(clubs_path)
    worksheet = workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[1]]
    header_map = {name: idx + 1 for idx, name in enumerate(headers)}

    lookup_state = _read_club_lookup_state(clubs_path)
    alias_to_preferred = lookup_state["alias_to_preferred"]
    preferred_to_divisions = lookup_state["preferred_to_divisions"]

    written = 0
    skipped_existing = 0
    skipped_conflicts = 0

    for item in selected:
        existing = alias_to_preferred.get(item.current_club.lower())
        if existing:
            if existing == item.proposed_club:
                skipped_existing += 1
            else:
                skipped_conflicts += 1
            continue

        divisions = preferred_to_divisions.get(item.proposed_club)
        if divisions is None:
            skipped_conflicts += 1
            continue

        next_row = worksheet.max_row + 1
        worksheet.cell(next_row, header_map["Club"], item.current_club)
        worksheet.cell(next_row, header_map["Preferred name"], item.proposed_club)
        worksheet.cell(next_row, header_map["Team A"], divisions[0])
        worksheet.cell(next_row, header_map["Team B"], divisions[1])
        written += 1
        alias_to_preferred[item.current_club.lower()] = item.proposed_club

    workbook.save(clubs_path)
    return {
        "written": written,
        "skipped_existing": skipped_existing,
        "skipped_conflicts": skipped_conflicts,
        "path": clubs_path,
    }


def _build_write_result_text(result: dict) -> str:
    lines = [f"Club Lookup updated: {result['path']}", ""]
    lines.append(f"Conversions written: {result['written']}")
    lines.append(f"Exact existing matches skipped: {result['skipped_existing']}")
    lines.append(f"Conflicting selections skipped: {result['skipped_conflicts']}")
    lines.append("")
    lines.append("Re-run audit to refresh club findings against the updated lookup.")
    return "\n".join(lines)


def _set_widget_tree_bg(widget: tk.Misc, bg: str) -> None:
    try:
        widget.configure(bg=bg)
    except tk.TclError:
        pass
    for child in widget.winfo_children():
        _set_widget_tree_bg(child, bg)