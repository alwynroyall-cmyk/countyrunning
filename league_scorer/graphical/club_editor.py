"""
club_editor.py — Embedded panel for editing runner–club assignments in race files.

Allows the user to pick a race workbook, browse its runners, assign or correct
a club from the eligible clubs list, and save the changes back to the xlsx file.
"""

import tkinter as tk
from pathlib import Path
from tkinter import messagebox, ttk

import openpyxl
import pandas as pd

from ..common_files import race_discovery_exclusions
from ..manual_data_audit import log_manual_data_changes
from ..session_config import config as session_config
from .dashboard import WRRL_GREEN, WRRL_LIGHT, WRRL_NAVY, WRRL_WHITE


class ClubEditorPanel(tk.Frame):
    """Embedded panel for editing runner–club assignments in race workbooks."""

    def __init__(self, parent: tk.Misc, back_callback=None) -> None:
        super().__init__(parent, bg=WRRL_LIGHT)
        self._back_callback = back_callback

        # Workbook state
        self._wb = None
        self._ws = None
        self._name_col: int | tuple[int, int] | None = None
        self._club_col: int | None = None
        self._rows: list[dict] = []        # [{row_idx, name, club}] row_idx is 1-based
        self._pending_changes: dict[int, str] = {}  # row_idx -> new club value
        self._current_file: Path | None = None

        # Listbox view (after filter applied)
        self._listbox_rows: list[dict] = []

        # Eligible clubs (preferred names from clubs.xlsx)
        self._eligible_clubs: list[str] = []
        self._load_eligible_clubs()

        self._build_ui()

    # ── eligible clubs ────────────────────────────────────────────────────────

    def _load_eligible_clubs(self) -> None:
        clubs_path = (
            session_config.control_dir / "clubs.xlsx"
            if session_config.control_dir
            else None
        )
        if not clubs_path or not clubs_path.exists():
            return
        try:
            df = pd.read_excel(clubs_path, dtype=str)
        except Exception as exc:
            messagebox.showwarning(
                "Clubs File Warning",
                f"Could not load eligible clubs from {clubs_path.name}.\n"
                f"The club list will be empty until this is fixed.\n\n"
                f"Details: {exc}",
                parent=self,
            )
            self._eligible_clubs = []
            return
        col = next(
            (c for c in df.columns if "preferred" in c.lower()),
            next((c for c in df.columns if "club" in c.lower()), None),
        )
        if col:
            self._eligible_clubs = sorted(
                {v.strip() for v in df[col].dropna() if v.strip()}
            )

    # ── race file loading ─────────────────────────────────────────────────────

    def _close_loaded_workbook(self) -> None:
        if self._wb is None:
            return
        try:
            self._wb.close()
        except Exception:
            pass
        finally:
            self._wb = None
            self._ws = None

    def _load_race_file(self, path: Path) -> None:
        self._close_loaded_workbook()
        try:
            wb = openpyxl.load_workbook(path)
        except Exception as exc:
            messagebox.showerror("Load Error", f"Could not open workbook:\n{exc}", parent=self)
            return
        ws = wb.active
        headers = [
            str(c.value).strip().lower() if c.value is not None else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]

        # Find name column (full name preferred; fall back to first+last)
        name_idx: int | tuple[int, int] | None = next(
            (
                i for i, h in enumerate(headers)
                if "name" in h and "first" not in h and "last" not in h
            ),
            None,
        )
        if name_idx is None:
            first = next((i for i, h in enumerate(headers) if "first" in h), None)
            last = next((i for i, h in enumerate(headers) if "last" in h), None)
            if first is not None and last is not None:
                name_idx = (first, last)

        if name_idx is None:
            messagebox.showerror("Format Error", "Could not find a Name column.", parent=self)
            wb.close()
            return

        club_idx = next((i for i, h in enumerate(headers) if "club" in h), None)
        if club_idx is None:
            messagebox.showerror("Format Error", "Could not find a Club column.", parent=self)
            wb.close()
            return

        rows: list[dict] = []
        for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            if isinstance(name_idx, tuple):
                fn = str(row[name_idx[0]].value or "").strip()
                ln = str(row[name_idx[1]].value or "").strip()
                name = f"{fn} {ln}".strip()
            else:
                name = str(row[name_idx].value or "").strip()

            raw_club = row[club_idx].value
            club = str(raw_club).strip() if raw_club is not None else ""
            if club.lower() in {"none", "nan"}:
                club = ""

            if name:
                rows.append({"row_idx": row_no, "name": name, "club": club})

        self._wb = wb
        self._ws = ws
        self._name_col = name_idx
        self._club_col = club_idx
        self._rows = rows
        self._pending_changes = {}
        self._current_file = path

        self._refresh_listbox()
        self._refresh_changes_list()
        self._save_btn.config(state="disabled")
        blank_count = sum(1 for r in rows if not r["club"])
        self._status_var.set(
            f"Loaded {len(rows)} runners from {path.name}  "
            f"({blank_count} with blank club)"
        )

    # ── UI construction ───────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        # ── Title bar ─────────────────────────────────────────────────────────
        title_row = tk.Frame(self, bg=WRRL_NAVY, padx=16, pady=10)
        title_row.pack(fill="x")
        tk.Label(
            title_row,
            text="Edit Club Assignments",
            font=("Segoe UI", 16, "bold"),
            bg=WRRL_NAVY,
            fg=WRRL_WHITE,
        ).pack(side="left")

        if self._back_callback:
            tk.Button(
                title_row,
                text="\u25c4 Dashboard",
                font=("Segoe UI", 10, "bold"),
                bg=WRRL_GREEN,
                fg=WRRL_WHITE,
                relief="flat",
                padx=10,
                pady=4,
                cursor="hand2",
                command=self._on_back,
                activebackground="#1f5632",
                activeforeground=WRRL_WHITE,
            ).pack(side="right")

        # ── File selector row ─────────────────────────────────────────────────
        selector_row = tk.Frame(self, bg=WRRL_LIGHT, padx=16, pady=8)
        selector_row.pack(fill="x")
        tk.Label(
            selector_row, text="Race file:",
            font=("Segoe UI", 10, "bold"),
            bg=WRRL_LIGHT, fg=WRRL_NAVY,
        ).pack(side="left", padx=(0, 8))

        self._file_var = tk.StringVar()
        self._file_combo = ttk.Combobox(
            selector_row, textvariable=self._file_var,
            state="readonly", width=55,
        )
        self._file_combo.pack(side="left")
        self._file_combo.bind("<<ComboboxSelected>>", self._on_file_selected)
        self._file_paths: list[Path] = []
        self._populate_file_list()

        self._blank_only_var = tk.BooleanVar(value=False)
        tk.Checkbutton(
            selector_row,
            text="Show blank-club only",
            variable=self._blank_only_var,
            command=self._refresh_listbox,
            bg=WRRL_LIGHT, fg=WRRL_NAVY,
            font=("Segoe UI", 10),
            activebackground=WRRL_LIGHT,
        ).pack(side="left", padx=16)

        self._changed_only_var = tk.BooleanVar(value=False)
        tk.Checkbutton(
            selector_row,
            text="Show changed only",
            variable=self._changed_only_var,
            command=self._refresh_listbox,
            bg=WRRL_LIGHT, fg=WRRL_NAVY,
            font=("Segoe UI", 10),
            activebackground=WRRL_LIGHT,
        ).pack(side="left", padx=(0, 16))

        # ── Search bar ────────────────────────────────────────────────────────
        search_row = tk.Frame(self, bg=WRRL_LIGHT, padx=16, pady=4)
        search_row.pack(fill="x")
        tk.Label(
            search_row, text="Search:",
            font=("Segoe UI", 10, "bold"),
            bg=WRRL_LIGHT, fg=WRRL_NAVY,
        ).pack(side="left", padx=(0, 6))
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._refresh_listbox())
        search_entry = tk.Entry(
            search_row, textvariable=self._search_var,
            font=("Segoe UI", 10), width=40, relief="sunken", bd=1,
        )
        search_entry.pack(side="left")
        tk.Button(
            search_row, text="✕",
            font=("Segoe UI", 9), relief="flat", bd=0,
            bg=WRRL_LIGHT, fg="#888888", cursor="hand2",
            command=lambda: self._search_var.set(""),
        ).pack(side="left", padx=(4, 0))
        self._list_count_var = tk.StringVar(value="")
        tk.Label(
            search_row, textvariable=self._list_count_var,
            font=("Segoe UI", 9, "italic"),
            bg=WRRL_LIGHT, fg="#666666",
        ).pack(side="left", padx=12)

        # Bottom bar must be packed before the expanding body (tkinter pack order)
        self._build_bottom_bar()

        # ── Main body (list + editor) ─────────────────────────────────────────
        body = tk.Frame(self, bg=WRRL_LIGHT)
        body.pack(fill="both", expand=True, padx=16, pady=(0, 8))

        # Left — runner listbox
        left = tk.Frame(body, bg=WRRL_LIGHT)
        left.pack(side="left", fill="both", expand=True, padx=(0, 12))

        tk.Label(
            left, text="Runners",
            font=("Segoe UI", 11, "bold"),
            bg=WRRL_LIGHT, fg=WRRL_NAVY,
        ).pack(anchor="w", pady=(4, 2))

        list_frame = tk.Frame(left, bg=WRRL_LIGHT)
        list_frame.pack(fill="both", expand=True)

        scrollbar = tk.Scrollbar(list_frame, orient="vertical")
        self._listbox = tk.Listbox(
            list_frame,
            yscrollcommand=scrollbar.set,
            font=("Consolas", 10),
            selectbackground=WRRL_GREEN,
            selectforeground=WRRL_WHITE,
            activestyle="none",
            exportselection=False,
            bd=1,
            relief="sunken",
        )
        scrollbar.config(command=self._listbox.yview)
        scrollbar.pack(side="right", fill="y")
        self._listbox.pack(side="left", fill="both", expand=True)
        self._listbox.bind("<<ListboxSelect>>", self._on_runner_selected)

        # Right — editor panel
        right = tk.Frame(body, bg=WRRL_LIGHT, width=300)
        right.pack(side="right", fill="y")
        right.pack_propagate(False)

        tk.Label(
            right, text="Selected Runner",
            font=("Segoe UI", 11, "bold"),
            bg=WRRL_LIGHT, fg=WRRL_NAVY,
        ).pack(anchor="w", pady=(4, 8))

        self._runner_name_var = tk.StringVar(value="—")
        tk.Label(
            right, textvariable=self._runner_name_var,
            font=("Segoe UI", 12, "bold"),
            bg=WRRL_LIGHT, fg=WRRL_NAVY,
            wraplength=280, justify="left",
        ).pack(anchor="w", pady=(0, 4))

        tk.Label(right, text="Current club:", font=("Segoe UI", 9),
                 bg=WRRL_LIGHT, fg="#666666").pack(anchor="w")
        self._current_club_var = tk.StringVar(value="—")
        tk.Label(
            right, textvariable=self._current_club_var,
            font=("Segoe UI", 10, "italic"),
            bg=WRRL_LIGHT, fg="#555555",
            wraplength=280, justify="left",
        ).pack(anchor="w", pady=(0, 16))

        tk.Label(right, text="Assign club:", font=("Segoe UI", 10, "bold"),
                 bg=WRRL_LIGHT, fg=WRRL_NAVY).pack(anchor="w")
        self._club_var = tk.StringVar()
        self._club_combo = ttk.Combobox(
            right, textvariable=self._club_var,
            values=self._eligible_clubs,
            state="readonly", width=35,
        )
        self._club_combo.pack(anchor="w", pady=(4, 8))

        self._apply_btn = tk.Button(
            right, text="Apply Change",
            font=("Segoe UI", 10, "bold"),
            bg=WRRL_GREEN, fg=WRRL_WHITE,
            relief="flat", padx=12, pady=6,
            cursor="hand2",
            state="disabled",
            command=self._on_apply,
            activebackground="#1f5632",
            activeforeground=WRRL_WHITE,
        )
        self._apply_btn.pack(anchor="w", pady=(0, 16))

        # Separator
        tk.Frame(right, bg="#cccccc", height=1).pack(fill="x", pady=(4, 8))

        self._clear_btn = tk.Button(
            right, text="Clear This Change",
            font=("Segoe UI", 9),
            bg="#dbe1e8", fg=WRRL_NAVY,
            relief="flat", padx=8, pady=4,
            cursor="hand2",
            state="disabled",
            command=self._on_clear_change,
        )
        self._clear_btn.pack(anchor="w", pady=(0, 8))

        # ── Pending changes list ───────────────────────────────────────────────
        tk.Frame(right, bg="#cccccc", height=1).pack(fill="x", pady=(0, 6))
        self._pending_var = tk.StringVar(value="No unsaved changes.")
        tk.Label(
            right, textvariable=self._pending_var,
            font=("Segoe UI", 9, "bold"),
            bg=WRRL_LIGHT, fg=WRRL_NAVY,
        ).pack(anchor="w", pady=(0, 4))

        changes_frame = tk.Frame(right, bg=WRRL_LIGHT)
        changes_frame.pack(fill="both", expand=True)
        changes_scroll = tk.Scrollbar(changes_frame, orient="vertical")
        self._changes_listbox = tk.Listbox(
            changes_frame,
            yscrollcommand=changes_scroll.set,
            font=("Consolas", 9),
            selectbackground="#dbe1e8",
            selectforeground=WRRL_NAVY,
            activestyle="none",
            bd=1,
            relief="sunken",
            height=10,
        )
        changes_scroll.config(command=self._changes_listbox.yview)
        changes_scroll.pack(side="right", fill="y")
        self._changes_listbox.pack(side="left", fill="both", expand=True)
        self._changes_listbox.bind("<<ListboxSelect>>", self._on_change_item_selected)

    def _build_bottom_bar(self) -> None:
        """Pack the bottom action bar. Must be called BEFORE _build_body so that
        tkinter's packer allocates its space before the expanding body frame."""
        bottom = tk.Frame(self, bg=WRRL_NAVY, padx=16, pady=8)
        bottom.pack(side="bottom", fill="x")

        self._save_btn = tk.Button(
            bottom, text="💾  Save to File",
            font=("Segoe UI", 11, "bold"),
            bg=WRRL_GREEN, fg=WRRL_WHITE,
            relief="flat", padx=16, pady=6,
            cursor="hand2",
            state="disabled",
            command=self._on_save,
            activebackground="#1f5632",
            activeforeground=WRRL_WHITE,
        )
        self._save_btn.pack(side="left", padx=(0, 12))

        self._status_var = tk.StringVar(value="Select a race file to begin.")
        tk.Label(
            bottom, textvariable=self._status_var,
            font=("Segoe UI", 9, "italic"),
            bg=WRRL_NAVY, fg="#a0b0c0",
        ).pack(side="right")

    # ── file list helpers ─────────────────────────────────────────────────────

    def _populate_file_list(self) -> None:
        skip = set(race_discovery_exclusions())
        raw_data_dir = session_config.raw_data_dir
        if not raw_data_dir or not raw_data_dir.exists():
            return
        files = sorted(
            [f for f in raw_data_dir.glob("*.xlsx") if f.name.lower() not in skip],
            key=lambda f: f.stem,
        )
        self._file_paths = files
        self._file_combo["values"] = [f.name for f in files]

    # ── listbox helpers ───────────────────────────────────────────────────────

    def _refresh_listbox(self) -> None:
        self._listbox.delete(0, "end")
        self._listbox_rows = []
        if not self._rows:
            self._list_count_var.set("")
            return
        blank_only   = self._blank_only_var.get()
        changed_only = self._changed_only_var.get()
        needle = self._search_var.get().strip().lower()
        for row in self._rows:
            effective_club = self._pending_changes.get(row["row_idx"], row["club"])
            if blank_only and effective_club:
                continue
            if changed_only and row["row_idx"] not in self._pending_changes:
                continue
            if needle and needle not in row["name"].lower() and needle not in effective_club.lower():
                continue
            self._listbox_rows.append(row)
            changed = row["row_idx"] in self._pending_changes
            marker = "* " if changed else "  "
            club_str = effective_club if effective_club else "[blank]"
            self._listbox.insert("end", f"{marker}{row['name']:<35}  {club_str}")
            if changed:
                self._listbox.itemconfig("end", fg=WRRL_GREEN)
            elif not row["club"]:
                self._listbox.itemconfig("end", fg="#cc6600")
        total = len(self._listbox_rows)
        self._list_count_var.set(f"{total} runner{'s' if total != 1 else ''} shown")

    def _refresh_changes_list(self) -> None:
        """Rebuild the pending-changes listbox in the right panel."""
        self._changes_listbox.delete(0, "end")
        if not self._pending_changes:
            self._pending_var.set("No unsaved changes.")
            return
        n = len(self._pending_changes)
        self._pending_var.set(f"{n} unsaved change{'s' if n != 1 else ''}:")
        # Build a lookup for names by row_idx
        name_by_row = {r["row_idx"]: r["name"] for r in self._rows}
        for row_idx, new_club in sorted(
            self._pending_changes.items(),
            key=lambda kv: name_by_row.get(kv[0], ""),
        ):
            name = name_by_row.get(row_idx, f"row {row_idx}")
            self._changes_listbox.insert("end", f"{name}  →  {new_club}")
            self._changes_listbox.itemconfig("end", fg=WRRL_GREEN)

    def _on_change_item_selected(self, _event=None) -> None:
        """Clicking a row in the changes list jumps to that runner in the main listbox."""
        sel = self._changes_listbox.curselection()
        if not sel:
            return
        # Find the row_idx for the selected change (sorted by name, same order as list)
        name_by_row = {r["row_idx"]: r["name"] for r in self._rows}
        sorted_changes = sorted(
            self._pending_changes.items(),
            key=lambda kv: name_by_row.get(kv[0], ""),
        )
        row_idx = sorted_changes[sel[0]][0]
        # Jump to that runner in the main listbox
        for i, r in enumerate(self._listbox_rows):
            if r["row_idx"] == row_idx:
                self._listbox.selection_clear(0, "end")
                self._listbox.selection_set(i)
                self._listbox.see(i)
                self._on_runner_selected()
                break
        else:
            # Runner not visible under current filter — clear filters and retry
            self._search_var.set("")
            self._blank_only_var.set(False)
            self._changed_only_var.set(False)
            self._refresh_listbox()
            for i, r in enumerate(self._listbox_rows):
                if r["row_idx"] == row_idx:
                    self._listbox.selection_clear(0, "end")
                    self._listbox.selection_set(i)
                    self._listbox.see(i)
                    self._on_runner_selected()
                    break

    def _on_file_selected(self, _event=None) -> None:
        if self._pending_changes:
            if not messagebox.askyesno(
                "Unsaved Changes",
                "You have unsaved changes. Discard them and load a different file?",
                parent=self,
            ):
                # Reset combobox to the current file
                if self._current_file:
                    names = [f.name for f in self._file_paths]
                    if self._current_file.name in names:
                        self._file_var.set(self._current_file.name)
                return
        idx = self._file_combo.current()
        if idx < 0:
            return
        self._load_race_file(self._file_paths[idx])
        self._runner_name_var.set("—")
        self._current_club_var.set("—")
        self._club_var.set("")
        self._apply_btn.config(state="disabled")
        self._clear_btn.config(state="disabled")

    def _on_runner_selected(self, _event=None) -> None:
        sel = self._listbox.curselection()
        if not sel:
            return
        row = self._listbox_rows[sel[0]]
        self._runner_name_var.set(row["name"])
        effective_club = self._pending_changes.get(row["row_idx"], row["club"])
        self._current_club_var.set(effective_club if effective_club else "[blank]")
        # Pre-fill the club combobox if the runner already has a club
        if effective_club and effective_club in self._eligible_clubs:
            self._club_var.set(effective_club)
        else:
            self._club_var.set("")
        self._apply_btn.config(state="normal")
        self._clear_btn.config(
            state="normal" if row["row_idx"] in self._pending_changes else "disabled"
        )

    def _on_apply(self) -> None:
        sel = self._listbox.curselection()
        if not sel:
            return
        row = self._listbox_rows[sel[0]]
        new_club = self._club_var.get().strip()
        if not new_club:
            messagebox.showwarning(
                "No Club Selected",
                "Please choose a club from the drop-down list.",
                parent=self,
            )
            return
        self._pending_changes[row["row_idx"]] = new_club
        self._current_club_var.set(new_club)
        self._save_btn.config(state="normal")
        self._clear_btn.config(state="normal")
        self._refresh_listbox()
        self._refresh_changes_list()
        # Re-select the same runner
        for i, r in enumerate(self._listbox_rows):
            if r["row_idx"] == row["row_idx"]:
                self._listbox.selection_set(i)
                self._listbox.see(i)
                break

    def _on_clear_change(self) -> None:
        sel = self._listbox.curselection()
        if not sel:
            return
        row = self._listbox_rows[sel[0]]
        self._pending_changes.pop(row["row_idx"], None)
        if not self._pending_changes:
            self._save_btn.config(state="disabled")
        self._clear_btn.config(state="disabled")
        # Restore display
        original = row["club"]
        self._current_club_var.set(original if original else "[blank]")
        if original and original in self._eligible_clubs:
            self._club_var.set(original)
        else:
            self._club_var.set("")
        self._refresh_listbox()
        self._refresh_changes_list()
        # Re-select
        for i, r in enumerate(self._listbox_rows):
            if r["row_idx"] == row["row_idx"]:
                self._listbox.selection_set(i)
                self._listbox.see(i)
                break

    def _on_save(self) -> None:
        if not self._pending_changes or self._ws is None or self._current_file is None:
            return
        if self._club_col is None:
            messagebox.showerror("Error", "Club column not identified.", parent=self)
            return

        row_lookup = {row["row_idx"]: row for row in self._rows}
        audit_changes: list[dict] = []
        for row_idx, new_club in self._pending_changes.items():
            row = row_lookup.get(row_idx, {})
            old_club = str(row.get("club", "") or "")
            if old_club == new_club:
                continue
            audit_changes.append(
                {
                    "runner": row.get("name", ""),
                    "field": "club",
                    "old_value": old_club,
                    "new_value": new_club,
                    "file_path": self._current_file,
                    "row_idx": row_idx,
                }
            )

        excel_col = self._club_col + 1  # openpyxl uses 1-based columns
        for row_idx, new_club in self._pending_changes.items():
            self._ws.cell(row=row_idx, column=excel_col).value = new_club
        try:
            self._wb.save(self._current_file)
        except Exception as exc:
            messagebox.showerror("Save Error", f"Could not save workbook:\n{exc}", parent=self)
            return
        # Commit changes into the in-memory row cache
        for row in self._rows:
            if row["row_idx"] in self._pending_changes:
                row["club"] = self._pending_changes[row["row_idx"]]
        n = len(self._pending_changes)
        self._pending_changes = {}
        self._save_btn.config(state="disabled")
        self._clear_btn.config(state="disabled")
        self._refresh_listbox()
        self._refresh_changes_list()

        log_error = log_manual_data_changes(
            audit_changes,
            source="Club Editor",
            action="Update club",
        )
        if log_error:
            messagebox.showwarning(
                "Audit Logging Warning",
                "Changes were saved, but manual audit logging failed:\n"
                f"{log_error}",
                parent=self,
            )

        self._status_var.set(
            f"Saved {n} change{'s' if n != 1 else ''} to {self._current_file.name}"
        )

    def _on_back(self) -> None:
        if self._pending_changes:
            if not messagebox.askyesno(
                "Unsaved Changes",
                "You have unsaved changes. Discard them and go back?",
                parent=self,
            ):
                return
        self._close_loaded_workbook()
        if self._back_callback:
            self._back_callback()
