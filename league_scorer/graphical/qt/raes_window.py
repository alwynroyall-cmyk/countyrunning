from __future__ import annotations

import datetime
import json
import os
import queue
import re
import threading
import urllib.parse
import webbrowser
from pathlib import Path
from typing import Dict, List, Optional

import openpyxl
import pandas as pd
from PySide6.QtCore import QTimer, Qt, Slot
from PySide6.QtGui import QFont, QGuiApplication
from PySide6.QtWidgets import (
    QCheckBox,
    QComboBox,
    QGroupBox,
    QHBoxLayout,
    QLabel,
    QLineEdit,
    QMainWindow,
    QMessageBox,
    QPushButton,
    QScrollArea,
    QSizePolicy,
    QSplitter,
    QStyle,
    QTreeWidget,
    QTreeWidgetItem,
    QVBoxLayout,
    QWidget,
    QTextEdit,
)

from ...manual_edit_service import _atomic_save, _find_columns, _row_name_value
from ...raes.raes_write_service import apply_field_to_files, find_candidate_source_files
from ...audit_data_service import find_latest_audit_workbook
from ...manual_data_audit import log_manual_data_changes
from ...name_lookup import append_name_corrections
from ..results_workbook import find_latest_results_workbook, sorted_race_sheet_names
from ...session_config import config as session_config
from ...structured_logging import log_event

RAES_LIGHT = "#f5f5f5"
RAES_NAVY = "#3a4658"
RAES_GREEN = "#2d7a4a"


def proper_case(s: str) -> str:
    import re

    return re.sub(
        r"([A-Za-z])([A-Za-z']*)",
        lambda m: m.group(1).upper() + m.group(2).lower(),
        s,
    )


def _processed_state_path() -> Path | None:
    out = session_config.output_dir
    if out is None:
        return None
    d = Path(out) / "raes"
    d.mkdir(parents=True, exist_ok=True)
    return d / "processed_state.json"


def load_processed_state() -> Dict[str, bool]:
    p = _processed_state_path()
    if p is None or not p.exists():
        return {}
    try:
        with open(p, "r", encoding="utf-8") as fh:
            data = json.load(fh)
            if isinstance(data, dict):
                return {str(k): bool(v) for k, v in data.items()}
    except Exception:
        return {}
    return {}


def save_processed_state(state: Dict[str, bool]) -> None:
    p = _processed_state_path()
    if p is None:
        return
    try:
        with open(p, "w", encoding="utf-8") as fh:
            json.dump(state, fh, ensure_ascii=False, indent=2)
    except Exception:
        pass


def set_runner_processed(runner: str, processed: bool) -> None:
    st = load_processed_state()
    st[runner] = bool(processed)
    save_processed_state(st)


def _scan_workbook_for_runner_state() -> Dict[str, Dict[str, set]]:
    if session_config.output_dir is None:
        return {}

    workbook = find_latest_results_workbook(session_config.output_dir)
    if workbook is None:
        return {}

    runner_state: Dict[str, Dict[str, set]] = {}
    try:
        xl = pd.ExcelFile(workbook)
        for sheet in sorted_race_sheet_names(xl):
            df = xl.parse(sheet).fillna("")
            if "Name" not in df.columns:
                continue
            for _, row in df.iterrows():
                name = str(row.get("Name", "")).strip()
                if not name:
                    continue
                norm = name.lower()
                state = runner_state.setdefault(
                    norm,
                    {
                        "club": set(),
                        "gender": set(),
                        "category": set(),
                        "raw_names": set(),
                    },
                )
                state["raw_names"].add(name)

                club = str(row.get("Club", "")).strip()
                if club:
                    state["club"].add(club)

                gender = str(row.get("Gender", "")).strip().upper()
                if gender:
                    state["gender"].add(gender)

                category = str(row.get("Category", "")).strip()
                if category:
                    state["category"].add(category)
    except Exception:
        return {}

    return runner_state


def _detect_runner_anomalies(runner_state: Dict[str, Dict[str, set]]) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    for norm, state in runner_state.items():
        flags: List[str] = []
        detail_parts: List[str] = []

        if len(state["club"]) > 1:
            flags.append("Club")
            detail_parts.append("clubs=" + ", ".join(sorted(state["club"])))

        if len(state["gender"]) > 1:
            flags.append("Gender")
            detail_parts.append("gender=" + ", ".join(sorted(state["gender"])))

        if len(state["category"]) > 1:
            flags.append("Category")
            preview = sorted(state["category"])
            detail_parts.append(
                "categories="
                + ", ".join(preview[:6])
                + ("..." if len(preview) > 6 else "")
            )

        if not flags:
            continue

        raw_name = next(iter(state["raw_names"]))
        rows.append(
            {
                "runner": proper_case(raw_name),
                "anomalies": "/".join(flags),
                "details": " | ".join(detail_parts),
                "processed": bool(load_processed_state().get(raw_name)),
            }
        )

    rows.sort(key=lambda item: item["runner"].lower())
    return rows


def _build_points_map() -> Optional[Dict[str, float]]:
    if session_config.output_dir is None:
        return None
    try:
        wb = find_latest_results_workbook(session_config.output_dir)
        if wb is None:
            return None
        points_map: Dict[str, float] = {}
        xl = pd.ExcelFile(wb)
        for sheet in xl.sheet_names:
            if str(sheet).strip().lower() not in ("male", "female"):
                continue
            try:
                df = xl.parse(sheet).fillna("")
            except Exception:
                continue
            if "Name" not in df.columns:
                continue
            for _, row in df.iterrows():
                name = str(row.get("Name", "")).strip()
                if not name:
                    continue
                norm = name.lower()
                pts = None
                for cname in row.index:
                    if cname and str(cname).lower() in ("total points", "total_points", "points", "pts"):
                        pts = row.get(cname)
                        break
                try:
                    pval = float(pts) if pts is not None and str(pts).strip() != "" else 0.0
                except Exception:
                    pval = 0.0
                points_map[norm] = max(points_map.get(norm, 0.0), pval)
        return points_map
    except Exception:
        return None


def build_raes_runner_rows(show_all: bool = False) -> List[Dict[str, object]]:
    runner_state = _scan_workbook_for_runner_state()
    if not runner_state:
        return []

    anomalies = _detect_runner_anomalies(runner_state)
    processed_map = load_processed_state()
    points_map = _build_points_map()

    rows: List[Dict[str, object]] = []
    for item in anomalies:
        name = item.get("runner", "")
        if points_map is not None and not show_all:
            if points_map.get(name.lower(), 0.0) <= 0.0:
                continue
        rows.append(
            {
                "runner": name,
                "anomalies": item.get("anomalies", ""),
                "details": item.get("details", ""),
                "processed": bool(processed_map.get(name)),
            }
        )

    rows.sort(key=lambda r: r["runner"].lower())
    return rows


class RAESWindow(QMainWindow):
    def __init__(self, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.setWindowTitle("RAES Manual Corrections")
        self.setMinimumSize(980, 720)
        self._rows: List[Dict[str, object]] = []
        self._selected_runner: Optional[str] = None
        self._candidate_checkboxes: Dict[str, QCheckBox] = {}
        self._file_value_labels: Dict[str, QLabel] = {}
        self._candidate_paths: list[Path] = []
        self._source_value_cache: Dict[tuple[str, str, str], str] = {}
        self._scan_queue: Optional[queue.Queue] = None
        self._source_value_queue: Optional[queue.Queue] = None
        self._value_option_queue: Optional[queue.Queue] = None

        self._build_ui()
        self._refresh_list()
        self._start_dirty_timer()

    def _build_ui(self) -> None:
        central = QWidget(self)
        self.setCentralWidget(central)
        layout = QVBoxLayout(central)
        layout.setContentsMargins(12, 12, 12, 12)
        layout.setSpacing(12)

        button_style = (
            "QPushButton { background: #ffffff; color: #3a4658; border: 1px solid #ccd7e3; border-radius: 8px; padding: 8px 14px; }"
            "QPushButton:hover { background: #eef2f7; }"
        )

        title_row = QWidget(self)
        title_row.setStyleSheet("background: #ffffff; border-radius: 12px;")
        title_layout = QHBoxLayout(title_row)
        title_layout.setContentsMargins(16, 16, 16, 16)
        title_layout.setSpacing(12)

        title = QLabel("RAES Manual Corrections", self)
        title.setFont(QFont("Segoe UI", 16, QFont.Bold))
        title.setStyleSheet("color: #2d7a4a;")
        title_layout.addWidget(title)
        title_layout.addStretch(1)
        layout.addWidget(title_row)

        button_bar = QWidget(self)
        button_bar.setStyleSheet("background: #ffffff; border-radius: 12px;")
        button_layout = QHBoxLayout(button_bar)
        button_layout.setContentsMargins(16, 12, 16, 12)
        button_layout.setSpacing(12)

        self._refresh_btn = QPushButton("Refresh", button_bar)
        self._refresh_btn.setCursor(Qt.PointingHandCursor)
        self._refresh_btn.setStyleSheet(button_style)
        self._refresh_btn.setIcon(self.style().standardIcon(QStyle.SP_BrowserReload))
        self._refresh_btn.clicked.connect(self._refresh_list)
        button_layout.addWidget(self._refresh_btn)

        button_layout.addWidget(QLabel("Mode:", button_bar))
        self._mode_combo = QComboBox(button_bar)
        self._mode_combo.addItems(["Anomalies", "Name Review"])
        self._mode_combo.currentTextChanged.connect(self._on_mode_changed)
        button_layout.addWidget(self._mode_combo)

        self._show_all_cb = QCheckBox("Show all runners (include non-league / zero-point runners)", button_bar)
        self._show_all_cb.stateChanged.connect(self._refresh_list)
        button_layout.addWidget(self._show_all_cb)

        self._count_label = QLabel("Anomalies: 0", button_bar)
        self._count_label.setFont(QFont("Segoe UI", 10, QFont.Bold))
        button_layout.addWidget(self._count_label)

        close_btn = QPushButton("🏠 Close", button_bar)
        close_btn.setCursor(Qt.PointingHandCursor)
        close_btn.setStyleSheet(button_style)
        close_btn.clicked.connect(self.close)
        button_layout.addWidget(close_btn)

        button_layout.addStretch(1)

        layout.addWidget(button_bar)

        splitter = QSplitter(Qt.Horizontal, self)
        splitter.setHandleWidth(4)
        splitter.setContentsMargins(0, 0, 0, 0)

        left = QWidget(self)
        left_layout = QVBoxLayout(left)
        left_layout.setContentsMargins(0, 0, 0, 0)
        left_layout.setSpacing(8)

        self._runner_tree = QTreeWidget(self)
        self._runner_tree.setHeaderLabels(["Runner", "Processed"])
        self._runner_tree.setRootIsDecorated(False)
        self._runner_tree.setSelectionBehavior(QTreeWidget.SelectRows)
        self._runner_tree.setSelectionMode(QTreeWidget.SingleSelection)
        self._runner_tree.itemSelectionChanged.connect(self._on_runner_selected)
        left_layout.addWidget(self._runner_tree)

        splitter.addWidget(left)

        right = QWidget(self)
        right_layout = QVBoxLayout(right)
        right_layout.setContentsMargins(0, 0, 0, 0)
        right_layout.setSpacing(10)

        self._summary_label = QLabel("Select a runner from the left to review anomalies.", self)
        self._summary_label.setWordWrap(True)
        self._summary_label.setStyleSheet(f"color: {RAES_NAVY};")
        self._summary_label.setFont(QFont("Segoe UI", 11))
        right_layout.addWidget(self._summary_label)

        self._source_group = QGroupBox("Source Files", self)
        self._source_group.setSizePolicy(QSizePolicy.Preferred, QSizePolicy.Expanding)
        self._source_group.setMinimumHeight(240)
        source_layout = QVBoxLayout(self._source_group)
        source_layout.setContentsMargins(8, 8, 8, 8)
        source_layout.setSpacing(8)

        self._source_scroll = QScrollArea(self)
        self._source_scroll.setWidgetResizable(True)
        self._source_scroll.setMinimumHeight(220)
        self._source_widget = QWidget(self)
        self._source_widget.setLayout(QVBoxLayout())
        self._source_widget.layout().setContentsMargins(0, 0, 0, 0)
        self._source_widget.layout().setSpacing(8)
        self._source_scroll.setWidget(self._source_widget)
        source_layout.addWidget(self._source_scroll)

        right_layout.addWidget(self._source_group)

        action_row = QWidget(self)
        action_layout = QHBoxLayout(action_row)
        action_layout.setContentsMargins(0, 0, 0, 0)
        action_layout.setSpacing(10)

        self._field_label = QLabel("Field:", self)
        action_layout.addWidget(self._field_label)
        self._field_combo = QComboBox(self)
        self._field_combo.addItems(["category", "club", "gender"])
        self._field_combo.currentTextChanged.connect(self._on_field_changed)
        action_layout.addWidget(self._field_combo)

        self._value_label = QLabel("Value:", self)
        action_layout.addWidget(self._value_label)
        self._value_combo = QComboBox(self)
        action_layout.addWidget(self._value_combo)

        self._name_label = QLabel("New name:", self)
        self._name_label.setVisible(False)
        action_layout.addWidget(self._name_label)
        self._name_line = QLineEdit(self)
        self._name_line.setPlaceholderText("Enter corrected runner name")
        self._name_line.setVisible(False)
        action_layout.addWidget(self._name_line)

        self._value_status = QLabel("", self)
        self._value_status.setStyleSheet("color: #888888; font-size: 10px;")
        action_layout.addWidget(self._value_status)

        self._apply_btn = QPushButton("Apply to selected files", self)
        self._apply_btn.setCursor(Qt.PointingHandCursor)
        self._apply_btn.setStyleSheet(button_style)
        self._apply_btn.clicked.connect(self._on_apply)
        action_layout.addWidget(self._apply_btn)

        mark_btn = QPushButton("Mark Reviewed", self)
        mark_btn.setCursor(Qt.PointingHandCursor)
        mark_btn.setStyleSheet(button_style)
        mark_btn.clicked.connect(self._on_mark_reviewed)
        action_layout.addWidget(mark_btn)

        self._power10_btn = QPushButton("Power of 10", self)
        self._power10_btn.setCursor(Qt.PointingHandCursor)
        self._power10_btn.setStyleSheet(button_style)
        self._power10_btn.setToolTip("Open Power of 10 athlete search for the selected runner and copy their details.")
        self._power10_btn.clicked.connect(self._on_power10_search)
        action_layout.addWidget(self._power10_btn)

        action_layout.addStretch(1)
        right_layout.addWidget(action_row)

        diag_panel = QWidget(self)
        diag_panel.setStyleSheet("background: #1e1e1e; border-radius: 12px;")
        diag_layout = QVBoxLayout(diag_panel)
        diag_layout.setContentsMargins(16, 16, 16, 16)
        diag_layout.setSpacing(8)

        diag_title = QLabel("Diagnostics", diag_panel)
        diag_title.setFont(QFont("Segoe UI", 12, QFont.Bold))
        diag_title.setStyleSheet("color: #ffffff;")
        diag_layout.addWidget(diag_title)

        self._diag_text = QTextEdit(diag_panel)
        self._diag_text.setReadOnly(True)
        self._diag_text.setStyleSheet("background: transparent; color: #ffffff; border: none;")
        diag_layout.addWidget(self._diag_text)
        right_layout.addWidget(diag_panel, 1)

        splitter.addWidget(right)
        splitter.setStretchFactor(0, 1)
        splitter.setStretchFactor(1, 2)
        layout.addWidget(splitter)

        status_row = QWidget(self)
        status_layout = QHBoxLayout(status_row)
        status_layout.setContentsMargins(0, 0, 0, 0)
        status_layout.setSpacing(12)

        self._last_updated_label = QLabel("Last updated: -", self)
        status_layout.addWidget(self._last_updated_label)
        self._workbook_label = QLabel("Workbook: -", self)
        status_layout.addWidget(self._workbook_label)
        status_layout.addStretch(1)
        self._dirty_label = QLabel("", self)
        self._dirty_label.setStyleSheet("color: #b22222;")
        status_layout.addWidget(self._dirty_label)

        layout.addWidget(status_row)

        self._field_combo.setCurrentText("category")
        self._populate_value_options("category")

    def _refresh_list(self) -> None:
        if self._scan_queue is not None:
            return

        self._runner_tree.clear()
        self._count_label.setText("Scanning…")
        self._selected_runner = None
        if self._mode_combo.currentText() == "Name Review":
            self._summary_label.setText("Select a variant from the left to review name correction candidates.")
        else:
            self._summary_label.setText("Select a runner from the left to review anomalies.")
        self._clear_source_checkboxes()
        self._diag_text.clear()
        self._refresh_btn.setEnabled(False)
        self._refresh_btn.setText("Refreshing…")

        q: queue.Queue = queue.Queue()
        self._scan_queue = q

        def worker() -> None:
            try:
                rows = self._load_current_mode_rows(self._show_all_cb.isChecked())
                workbook = find_latest_results_workbook(session_config.output_dir)
                q.put((rows, workbook))
            except Exception as exc:
                q.put((None, exc))

        threading.Thread(target=worker, daemon=True).start()
        QTimer.singleShot(100, self._poll_scan)

    def _on_mode_changed(self, mode: str) -> None:
        self._field_label.setVisible(mode == "Anomalies")
        self._field_combo.setVisible(mode == "Anomalies")
        self._value_label.setVisible(mode == "Anomalies")
        self._value_combo.setVisible(mode == "Anomalies")
        self._value_status.setVisible(mode == "Anomalies")
        self._name_label.setVisible(mode == "Name Review")
        self._name_line.setVisible(mode == "Name Review")
        self._show_all_cb.setEnabled(mode == "Anomalies")
        self._apply_btn.setText("Apply name change to selected files" if mode == "Name Review" else "Apply to selected files")
        if mode == "Name Review":
            self._name_line.setText("")
        self._refresh_list()

    def _load_current_mode_rows(self, show_all: bool) -> List[Dict[str, object]]:
        if self._mode_combo.currentText() == "Name Review":
            return self._load_name_review_rows()
        return build_raes_runner_rows(show_all)

    def _load_name_review_rows(self) -> List[Dict[str, object]]:
        workbook = find_latest_audit_workbook()
        if workbook is None:
            return []

        try:
            with pd.ExcelFile(workbook) as xl:
                if "Runner Audit" not in xl.sheet_names:
                    return []
                df = xl.parse("Runner Audit", dtype=str).fillna("")
        except Exception:
            return []

        rows: List[Dict[str, object]] = []
        processed_map = load_processed_state()
        for _, item in df.iterrows():
            if str(item.get("Issue Code", "")).strip() != "AUD-RUNNER-005":
                continue
            display_name = str(item.get("Display Name", "")).strip()
            if not display_name:
                continue
            variant_names = [name.strip() for name in re.split(r"\s*/\s*", display_name) if name.strip()]
            rows.append(
                {
                    "runner": display_name,
                    "variant_names": variant_names,
                    "clubs_seen": str(item.get("Clubs Seen", "")).strip(),
                    "categories_seen": str(item.get("Categories Seen", "")).strip(),
                    "races_seen": str(item.get("Races Seen", "")).strip(),
                    "message": str(item.get("Message", "")).strip(),
                    "next_step": str(item.get("Next Step", "")).strip(),
                    "processed": bool(processed_map.get(display_name)),
                }
            )
        return rows

    def _poll_scan(self) -> None:
        if self._scan_queue is None:
            return
        try:
            payload = self._scan_queue.get_nowait()
        except queue.Empty:
            QTimer.singleShot(100, self._poll_scan)
            return

        self._scan_queue = None
        self._refresh_btn.setEnabled(True)
        self._refresh_btn.setText("Refresh")
        if payload is None or payload[0] is None and isinstance(payload[1], Exception):
            QMessageBox.critical(self, "RAES Error", str(payload[1] if payload else "Unknown scan error."))
            self._count_label.setText("Anomalies: 0")
            return

        rows, workbook = payload
        self._rows = rows or []
        for row in self._rows:
            item = QTreeWidgetItem([str(row.get("runner", "")), "✓" if row.get("processed") else ""])
            self._runner_tree.addTopLevelItem(item)

        if self._mode_combo.currentText() == "Name Review":
            self._count_label.setText(f"Name review items: {len(self._rows)}")
        else:
            self._count_label.setText(f"Anomalies: {len(self._rows)}")
        if workbook is not None:
            age = int((datetime.datetime.now() - datetime.datetime.fromtimestamp(Path(workbook).stat().st_mtime)).total_seconds() // 60)
            self._last_updated_label.setText(f"Last updated: {age} minute(s) ago")
            self._workbook_label.setText(f"Workbook: {Path(workbook).name}")
        else:
            self._last_updated_label.setText("Last updated: -")
            self._workbook_label.setText("Workbook: -")
        self._refresh_btn.setEnabled(True)
        self._refresh_btn.setText("Refresh")

    def _on_runner_selected(self) -> None:
        items = self._runner_tree.selectedItems()
        if not items:
            return
        runner = items[0].text(0)
        self._selected_runner = runner
        self._populate_runner_detail(runner)

    def _populate_runner_detail(self, runner: str) -> None:
        row = next((r for r in self._rows if r.get("runner") == runner), None)
        if row is None:
            return

        if self._mode_combo.currentText() == "Name Review":
            clubs_seen = row.get("clubs_seen", "")
            categories_seen = row.get("categories_seen", "")
            races_seen = row.get("races_seen", "")
            self._summary_label.setText(
                f"<b>{runner}</b><br><br>"
                f"Possible name variant pair: {', '.join(row.get('variant_names', []))}<br>"
                f"Clubs seen: {clubs_seen or 'none'}<br>"
                f"Categories seen: {categories_seen or 'none'}<br>"
                f"Races seen: {races_seen or 'none'}<br>"
                f"Message: {row.get('message', '')}<br>"
                f"Next Step: {row.get('next_step', '')}<br>"
                f"Processed: {'Yes' if row.get('processed') else 'No'}"
            )
        else:
            self._summary_label.setText(
                f"<b>{runner}</b><br><br>"
                f"Anomalies: {row.get('anomalies', '')}<br>"
                f"Details: {row.get('details', '')}<br>"
                f"Processed: {'Yes' if row.get('processed') else 'No'}"
            )
        self._summary_label.setTextFormat(Qt.RichText)

        self._populate_source_files(runner)
        if self._mode_combo.currentText() != "Name Review":
            self._on_field_changed(self._field_combo.currentText())
        self._refresh_diagnostics(runner)

    def _populate_source_files(self, runner: str) -> None:
        self._clear_source_checkboxes()
        try:
            if self._mode_combo.currentText() == "Name Review":
                row = next((r for r in self._rows if r.get("runner") == runner), None)
                candidate_names = row.get("variant_names", []) if row else []
                candidates = self._find_name_review_candidates(candidate_names)
            else:
                candidates = find_candidate_source_files(runner)
            if not candidates.get("series") and not candidates.get("raw"):
                label = QLabel("No candidate source files found for this runner.", self._source_widget)
                self._source_widget.layout().addWidget(label)
                return

            def add_file_row(prefix: str, p: Path) -> None:
                row = QWidget(self._source_widget)
                row.setMinimumHeight(34)
                row_layout = QHBoxLayout(row)
                row_layout.setContentsMargins(0, 0, 0, 0)
                row_layout.setSpacing(10)

                checkbox = QCheckBox(f"{prefix}: {p.name}", row)
                checkbox.setChecked(False)
                checkbox.setFont(QFont("Segoe UI", 10))
                self._candidate_checkboxes[str(p)] = checkbox
                row_layout.addWidget(checkbox)

                value_lbl = QLabel("", row)
                value_lbl.setStyleSheet("color: #444444;")
                value_lbl.setFont(QFont("Segoe UI", 10, QFont.DemiBold))
                row_layout.addWidget(value_lbl)
                row_layout.addStretch(1)
                self._file_value_labels[str(p)] = value_lbl

                self._source_widget.layout().addWidget(row)

            self._candidate_paths = list(candidates.get("series", [])) + list(candidates.get("raw", []))
            for p in candidates.get("series", []):
                add_file_row("Series", p)
            for p in candidates.get("raw", []):
                add_file_row("Raw", p)

            self._refresh_source_file_values_async()
        except Exception:
            label = QLabel("Failed to load candidate source files.", self._source_widget)
            self._source_widget.layout().addWidget(label)

    def _find_name_review_candidates(self, names: List[str]) -> dict:
        candidates = {"series": [], "raw": []}
        seen: set[Path] = set()
        for name in names:
            if not name:
                continue
            result = find_candidate_source_files(name)
            for kind in ("series", "raw"):
                for path in result.get(kind, []):
                    if path not in seen:
                        candidates[kind].append(path)
                        seen.add(path)
        return candidates

    def _clear_source_checkboxes(self) -> None:
        self._candidate_checkboxes.clear()
        self._file_value_labels.clear()
        layout = self._source_widget.layout()
        while layout.count():
            item = layout.takeAt(0)
            widget = item.widget()
            if widget is not None:
                widget.deleteLater()

    def _ensure_raes_output_dir(self) -> Path | None:
        out = session_config.output_dir
        if out is None:
            return None
        d = Path(out) / "raes"
        d.mkdir(parents=True, exist_ok=True)
        return d

    def _write_raes_changes_json(self, audit: List[Dict[str, object]]) -> None:
        outdir = self._ensure_raes_output_dir()
        if outdir is None:
            return
        changes_file = outdir / "changes.json"
        existing = []
        try:
            if changes_file.exists():
                existing = json.loads(changes_file.read_text(encoding="utf-8"))
        except Exception:
            existing = []
        existing.extend(audit)
        try:
            changes_file.write_text(json.dumps(existing, indent=2), encoding="utf-8")
        except Exception:
            pass

    def _append_name_corrections(self, variants: List[str], target_name: str) -> None:
        if session_config.control_dir is None:
            return
        name_lookup_path = Path(session_config.control_dir) / "name_corrections.xlsx"
        selected = [
            {"current_name": variant, "proposed_name": target_name}
            for variant in variants
            if variant and variant.strip().lower() != target_name.strip().lower()
        ]
        if not selected:
            return
        try:
            append_name_corrections(name_lookup_path, selected)
        except Exception:
            pass

    def _populate_value_options(self, field: str) -> bool:
        if self._mode_combo.currentText() == "Name Review":
            return False
        self._value_combo.clear()
        self._set_value_status("")
        self._value_combo.setEnabled(True)
        if self._selected_runner is None:
            return False

        if field == "category":
            opts = ["Jun", "Sen", "V40", "V50", "V60", "V70"]
            self._value_combo.addItems(opts)
        elif field == "gender":
            self._value_combo.addItems(["Male", "Female"])
        else:
            self._load_club_value_options_async()
            return False

        if self._value_combo.count() > 0:
            self._value_combo.setCurrentIndex(0)
        return True

    def _on_field_changed(self, field: str) -> None:
        if self._mode_combo.currentText() == "Name Review":
            return
        if self._populate_value_options(field) and self._selected_runner:
            self._refresh_source_file_values_async()

    def _on_power10_search(self) -> None:
        if self._selected_runner is None:
            QMessageBox.information(self, "Power of 10", "Select a runner before opening Power of 10.")
            return

        runner_name = self._selected_runner.strip()
        club_name = self._find_runner_club()
        url = self._build_power10_url(runner_name, club_name)
        webbrowser.open(url)

        clipboard_text = self._build_power10_clipboard_text(runner_name, club_name)
        QGuiApplication.clipboard().setText(clipboard_text)
        QMessageBox.information(
            self,
            "Power of 10",
            "Opened Power of 10 athlete search. Runner details have been copied to the clipboard.\nPaste them into the search form if needed.",
        )

    def _find_runner_club(self) -> str:
        if self._selected_runner is None:
            return ""
        try:
            state = _scan_workbook_for_runner_state().get(self._selected_runner.lower(), {})
            clubs = sorted(state.get("club", []))
            return clubs[0] if clubs else ""
        except Exception:
            return ""

    def _build_power10_url(self, runner_name: str, club_name: str) -> str:
        last_name, first_name = self._split_runner_name(runner_name)
        params = {
            "searchLastName": last_name,
            "searchFirstName": first_name,
            "searchClubName": club_name,
        }
        query = urllib.parse.urlencode({k: v for k, v in params.items() if v})
        return f"https://www.powerof10.uk/Home/AthleteSearch?{query}"

    def _build_power10_clipboard_text(self, runner_name: str, club_name: str) -> str:
        last_name, first_name = self._split_runner_name(runner_name)
        lines = [f"Last Name: {last_name}", f"First Name: {first_name}"]
        if club_name:
            lines.append(f"Club Name: {club_name}")
        return "\n".join(lines)

    def _split_runner_name(self, runner_name: str) -> tuple[str, str]:
        parts = runner_name.split()
        if len(parts) <= 1:
            return runner_name, ""
        return parts[-1], " ".join(parts[:-1])

    def _load_club_value_options_async(self) -> None:
        self._value_combo.clear()
        self._value_combo.addItem("Loading…")
        self._value_combo.setEnabled(False)
        self._set_value_status("Scanning candidate files for club values…")

        runner = self._selected_runner
        paths = list(self._candidate_paths)
        q = queue.Queue()
        self._value_option_queue = q

        def worker() -> None:
            opts_set: set[str] = set()
            for p in paths:
                try:
                    wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
                except Exception:
                    continue
                try:
                    for sname in wb.sheetnames:
                        try:
                            ws = wb[sname]
                            name_col, club_col = _find_columns(ws, "club")
                            if name_col is None or club_col is None:
                                continue
                            for ri in range(2, ws.max_row + 1):
                                nm = _row_name_value(ws, ri, name_col)
                                if nm is None or nm.strip().lower() != self._selected_runner.lower():
                                    continue
                                val = "" if ws.cell(row=ri, column=club_col).value is None else str(ws.cell(row=ri, column=club_col).value).strip()
                                if val:
                                    opts_set.add(val)
                        except Exception:
                            continue
                finally:
                    try:
                        wb.close()
                    except Exception:
                        pass
            q.put((runner, sorted(opts_set)))

        threading.Thread(target=worker, daemon=True).start()
        QTimer.singleShot(100, self._poll_value_option_results)

    def _poll_value_option_results(self) -> None:
        if self._value_option_queue is None:
            return
        try:
            runner, opts = self._value_option_queue.get_nowait()
        except queue.Empty:
            QTimer.singleShot(100, self._poll_value_option_results)
            return

        self._value_option_queue = None
        if self._selected_runner != runner or self._field_combo.currentText() != "club":
            return

        self._value_combo.clear()
        if opts:
            self._value_combo.addItems(opts)
        else:
            self._value_combo.addItem("")
        self._value_combo.setCurrentIndex(0)
        self._value_combo.setEnabled(True)
        self._set_value_status("")

        if self._selected_runner:
            self._refresh_source_file_values_async()

    def _set_value_status(self, text: str) -> None:
        self._value_status.setText(text)

    def _refresh_source_file_values_async(self) -> None:
        if not self._selected_runner:
            return
        runner = self._selected_runner
        mode = self._mode_combo.currentText()
        field = self._field_combo.currentText()
        for label in self._file_value_labels.values():
            label.setText("[loading]")

        q = queue.Queue()
        self._source_value_queue = q
        paths = list(self._file_value_labels.keys())

        def worker() -> None:
            results: Dict[str, str] = {}
            for path_str in paths:
                try:
                    if mode == "Name Review":
                        value = self._read_name_review_value_for_file(Path(path_str), runner)
                    else:
                        value = self._read_runner_value_for_file(Path(path_str), runner, field)
                except Exception:
                    value = "[error]"
                results[path_str] = value
            q.put((runner, mode, results))

        threading.Thread(target=worker, daemon=True).start()
        QTimer.singleShot(100, self._poll_source_value_results)

    def _poll_source_value_results(self) -> None:
        if self._source_value_queue is None:
            return
        try:
            runner, mode, results = self._source_value_queue.get_nowait()
        except queue.Empty:
            QTimer.singleShot(100, self._poll_source_value_results)
            return

        if runner != self._selected_runner or mode != self._mode_combo.currentText():
            return

        for path_str, label in self._file_value_labels.items():
            value = results.get(path_str, "")
            if value == "":
                label.setText("[missing]")
            elif value == "[error]":
                label.setText("[error]")
            else:
                label.setText(f"[{value}]")

    def _read_runner_value_for_file(self, path: Path, runner: str, field: str) -> str:
        cache_key = (runner, field, str(path))
        if cache_key in self._source_value_cache:
            return self._source_value_cache[cache_key]

        runner_key = runner.lower()
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            self._source_value_cache[cache_key] = ""
            return ""

        try:
            for sname in wb.sheetnames:
                ws = wb[sname]
                name_col, field_col = _find_columns(ws, field)
                if name_col is None or field_col is None:
                    continue
                for ri in range(2, ws.max_row + 1):
                    nm = _row_name_value(ws, ri, name_col)
                    if nm is None or nm.strip().lower() != runner_key:
                        continue
                    cell = ws.cell(row=ri, column=field_col)
                    result = "" if cell.value is None else str(cell.value).strip()
                    self._source_value_cache[cache_key] = result
                    return result
        finally:
            try:
                wb.close()
            except Exception:
                pass

        self._source_value_cache[cache_key] = ""
        return ""

    def _read_name_review_value_for_file(self, path: Path, runner: str) -> str:
        cache_key = (runner, "name_review", str(path))
        if cache_key in self._source_value_cache:
            return self._source_value_cache[cache_key]

        row = next((r for r in self._rows if r.get("runner") == runner), None)
        variants = [name.lower() for name in row.get("variant_names", [])] if row else [runner.lower()]
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        except Exception:
            self._source_value_cache[cache_key] = ""
            return ""

        try:
            for sname in wb.sheetnames:
                ws = wb[sname]
                name_col = self._find_name_columns(ws)
                if name_col is None:
                    continue
                for ri in range(2, ws.max_row + 1):
                    nm = _row_name_value(ws, ri, name_col)
                    if nm is None or nm.strip().lower() not in variants:
                        continue
                    result = nm.strip()
                    self._source_value_cache[cache_key] = result
                    return result
        finally:
            try:
                wb.close()
            except Exception:
                pass

        self._source_value_cache[cache_key] = ""
        return ""

    def _find_name_columns(self, ws) -> int | tuple[int, int] | None:
        headers = [
            str(c.value).strip().lower() if c.value is not None else ""
            for c in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        name_col = next(
            (i + 1 for i, h in enumerate(headers) if "name" in h and "first" not in h and "last" not in h),
            None,
        )
        if name_col is not None:
            return name_col

        first_col = next((i + 1 for i, h in enumerate(headers) if "first" in h), None)
        last_col = next((i + 1 for i, h in enumerate(headers) if "last" in h), None)
        if first_col is not None and last_col is not None:
            return (first_col, last_col)
        return None

    def _set_name_value(self, ws, row_idx: int, name_col: int | tuple[int, int], target_name: str) -> None:
        if isinstance(name_col, tuple):
            parts = target_name.split()
            first_value = parts[0] if parts else ""
            last_value = " ".join(parts[1:]) if len(parts) > 1 else ""
            ws.cell(row=row_idx, column=name_col[0]).value = first_value
            ws.cell(row=row_idx, column=name_col[1]).value = last_value
        else:
            ws.cell(row=row_idx, column=name_col).value = target_name

    def _apply_name_to_files(self, files: List[Path], runner: str, target_name: str) -> List[Dict[str, object]]:
        audit: List[Dict[str, object]] = []
        row = next((r for r in self._rows if r.get("runner") == runner), None)
        variants = [name.lower() for name in row.get("variant_names", [])] if row else [runner.lower()]

        for path in files:
            try:
                wb = openpyxl.load_workbook(path)
            except Exception as exc:
                audit.append({"file": str(path), "error": str(exc)})
                continue

            try:
                changed = False
                for sname in wb.sheetnames:
                    ws = wb[sname]
                    name_col = self._find_name_columns(ws)
                    if name_col is None:
                        continue
                    for ri in range(2, ws.max_row + 1):
                        current = _row_name_value(ws, ri, name_col)
                        if current is None or current.strip().lower() not in variants:
                            continue
                        old = str(current).strip()
                        self._set_name_value(ws, ri, name_col, target_name)
                        changed = True
                        audit.append(
                            {
                                "timestamp": datetime.datetime.utcnow().isoformat() + "Z",
                                "file": str(path),
                                "file_path": str(path),
                                "sheet": sname,
                                "runner": runner,
                                "field": "name",
                                "old_value": old,
                                "new_value": target_name,
                                "row_idx": ri,
                            }
                        )
                if changed:
                    _atomic_save(wb, path)
            except Exception as exc:
                audit.append({"file": str(path), "error": str(exc)})
            finally:
                try:
                    wb.close()
                except Exception:
                    pass

        if audit:
            self._write_raes_changes_json(audit)
            try:
                err = log_manual_data_changes(audit, source="RAES", action="raes_applied_name")
                if err:
                    log_event("raes_manual_audit_error", year=session_config.year, error=err)
                else:
                    log_event("raes_manual_audit_logged", year=session_config.year, file_changes=len(audit))
            except Exception:
                pass
            self._append_name_corrections(variants, target_name)

        return audit

    def _on_apply(self) -> None:
        if self._selected_runner is None:
            QMessageBox.warning(self, "Apply", "Select a runner before applying changes.")
            return
        files = [Path(path) for path, cb in self._candidate_checkboxes.items() if cb.isChecked()]
        if not files:
            QMessageBox.warning(self, "Apply", "No source files selected. Please check at least one source file.")
            return

        if self._mode_combo.currentText() == "Name Review":
            value = self._name_line.text().strip()
            if not value:
                QMessageBox.warning(self, "Apply", "Enter a corrected runner name before applying.")
                return
            if not QMessageBox.question(
                self,
                "Apply Name Change",
                f"Apply name change: set selected variants to '{value}' in {len(files)} file(s)?",
                QMessageBox.Yes | QMessageBox.No,
            ) == QMessageBox.Yes:
                return
            audit = self._apply_name_to_files(files, self._selected_runner, value)
            self._source_value_cache.clear()
            changed = [a for a in audit if a.get("runner")]
            errors = [a for a in audit if a.get("error")]
            msg = f"Applied edits: {len(changed)} change(s)."
            if errors:
                msg += f" {len(errors)} error(s) occurred. See RAES changes file for details."
            QMessageBox.information(self, "Apply Results", msg)
            set_runner_processed(self._selected_runner, True)
            self._mark_runner_processed(self._selected_runner)
            self._refresh_diagnostics(self._selected_runner)
            return

        field = self._field_combo.currentText()
        value = self._value_combo.currentText()
        if not field or not value:
            QMessageBox.warning(self, "Apply", "Field and value must be selected before applying.")
            return

        if not QMessageBox.question(
            self,
            "Apply Changes",
            f"Apply change: set {field} → {value} for {self._selected_runner} in {len(files)} file(s)?",
            QMessageBox.Yes | QMessageBox.No,
        ) == QMessageBox.Yes:
            return

        try:
            audit = apply_field_to_files(files, self._selected_runner, field, value)
        except Exception as exc:
            QMessageBox.critical(self, "Apply Error", f"Failed to apply changes: {exc}")
            return
        self._source_value_cache.clear()

        changed = [a for a in audit if a.get("runner")]
        errors = [a for a in audit if a.get("error")]
        msg = f"Applied edits: {len(changed)} change(s)."
        if errors:
            msg += f" {len(errors)} error(s) occurred. See RAES changes file for details."
        QMessageBox.information(self, "Apply Results", msg)
        set_runner_processed(self._selected_runner, True)
        self._mark_runner_processed(self._selected_runner)
        self._refresh_diagnostics(self._selected_runner)

    def _set_runner_item_processed(self, runner: str, processed: bool) -> None:
        for row in self._rows:
            if row.get("runner") == runner:
                row["processed"] = processed
                break
        for i in range(self._runner_tree.topLevelItemCount()):
            item = self._runner_tree.topLevelItem(i)
            if item.text(0) == runner:
                item.setText(1, "✓" if processed else "")
                break

    def _mark_runner_processed(self, runner: str) -> None:
        self._set_runner_item_processed(runner, True)
        if self._selected_runner == runner:
            self._summary_label.setText(
                self._summary_label.text().replace("Processed: No", "Processed: Yes")
            )

    def _on_mark_reviewed(self) -> None:
        if self._selected_runner is None:
            QMessageBox.warning(self, "Mark Reviewed", "Select a runner before marking reviewed.")
            return
        set_runner_processed(self._selected_runner, True)
        self._mark_runner_processed(self._selected_runner)
        log_event("raes_mark_reviewed", year=session_config.year, runner=self._selected_runner)

    def _refresh_diagnostics(self, runner: str) -> None:
        lines: List[str] = []
        changes_file = None
        if session_config.output_dir is not None:
            changes_file = Path(session_config.output_dir) / "raes" / "changes.json"
        recent_changes = []
        if changes_file is not None and changes_file.exists():
            try:
                recent_changes = json.loads(changes_file.read_text(encoding="utf-8"))
            except Exception:
                recent_changes = []
        rc_for_runner = [c for c in recent_changes if str(c.get("runner", "")).lower() == runner.lower()]
        if rc_for_runner:
            lines.append("Recent Changes:")
            for ch in rc_for_runner[-6:]:
                ts = ch.get("timestamp", "")
                file_name = Path(str(ch.get("file", ch.get("file_path", "")))).name
                fld = ch.get("field", "")
                old = ch.get("old_value", ch.get("old", ""))
                nv = ch.get("new_value", ch.get("new", ""))
                lines.append(f"{ts} — {file_name} — {fld}: {old} → {nv}")
        else:
            lines.append("Recent Changes: none")

        try:
            runner_state = _scan_workbook_for_runner_state()
            state = runner_state.get(runner.lower())
            if state:
                if len(state.get("club", set())) == 1:
                    lines.append(f"Club: consistent — {next(iter(state['club']))}")
                if len(state.get("gender", set())) == 1:
                    lines.append(f"Gender: consistent — {next(iter(state['gender']))}")
                if len(state.get("category", set())) == 1:
                    lines.append(f"Category: consistent — {next(iter(state['category']))}")
                if any(not v for v in [state.get("club"), state.get("gender"), state.get("category")]):
                    lines.append("Warning: Some fields may be missing in the latest results workbook.")
        except Exception:
            pass

        self._diag_text.setPlainText("\n".join(lines))

    def _start_dirty_timer(self) -> None:
        QTimer.singleShot(1000, self._update_dirty_indicator)

    def _update_dirty_indicator(self) -> None:
        is_dirty = False
        if session_config.output_dir is not None:
            ap_flag = Path(session_config.output_dir) / "autopilot" / "dirty"
            raes_flag = Path(session_config.output_dir) / "raes" / "dirty"
            is_dirty = ap_flag.exists() or raes_flag.exists()
        self._dirty_label.setText("DATA DIRTY — run Autopilot" if is_dirty else "")
        QTimer.singleShot(1000, self._update_dirty_indicator)
