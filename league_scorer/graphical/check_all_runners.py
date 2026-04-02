"""
check_all_runners.py — Suggest and apply club backfills across all race files.

Finds runners with blank club entries in some races but a valid eligible club in
other races, then lets the user manually approve suggested fixes before writing.
"""

from __future__ import annotations

import tkinter as tk
from collections import Counter, defaultdict
from pathlib import Path
from tkinter import messagebox, ttk

import openpyxl
import pandas as pd

from ..common_files import race_discovery_exclusions
from ..manual_data_audit import log_manual_data_changes
from ..manual_edit_service import apply_club_suggestions
from ..source_loader import discover_race_files
from ..session_config import config as session_config
from .dashboard import WRRL_GREEN, WRRL_LIGHT, WRRL_NAVY, WRRL_WHITE


class CheckAllRunnersPanel(tk.Frame):
    """Embedded panel for cross-race blank-club checks with manual approval."""

    def __init__(self, parent: tk.Misc, back_callback=None) -> None:
        super().__init__(parent, bg=WRRL_LIGHT)
        self._back_callback = back_callback

        self._eligible_clubs: set[str] = set()
        self._suggestions: list[dict] = []
        self._selected: dict[str, bool] = {}

        self._build_ui()
        self._scan_all_races()

    # ── UI ───────────────────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        header = tk.Frame(self, bg=WRRL_NAVY, padx=16, pady=10)
        header.pack(fill="x")
        tk.Label(
            header,
            text="Check All Runners",
            font=("Segoe UI", 16, "bold"),
            bg=WRRL_NAVY,
            fg=WRRL_WHITE,
        ).pack(side="left")

        if self._back_callback:
            tk.Button(
                header,
                text="\u25c4 Dashboard",
                command=self._on_back,
                font=("Segoe UI", 10, "bold"),
                bg=WRRL_GREEN,
                fg=WRRL_WHITE,
                relief="flat",
                padx=10,
                pady=4,
                cursor="hand2",
                activebackground="#1f5632",
                activeforeground=WRRL_WHITE,
            ).pack(side="right")

        top = tk.Frame(self, bg=WRRL_LIGHT, padx=16, pady=10)
        top.pack(fill="x")

        tk.Label(
            top,
            text=(
                "Find runners with blank club entries in some races but valid clubs in others. "
                "Review the suggestions, tick/untick manually, then apply selected changes."
            ),
            font=("Segoe UI", 10),
            bg=WRRL_LIGHT,
            fg=WRRL_NAVY,
            wraplength=980,
            justify="left",
            anchor="w",
        ).pack(fill="x")

        actions = tk.Frame(self, bg=WRRL_LIGHT, padx=16, pady=6)
        actions.pack(fill="x")

        tk.Button(
            actions,
            text="Rescan",
            command=self._scan_all_races,
            font=("Segoe UI", 10),
            bg="#dbe1e8",
            fg=WRRL_NAVY,
            relief="flat",
            padx=10,
            pady=4,
            cursor="hand2",
        ).pack(side="left")

        tk.Button(
            actions,
            text="Select All",
            command=lambda: self._set_all_selected(True),
            font=("Segoe UI", 10),
            bg="#dbe1e8",
            fg=WRRL_NAVY,
            relief="flat",
            padx=10,
            pady=4,
            cursor="hand2",
        ).pack(side="left", padx=(8, 0))

        tk.Button(
            actions,
            text="Clear All",
            command=lambda: self._set_all_selected(False),
            font=("Segoe UI", 10),
            bg="#dbe1e8",
            fg=WRRL_NAVY,
            relief="flat",
            padx=10,
            pady=4,
            cursor="hand2",
        ).pack(side="left", padx=(8, 0))

        self._apply_btn = tk.Button(
            actions,
            text="Apply Selected",
            command=self._apply_selected,
            font=("Segoe UI", 10, "bold"),
            bg=WRRL_GREEN,
            fg=WRRL_WHITE,
            relief="flat",
            padx=12,
            pady=4,
            cursor="hand2",
            state="disabled",
            activebackground="#1f5632",
            activeforeground=WRRL_WHITE,
        )
        self._apply_btn.pack(side="left", padx=(12, 0))

        table_frame = tk.Frame(self, bg=WRRL_LIGHT, padx=16, pady=12)
        table_frame.pack(fill="both", expand=True)

        cols = ("apply", "race", "file", "name", "suggested", "evidence")
        self._tree = ttk.Treeview(table_frame, columns=cols, show="headings")
        self._tree.heading("apply", text="Apply")
        self._tree.heading("race", text="Race")
        self._tree.heading("file", text="File")
        self._tree.heading("name", text="Runner")
        self._tree.heading("suggested", text="Suggested Club")
        self._tree.heading("evidence", text="Evidence")

        self._tree.column("apply", width=60, anchor="center")
        self._tree.column("race", width=55, anchor="center")
        self._tree.column("file", width=230, anchor="w")
        self._tree.column("name", width=220, anchor="w")
        self._tree.column("suggested", width=180, anchor="w")
        self._tree.column("evidence", width=360, anchor="w")

        y_scroll = ttk.Scrollbar(table_frame, orient="vertical", command=self._tree.yview)
        x_scroll = ttk.Scrollbar(table_frame, orient="horizontal", command=self._tree.xview)
        self._tree.configure(yscrollcommand=y_scroll.set, xscrollcommand=x_scroll.set)

        self._tree.grid(row=0, column=0, sticky="nsew")
        y_scroll.grid(row=0, column=1, sticky="ns")
        x_scroll.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        self._tree.bind("<Button-1>", self._on_tree_click)

        self._status_var = tk.StringVar(value="Scanning...")
        tk.Label(
            self,
            textvariable=self._status_var,
            font=("Segoe UI", 9, "italic"),
            bg=WRRL_LIGHT,
            fg="#666666",
            anchor="w",
        ).pack(fill="x", padx=16, pady=(0, 10))

    # ── scan logic ───────────────────────────────────────────────────────────

    def _load_eligible_clubs(self) -> None:
        self._eligible_clubs = set()
        input_dir = session_config.input_dir
        if not input_dir:
            return
        clubs_path = input_dir / "clubs.xlsx"
        if not clubs_path.exists():
            return
        try:
            df = pd.read_excel(clubs_path, dtype=str)
        except Exception:
            return

        for col in ("Club", "Preferred name"):
            if col in df.columns:
                self._eligible_clubs.update(
                    v.strip().lower() for v in df[col].dropna() if str(v).strip()
                )

    def _extract_rows(self, filepath: Path) -> tuple[list[dict], int | None]:
        """Read runner rows from workbook, return entries and 1-based club column."""
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True)
        except Exception:
            return [], None
        try:
            ws = wb.active
            header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_row:
                return [], None

            headers = [str(h).strip().lower() if h is not None else "" for h in header_row]
            name_idx = next(
                (i for i, h in enumerate(headers) if "name" in h and "first" not in h and "last" not in h),
                None,
            )
            if name_idx is None:
                first_idx = next((i for i, h in enumerate(headers) if "first" in h), None)
                last_idx = next((i for i, h in enumerate(headers) if "last" in h), None)
                if first_idx is not None and last_idx is not None:
                    name_idx = (first_idx, last_idx)

            club_idx = next((i for i, h in enumerate(headers) if "club" in h), None)
            if name_idx is None or club_idx is None:
                return [], None

            rows: list[dict] = []
            for row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if isinstance(name_idx, tuple):
                    first_name = str(row[name_idx[0]] or "").strip()
                    last_name = str(row[name_idx[1]] or "").strip()
                    name = f"{first_name} {last_name}".strip()
                else:
                    name = str(row[name_idx] or "").strip()

                raw_club = row[club_idx] if club_idx < len(row) else None
                club = str(raw_club).strip() if raw_club is not None else ""
                if club.lower() in {"nan", "none"}:
                    club = ""

                if name:
                    rows.append(
                        {
                            "row_idx": row_no,
                            "name": name,
                            "club": club,
                        }
                    )

            return rows, club_idx + 1
        finally:
            wb.close()

    def _scan_all_races(self) -> None:
        input_dir = session_config.input_dir
        if not input_dir or not input_dir.exists():
            self._status_var.set("Input directory not available.")
            return

        self._load_eligible_clubs()
        race_files = discover_race_files(
            input_dir,
            excluded_names=race_discovery_exclusions(),
        )

        by_name: dict[str, list[dict]] = defaultdict(list)
        for race_num, filepath in race_files.items():
            rows, club_col_1based = self._extract_rows(filepath)
            if not rows or club_col_1based is None:
                continue
            for entry in rows:
                by_name[entry["name"].strip().lower()].append(
                    {
                        "race": race_num,
                        "file": filepath,
                        "file_name": filepath.name,
                        "row_idx": entry["row_idx"],
                        "name": entry["name"],
                        "club": entry["club"],
                        "club_col": club_col_1based,
                    }
                )

        suggestions: list[dict] = []
        for _, entries in by_name.items():
            valid = [
                e["club"]
                for e in entries
                if e["club"] and e["club"].strip().lower() in self._eligible_clubs
            ]
            if not valid:
                continue

            counts = Counter(valid)
            suggested_club, count = sorted(counts.items(), key=lambda t: (-t[1], t[0].lower()))[0]
            evidence = f"Seen as '{suggested_club}' in {count}/{len(valid)} valid entries"

            for e in entries:
                if e["club"]:
                    continue
                key = f"{e['file']}::{e['row_idx']}"
                suggestions.append(
                    {
                        "key": key,
                        "race": e["race"],
                        "file": e["file"],
                        "file_name": e["file_name"],
                        "row_idx": e["row_idx"],
                        "name": e["name"],
                        "club_col": e["club_col"],
                        "suggested_club": suggested_club,
                        "evidence": evidence,
                    }
                )

        suggestions.sort(key=lambda s: (s["race"], s["file_name"].lower(), s["name"].lower(), s["row_idx"]))
        self._suggestions = suggestions
        self._selected = {s["key"]: True for s in suggestions}
        self._refresh_table()

    # ── table / selection ────────────────────────────────────────────────────

    def _refresh_table(self) -> None:
        for iid in self._tree.get_children():
            self._tree.delete(iid)

        selected_count = 0
        for idx, s in enumerate(self._suggestions):
            checked = self._selected.get(s["key"], False)
            selected_count += 1 if checked else 0
            marker = "☑" if checked else "☐"
            self._tree.insert(
                "",
                "end",
                iid=str(idx),
                values=(
                    marker,
                    s["race"],
                    s["file_name"],
                    s["name"],
                    s["suggested_club"],
                    s["evidence"],
                ),
            )

        self._apply_btn.config(state="normal" if selected_count > 0 else "disabled")
        self._status_var.set(
            f"Found {len(self._suggestions)} suggestions across {len({s['file'] for s in self._suggestions})} files; "
            f"{selected_count} selected."
            if self._suggestions
            else "No suggestions found."
        )

    def _on_tree_click(self, event) -> None:
        region = self._tree.identify("region", event.x, event.y)
        if region != "cell":
            return
        col = self._tree.identify_column(event.x)
        row_id = self._tree.identify_row(event.y)
        if not row_id or col != "#1":
            return

        idx = int(row_id)
        suggestion = self._suggestions[idx]
        key = suggestion["key"]
        self._selected[key] = not self._selected.get(key, False)
        self._refresh_table()

    def _set_all_selected(self, value: bool) -> None:
        for s in self._suggestions:
            self._selected[s["key"]] = value
        self._refresh_table()

    # ── apply ────────────────────────────────────────────────────────────────

    def _apply_selected(self) -> None:
        selected = [s for s in self._suggestions if self._selected.get(s["key"], False)]
        if not selected:
            messagebox.showinfo("Nothing Selected", "No suggestions are selected.", parent=self)
            return

        if not messagebox.askyesno(
            "Apply Club Suggestions",
            f"Apply {len(selected)} selected club updates across race files?",
            parent=self,
        ):
            return

        updates_by_file: dict[Path, list[dict]] = defaultdict(list)
        for s in selected:
            updates_by_file[s["file"]].append(s)

        applied, audit_changes, failed = apply_club_suggestions(updates_by_file)

        # Remove successfully applied suggestions from in-memory list
        if applied:
            selected_keys = {s["key"] for s in selected}
            self._suggestions = [s for s in self._suggestions if s["key"] not in selected_keys]
            for k in list(self._selected.keys()):
                if k in selected_keys:
                    self._selected.pop(k, None)

        self._refresh_table()

        log_error = log_manual_data_changes(
            audit_changes,
            source="Check All Runners",
            action="Bulk update club",
        )
        if log_error:
            failed.append(f"Manual_Data_Audit: {log_error}")

        if failed:
            messagebox.showwarning(
                "Applied With Issues",
                f"Applied {applied} updates, but some files failed:\n\n" + "\n".join(failed),
                parent=self,
            )
        else:
            messagebox.showinfo(
                "Apply Complete",
                f"Applied {applied} club update(s).",
                parent=self,
            )

    def _on_back(self) -> None:
        if self._back_callback:
            self._back_callback()
