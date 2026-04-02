"""Central issue-resolution helpers used by Issue Review and related UIs."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Mapping

import openpyxl

from .common_files import race_discovery_exclusions
from .issue_tracking import build_issue_identity
from .manual_data_audit import log_manual_data_changes
from .manual_edit_service import resolve_runner_field_across_files
from .normalisation import parse_time_to_seconds
from .race_processor import extract_race_number
from .source_loader import discover_race_files

_SUPPORTED_QUICK_FIX_CODES = {"AUD-ROW-002"}


@dataclass
class QuickFixResult:
    success: bool
    message: str
    updated_rows: int = 0
    touched_files: int = 0
    failures: list[str] | None = None
    verified_resolved: bool = False
    issue_identity: str = ""


def supports_quick_fix(issue_code: str) -> bool:
    return issue_code in _SUPPORTED_QUICK_FIX_CODES


def quick_fix_requires_input(issue_code: str) -> bool:
    return issue_code == "AUD-ROW-002"


def quick_fix_prompt(issue_code: str, runner_name: str) -> tuple[str, str]:
    if issue_code == "AUD-ROW-002":
        return (
            "Fix Invalid Time",
            f"Enter corrected time for '{runner_name}' (hh:mm:ss):",
        )
    return ("Quick Fix", "Enter value:")


def apply_quick_fix_for_issue(
    issue: Mapping[str, str],
    *,
    input_dir: Path | None,
    target_value: str | None = None,
) -> QuickFixResult:
    issue_code = str(issue.get("Issue Code", "")).strip()
    runner_name = str(issue.get("Name", "")).strip()
    issue_identity = build_issue_identity(issue)

    if not supports_quick_fix(issue_code):
        return QuickFixResult(
            success=False,
            message=f"No quick fix is available for issue code {issue_code}.",
            issue_identity=issue_identity,
        )

    if not input_dir or not input_dir.exists():
        return QuickFixResult(
            success=False,
            message="Active input directory is not available.",
            issue_identity=issue_identity,
        )

    race_files = discover_race_files(
        input_dir,
        excluded_names=race_discovery_exclusions(),
    )
    if not race_files:
        return QuickFixResult(
            success=False,
            message="No race files found in the active input directory.",
            issue_identity=issue_identity,
        )

    if issue_code == "AUD-ROW-002":
        if not runner_name:
            return QuickFixResult(
                success=False,
                message="Runner name is required for this quick fix.",
                issue_identity=issue_identity,
            )
        normalised = _normalise_time_value(target_value or "")
        if not normalised:
            return QuickFixResult(
                success=False,
                message="Invalid time format. Use hh:mm:ss.",
                issue_identity=issue_identity,
            )

        updated_rows, touched_files, audit_changes, failed_files = resolve_runner_field_across_files(
            race_files,
            selected_runner=runner_name,
            field_type="time",
            target_value=normalised,
        )

        log_error = log_manual_data_changes(
            audit_changes,
            source="Issue Review",
            action=f"Quick fix {issue_code}",
        )
        if log_error:
            failed_files.append(f"Manual_Data_Audit: {log_error}")

        if updated_rows <= 0:
            return QuickFixResult(
                success=False,
                message=f"No matching rows were updated for runner '{runner_name}'.",
                updated_rows=updated_rows,
                touched_files=touched_files,
                failures=failed_files,
                issue_identity=issue_identity,
            )

        verified_resolved, verification_note = _verify_issue_resolved(issue, race_files)
        if verified_resolved:
            resolution_log_error = log_manual_data_changes(
                [
                    {
                        "runner": runner_name,
                        "field": "audit_issue",
                        "old_value": "Open",
                        "new_value": "Resolved",
                        "file_path": race_files.get(_parse_race_num(issue.get("Race", "")), ""),
                        "row_idx": issue.get("Source Row", ""),
                        "issue_identity": issue_identity,
                        "issue_code": issue_code,
                        "issue_race": issue.get("Race", ""),
                        "issue_source_row": issue.get("Source Row", ""),
                        "resolution_status": "Resolved",
                        "verification_note": verification_note,
                    }
                ],
                source="Issue Review",
                action=f"Resolve {issue_code}",
            )
            if resolution_log_error:
                failed_files.append(f"Manual_Data_Audit: {resolution_log_error}")

        msg = f"Updated {updated_rows} row(s) across {touched_files} file(s)."
        if verified_resolved:
            msg += "\n\nResolution verified and issue marked as resolved in Manual_Data_Audit."
        else:
            msg += f"\n\nResolution could not be verified automatically: {verification_note}"
        if failed_files:
            msg += "\n\nSome files failed:\n" + "\n".join(failed_files)

        return QuickFixResult(
            success=True,
            message=msg,
            updated_rows=updated_rows,
            touched_files=touched_files,
            failures=failed_files,
            verified_resolved=verified_resolved,
            issue_identity=issue_identity,
        )

    return QuickFixResult(
        success=False,
        message=f"Issue code {issue_code} is not mapped to a quick fix.",
        issue_identity=issue_identity,
    )


def _normalise_time_value(value: str) -> str | None:
    text = str(value).strip()
    if not text:
        return None
    seconds = parse_time_to_seconds(text)
    if seconds is None or seconds <= 0:
        return None
    total = int(seconds)
    hours = total // 3600
    minutes = (total % 3600) // 60
    secs = total % 60
    return f"{hours:02d}:{minutes:02d}:{secs:02d}"


def _verify_issue_resolved(issue: Mapping[str, str], race_files: Mapping[int, Path]) -> tuple[bool, str]:
    issue_code = str(issue.get("Issue Code", "")).strip()
    if issue_code != "AUD-ROW-002":
        return False, "Automatic verification is not implemented for this issue code."

    race_num = _parse_race_num(issue.get("Race", ""))
    source_row = _parse_row_num(issue.get("Source Row", ""))
    if race_num is None or source_row is None:
        return False, "Race or source row is missing from the audit issue."

    filepath = race_files.get(race_num)
    if filepath is None or not filepath.exists():
        return False, f"Race file for race {race_num} was not found."

    try:
        wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    except Exception as exc:
        return False, f"Could not open race file for verification: {exc}"

    try:
        ws = wb.active
        time_col = _find_time_column(ws)
        if time_col is None:
            return False, "Time column could not be identified for verification."
        if source_row > ws.max_row:
            return False, f"Source row {source_row} is outside the worksheet."
        value = ws.cell(row=source_row, column=time_col).value
        value_text = "" if value is None else str(value).strip()
        if not value_text or value_text.upper() == "QRY":
            return False, f"Source row {source_row} still has an unresolved time value."
        if parse_time_to_seconds(value_text) is None:
            return False, f"Source row {source_row} still has an invalid time value '{value_text}'."
        return True, f"Verified race {race_num} row {source_row} now has valid time '{value_text}'."
    finally:
        wb.close()


def _find_time_column(ws) -> int | None:
    headers = [
        str(cell.value).strip().lower() if cell.value is not None else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    for idx, header in enumerate(headers, start=1):
        if "chip" in header and "time" in header:
            return idx
    for idx, header in enumerate(headers, start=1):
        if header == "time":
            return idx
    for idx, header in enumerate(headers, start=1):
        if "time" in header:
            return idx
    return None


def _parse_race_num(value: object) -> int | None:
    text = str(value or "").strip()
    if not text:
        return None
    if text.isdigit():
        return int(text)
    return extract_race_number(text)


def _parse_row_num(value: object) -> int | None:
    text = str(value or "").strip()
    if not text:
        return None
    try:
        return int(float(text))
    except Exception:
        return None
