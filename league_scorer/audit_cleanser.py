"""Build a reviewed, non-destructive cleansed workbook from a raw race file."""

import logging
import re
from difflib import SequenceMatcher
from pathlib import Path
from typing import Dict, Iterable, Tuple

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill

from .name_lookup import load_name_corrections
from .normalisation import find_time_column, normalise_gender, parse_time_to_seconds, time_display
from .race_processor import extract_race_number
from .source_loader import load_race_dataframe

log = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="3A4658")
_ALT_FILL = PatternFill("solid", fgColor="F5F7FA")
_HEADER_FONT = Font(color="FFFFFF", bold=True)

_COLUMN_ALIASES = {
    "position": ("Position", "Pos", "Place"),
    "bib": ("Bib#", "Race No", "Bib", "Number"),
    "name": ("Name", "Runner", "Runner Name"),
    "club": ("Club", "Team", "Affiliation"),
    "gender": ("Gender", "Sex"),
    "category": ("Category", "Cat", "Age Category"),
    "gun_time": ("Gun Time",),
}

_NAME_REVIEW_EXPANSIONS = {
    "chris": "Christopher",
    "dave": "David",
    "phil": "Philip",
    "daf": "Daphne",
}


def create_cleansed_race_file(
    filepath: Path,
    raw_to_preferred: Dict[str, str],
    preferred_clubs: Iterable[str],
    audited_dir: Path,
    control_dir: Path,
    overwrite_existing: bool = False,
) -> Path:
    """Write a new audited workbook into the season audited folder."""
    source_df = load_race_dataframe(filepath)
    col_map = _resolve_columns(source_df.columns)

    time_col = find_time_column(list(source_df.columns))
    if time_col is None:
        raise ValueError(f"'{filepath.name}' has no time-like column")

    name_corrections = load_name_corrections(control_dir / "name_corrections.xlsx")

    use_excel_time_format = _is_race_over_5k(filepath.stem)

    race_df, club_df, name_df = _build_output_frames(
        source_df,
        col_map,
        time_col,
        raw_to_preferred,
        sorted(preferred_clubs),
        name_corrections,
        prefer_excel_time_format=use_excel_time_format,
    )
    race_sheet_name = _build_race_sheet_name(filepath)

    audited_dir.mkdir(parents=True, exist_ok=True)
    output_path = audited_dir / f"{_build_audited_stem(filepath.stem)} (audited).xlsx"
    if output_path.exists():
        if not overwrite_existing:
            raise FileExistsError(f"Audited file already exists: {output_path.name}")
        output_path.unlink()

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        race_df.to_excel(writer, index=False, sheet_name=race_sheet_name)
        club_df.to_excel(writer, index=False, sheet_name="Club Review")
        name_df.to_excel(writer, index=False, sheet_name="Name Review")

        _style_sheet(
            writer.sheets[race_sheet_name],
            race_df,
            excel_time_columns=("Time", "Gun Time") if use_excel_time_format else (),
        )
        _style_sheet(writer.sheets["Club Review"], club_df)
        _style_sheet(writer.sheets["Name Review"], name_df)

    log.info("Audited cleansed workbook written: %s", output_path)
    return output_path


def _build_output_frames(
    source_df: pd.DataFrame,
    col_map: Dict[str, str],
    time_col: str,
    raw_to_preferred: Dict[str, str],
    preferred_clubs: list[str],
    name_corrections: Dict[str, str],
    prefer_excel_time_format: bool = False,
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    race_rows = []
    club_suggestions: Dict[str, dict] = {}
    name_suggestions: Dict[str, dict] = {}

    for idx, row in source_df.iterrows():
        source_row = idx + 2
        comments = []

        raw_name = _clean_text(row.get(col_map.get("name", "")))
        lookup_name = name_corrections.get(raw_name.lower(), "") if raw_name else ""
        if lookup_name:
            clean_name = lookup_name
            if clean_name != raw_name:
                comments.append(f"Name normalised from '{raw_name}' to '{clean_name}' using reviewed name lookup")
        else:
            clean_name = _clean_name(raw_name)
        if not lookup_name and clean_name != raw_name and clean_name:
            comments.append(f"Name cleaned from '{raw_name}' to '{clean_name}'")

        suggested_name = ""
        suggested_name_confidence = ""
        if raw_name and not lookup_name:
            suggested_name, suggested_name_confidence, name_review_message = _suggest_name_review(raw_name, clean_name)
            if name_review_message:
                name_suggestions.setdefault(
                    raw_name,
                    {
                        "Raw Name": raw_name,
                        "Suggested Name": suggested_name,
                        "Confidence": suggested_name_confidence,
                        "Occurrences": 0,
                        "Status": "Manual Review",
                        "Message": name_review_message,
                    },
                )
                name_suggestions[raw_name]["Occurrences"] += 1
                comments.append(name_review_message)

        raw_club = _clean_text(row.get(col_map.get("club", "")))
        preferred_club = raw_to_preferred.get(raw_club.lower(), "") if raw_club else ""
        suggested_club = ""
        confidence = ""
        if preferred_club:
            if preferred_club != raw_club:
                comments.append(f"Club normalised from '{raw_club}' to '{preferred_club}'")
        elif raw_club:
            suggested_name, suggested_score = _best_club_match(raw_club, preferred_clubs)
            suggested_club = suggested_name
            confidence = suggested_score
            club_suggestions.setdefault(
                raw_club,
                {
                    "Raw Club": raw_club,
                    "Suggested Club": suggested_name,
                    "Confidence": suggested_score,
                    "Occurrences": 0,
                    "Status": "Manual Review",
                    "Message": "Review this club match in the approval dialog before updating Club Lookup.",
                },
            )
            club_suggestions[raw_club]["Occurrences"] += 1
            comments.append(
                f"Club not in lookup; suggested '{suggested_name}' ({suggested_score}%) awaiting manual review"
            )
        else:
            comments.append("Blank club retained for manual review")

        raw_gender = _clean_text(row.get(col_map.get("gender", "")))
        clean_gender = normalise_gender(raw_gender)
        if raw_gender and clean_gender is None:
            comments.append(f"Unresolved gender '{raw_gender}'")

        raw_category = _clean_text(row.get(col_map.get("category", "")))
        clean_category, category_note = _derive_clean_category(raw_category, clean_gender)
        if category_note:
            comments.append(category_note)

        # Wheelchair entries from non-eligible clubs are excluded from the audited output
        # and logged so operators can trace what was removed.
        if _is_wheelchair_category(raw_category) and not preferred_club:
            log.warning(
                "Dropped wheelchair row from audited output (non-eligible club): source_row=%s name='%s' club='%s' category='%s'",
                source_row,
                clean_name or raw_name,
                raw_club,
                raw_category,
            )
            continue

        chip_time_value = row.get(time_col)
        time_seconds = parse_time_to_seconds(chip_time_value)
        if _clean_text(chip_time_value):
            if prefer_excel_time_format and time_seconds is not None and time_seconds > 0:
                chip_time = _seconds_to_excel_time(time_seconds)
            else:
                chip_time = time_display(chip_time_value)
        else:
            chip_time = ""

        gun_time = ""
        if col_map.get("gun_time"):
            gun_time_value = row.get(col_map["gun_time"])
            if _clean_text(gun_time_value):
                gun_seconds = parse_time_to_seconds(gun_time_value)
                if prefer_excel_time_format and gun_seconds is not None and gun_seconds > 0:
                    gun_time = _seconds_to_excel_time(gun_seconds)
                else:
                    gun_time = time_display(gun_time_value)

        include = "Yes"
        status = "Ready"
        if not clean_name:
            include = "No"
            status = "Exclude"
            comments.append("Missing runner name")
        elif time_seconds is None:
            include = "QRY"
            status = "QRY"
            chip_time = "QRY"
            gun_time = "QRY"
            comments.append("Missing or invalid chip time retained as QRY")
        elif time_seconds <= 0:
            include = "QRY"
            status = "QRY"
            chip_time = "QRY"
            gun_time = "QRY"
            comments.append("Zero chip time retained as QRY pending manual correction")

        if status not in {"Exclude", "QRY"} and (
            clean_gender is None
            or (raw_category and not clean_category)
            or (raw_club and not preferred_club)
            or not raw_club
        ):
            status = "Review"
            if include == "Yes":
                include = "Yes"

        eligible = "Yes" if preferred_club and clean_gender and clean_category and include == "Yes" else "No"
        comment_text = " | ".join(dict.fromkeys(comments))

        race_rows.append(
            {
                "Pos": _clean_text(row.get(col_map.get("position", ""))),
                "Bib": _clean_text(row.get(col_map.get("bib", ""))),
                "Name": clean_name,
                "Club": preferred_club,
                "Time": chip_time,
                "Category": clean_category,
                "Gender": clean_gender or "",
                "Comments": comment_text,
            }
        )

    race_df = pd.DataFrame(
        race_rows,
        columns=["Pos", "Bib", "Name", "Club", "Time", "Category", "Gender", "Comments"],
    )
    club_df = pd.DataFrame(
        sorted(
            club_suggestions.values(),
            key=lambda item: (-item["Occurrences"], item["Raw Club"].lower()),
        ),
        columns=["Raw Club", "Suggested Club", "Confidence", "Occurrences", "Status", "Message"],
    )
    name_df = pd.DataFrame(
        sorted(
            name_suggestions.values(),
            key=lambda item: (-item["Occurrences"], item["Raw Name"].lower()),
        ),
        columns=["Raw Name", "Suggested Name", "Confidence", "Occurrences", "Status", "Message"],
    )
    return race_df, club_df, name_df


def _resolve_columns(columns: Iterable[str]) -> Dict[str, str]:
    lookup = {_normalise_column_name(column): str(column).strip() for column in columns}
    resolved = {}
    for key, aliases in _COLUMN_ALIASES.items():
        match = next(
            (
                lookup[_normalise_column_name(alias)]
                for alias in aliases
                if _normalise_column_name(alias) in lookup
            ),
            "",
        )
        if match:
            resolved[key] = match
    return resolved


def _build_race_sheet_name(filepath: Path) -> str:
    race_num = extract_race_number(filepath.stem)
    return f"Race {race_num}" if race_num is not None else "Race"


def _derive_clean_category(raw_category: str, clean_gender: str | None) -> Tuple[str, str]:
    category = _clean_text(raw_category)
    if not category:
        return "", "Blank category retained for review"

    compact = re.sub(r"\s+", " ", category).strip().lower()
    stripped_category, category_gender = _strip_category_gender(category)
    stripped_compact = re.sub(r"\s+", " ", stripped_category).strip().lower()
    notes = []

    if category_gender and clean_gender and category_gender != clean_gender:
        notes.append(
            f"Category sex prefix in '{category}' does not match row gender '{clean_gender}'"
        )
    elif category_gender and clean_gender:
        notes.append(f"Category cleaned from '{category}' to '{stripped_category}' after matching row gender")
    elif category_gender:
        notes.append(f"Category cleaned from '{category}' to '{stripped_category}' but row gender is unresolved")

    deterministic = {
        "ages 20 - 34": "Sen",
        "ages 20 - 39": "Sen",
        "ages 35 - 44": "V35",
        "ages 40 - 49": "V40",
        "ages 45 - 54": "V45",
        "ages 50 - 59": "V50",
        "ages 55 +": "V55",
        "ages 60 +": "V60",
        "unknown": "Sen",
        "os": "Sen",
        "fs": "Sen",
        "ms": "Sen",
        "sm": "Sen",
        "sl": "Sen",
        "top 3": "FIX",
        "top 3 male": "FIX",
        "top 3 female": "FIX",
        "pacer": "FIX",
        "oj": "Jun",
        "fj": "Jun",
        "under 17": "Jun",
        "under 20": "Jun",
    }
    if stripped_compact in deterministic:
        derived = deterministic[stripped_compact]
        notes.append(f"Category derived from '{category}' to '{derived}'")
        return derived, " | ".join(dict.fromkeys(notes))

    work = re.sub(r"[\s_\.]", "", stripped_compact)
    work = work.replace("vet", "v")
    if re.match(r"^(jun|junior|youth|cadet)", work):
        return "Jun", " | ".join(dict.fromkeys(notes))

    junior_match = re.match(r"u(\d+)$", work)
    if junior_match and int(junior_match.group(1)) <= 21:
        if stripped_category != "Jun":
            notes.append(f"Category cleaned from '{category}' to 'Jun'")
        return "Jun", " | ".join(dict.fromkeys(notes))

    if re.search(r"(sen|senior|open|elite|adult)", work):
        if stripped_category != "Sen":
            notes.append(f"Category cleaned from '{category}' to 'Sen'")
        return "Sen", " | ".join(dict.fromkeys(notes))

    age_match = re.search(r"v(\d{2})(\+)?", work)
    if not age_match:
        age_match = re.search(r"(\d{2})(\+)?", work)

    if age_match and "ages" not in stripped_compact:
        age = int(age_match.group(1))
        has_plus = bool(age_match.group(2))
        if age <= 21:
            notes.append(f"Category cleaned from '{category}' to 'Jun'")
            return "Jun", " | ".join(dict.fromkeys(notes))
        if age >= 35:
            derived = _format_clean_veteran_category(age, has_plus)
            if derived != stripped_category:
                notes.append(f"Category cleaned from '{category}' to '{derived}'")
            return derived, " | ".join(dict.fromkeys(notes))

    notes.append(f"Category '{category}' needs manual review")
    return "", " | ".join(dict.fromkeys(notes))


def _clean_name(name: str) -> str:
    if not name:
        return ""
    cleaned = name.replace("\ufffd", "'").replace("’", "'").replace("`", "'")
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    cleaned = re.sub(r"\s*-\s*", "-", cleaned)
    cleaned = re.sub(r"\s*\([^)]*\)\s*$", "", cleaned).strip()

    tokens = cleaned.split(" ")
    if len(tokens) >= 2:
        previous = re.sub(r"[^A-Za-z]", "", tokens[-2]).lower()
        last = re.sub(r"[^A-Za-z]", "", tokens[-1]).lower()
        if previous and previous == last and tokens[-1].upper() == tokens[-1]:
            cleaned = " ".join(tokens[:-1])

    return cleaned


def _suggest_name_review(raw_name: str, clean_name: str) -> Tuple[str, str, str]:
    text = _clean_text(raw_name)
    if not text:
        return "", "", ""

    tokens = text.split()
    if not tokens:
        return "", "", ""

    first_key = re.sub(r"[^A-Za-z]", "", tokens[0]).lower()
    if first_key in _NAME_REVIEW_EXPANSIONS:
        proposed_tokens = clean_name.split() if clean_name and clean_name != raw_name else list(tokens)
        if not proposed_tokens:
            proposed_tokens = [_NAME_REVIEW_EXPANSIONS[first_key]]
        else:
            proposed_tokens[0] = _NAME_REVIEW_EXPANSIONS[first_key]
        suggested = " ".join(proposed_tokens)
        return suggested, "45", f"Possible long-name expansion for '{tokens[0]}' requires manual review"

    if re.search(r"(.)\1{3,}", text):
        suggestion = clean_name if clean_name != raw_name else ""
        return suggestion, "25", "Suspicious repeated character sequence in name; review manually"

    return "", "", ""


def _clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).replace("\xa0", " ").strip()
    return "" if text.lower() == "nan" else text


def _normalise_column_name(value: str) -> str:
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _strip_category_gender(category: str) -> Tuple[str, str | None]:
    match = re.match(r"^([mf])(?:\s*[-_ ]*)?(.*)$", category.strip(), flags=re.IGNORECASE)
    if not match:
        return category, None

    remainder = match.group(2).strip()
    if not remainder:
        return category, None

    leading = remainder.lower()
    if not (
        leading.startswith(("v", "vet", "u", "jun", "sen", "senior"))
        or re.match(r"^\d", leading)
    ):
        return category, None

    return remainder, match.group(1).upper()


def _is_wheelchair_category(raw_category: str) -> bool:
    stripped, _ = _strip_category_gender(raw_category)
    compact = re.sub(r"[^a-z0-9]", "", stripped.lower())
    return compact in {"wheelchair", "wheelchairracer"}


def _format_clean_veteran_category(age: int, has_plus: bool) -> str:
    if age >= 70:
        return "V70+" if has_plus else "V70"
    if age >= 65:
        return "V65+" if has_plus else "V65"
    if age >= 60:
        return "V60"
    if age >= 55:
        return "V55"
    if age >= 50:
        return "V50"
    if age >= 45:
        return "V45"
    if age >= 40:
        return "V40"
    return "V35"


def _best_club_match(raw_club: str, preferred_clubs: Iterable[str]) -> Tuple[str, int]:
    raw = raw_club.lower().strip()
    best_name = ""
    best_score = 0
    for preferred in preferred_clubs:
        score = int(round(SequenceMatcher(None, raw, preferred.lower()).ratio() * 100))
        if score > best_score:
            best_name = preferred
            best_score = score
    return best_name, best_score


def _build_audited_stem(stem: str) -> str:
    match = re.match(r"race\s*#?\s*(\d+)\s*[-–]?\s*(.*)$", stem, flags=re.IGNORECASE)
    if not match:
        return stem

    race_num = int(match.group(1))
    suffix = match.group(2).strip(" -")
    return f"Race {race_num} - {suffix}" if suffix else f"Race {race_num}"


def _is_race_over_5k(stem: str) -> bool:
    text = stem.lower()

    km_match = re.search(r"(\d+(?:\.\d+)?)\s*k\b", text)
    if km_match:
        try:
            return float(km_match.group(1)) > 5.0
        except ValueError:
            pass

    mile_match = re.search(r"(\d+(?:\.\d+)?)\s*mile", text)
    if mile_match:
        try:
            miles = float(mile_match.group(1))
            return miles > 3.10686
        except ValueError:
            pass

    if "half marathon" in text or re.search(r"\bmarathon\b", text):
        return True

    return False


def _seconds_to_excel_time(seconds: float) -> float:
    return float(seconds) / 86400.0


def _style_sheet(ws, df: pd.DataFrame, excel_time_columns: Iterable[str] = ()) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
        use_alt_fill = row_idx % 2 == 1
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            if use_alt_fill:
                cell.fill = _ALT_FILL

    if excel_time_columns:
        time_col_indices = {
            idx
            for idx, name in enumerate(df.columns, start=1)
            if str(name) in set(excel_time_columns)
        }
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                if cell.column in time_col_indices and isinstance(cell.value, (int, float)):
                    cell.number_format = "h:mm:ss.0"

    for index, column_name in enumerate(df.columns, start=1):
        values = [str(column_name)]
        values.extend(str(value) for value in df.iloc[:, index - 1].head(200))
        width = min(max(len(value) for value in values) + 2, 36)
        ws.column_dimensions[ws.cell(row=1, column=index).column_letter].width = width