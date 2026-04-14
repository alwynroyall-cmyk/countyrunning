"""
events_loader.py - Load and parse the WRRL Championship Events spreadsheet.

Expected file: WRRL_events.xlsx
Expected sheet: "Championship Events"
Columns: RaceRef, EventName, Category, Distance, Location, Organiser,
         DateType, ScheduledDates, EligibilityWindow, EntryFee,
         ScoringBasis, Notes, Status
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional
import datetime

import openpyxl

# Status constants
STATUS_CONFIRMED   = "Confirmed"
STATUS_PROVISIONAL = "Provisional"
STATUS_TBC         = "TBC"

VALID_STATUSES = {STATUS_CONFIRMED, STATUS_PROVISIONAL, STATUS_TBC}

SHEET_NAME = "Championship Events"

# Column header -> attribute name mapping
_REQUIRED_COLUMN_MAP = {
    "RaceRef":           "race_ref",
    "EventName":         "event_name",
    "Category":          "category",
    "Distance":          "distance",
    "Location":          "location",
    "Organiser":         "organiser",
    "DateType":          "date_type",
    "ScheduledDates":    "scheduled_dates",
    "EligibilityWindow": "eligibility_window",
    "EntryFee":          "entry_fee",
    "ScoringBasis":      "scoring_basis",
    "Notes":             "notes",
    "Status":            "status",
}
_OPTIONAL_COLUMN_MAP = {
    "Website":           "website",
}
_COL_MAP = {**_REQUIRED_COLUMN_MAP, **_OPTIONAL_COLUMN_MAP}


@dataclass
class EventEntry:
    """One row from the Championship Events sheet."""
    race_ref:           str
    event_name:         str
    category:           str
    distance:           str
    location:           str
    organiser:          str
    date_type:          str
    scheduled_dates:    str          # normalised to human-readable string
    eligibility_window: str
    entry_fee:          str
    scoring_basis:      str
    notes:              str
    status:             str
    website:            str = ""

    # ------------------------------------------------------------------
    # Helpers
    # ------------------------------------------------------------------
    @property
    def is_confirmed(self) -> bool:
        return self.status == STATUS_CONFIRMED

    @property
    def is_provisional(self) -> bool:
        return self.status == STATUS_PROVISIONAL

    @property
    def is_tbc(self) -> bool:
        return self.status == STATUS_TBC


@dataclass
class EventsSchedule:
    """Container for all loaded events."""
    events: list[EventEntry] = field(default_factory=list)
    source_path: Optional[Path] = None

    # ------------------------------------------------------------------
    # Convenience accessors
    # ------------------------------------------------------------------
    def by_status(self, status: str) -> list[EventEntry]:
        return [e for e in self.events if e.status == status]

    @property
    def confirmed(self) -> list[EventEntry]:
        return self.by_status(STATUS_CONFIRMED)

    @property
    def provisional(self) -> list[EventEntry]:
        return self.by_status(STATUS_PROVISIONAL)

    @property
    def tbc(self) -> list[EventEntry]:
        return self.by_status(STATUS_TBC)

    def __len__(self) -> int:
        return len(self.events)


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def load_events(path: Path | str) -> EventsSchedule:
    """
    Load the Championship Events spreadsheet and return an EventsSchedule.

    Parameters
    ----------
    path:
        Full path to the WRRL_events.xlsx file.

    Returns
    -------
    EventsSchedule
        Populated schedule; empty if the sheet has no data rows.

    Raises
    ------
    FileNotFoundError
        If *path* does not exist.
    ValueError
        If the expected sheet is not found, or required columns are missing.
    """
    path = Path(path)
    if not path.exists():
        raise FileNotFoundError(f"Events file not found: {path}")

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            available = ", ".join(wb.sheetnames)
            raise ValueError(
                f"Sheet '{SHEET_NAME}' not found in {path.name}. "
                f"Available sheets: {available}"
            )

        ws = wb[SHEET_NAME]
        rows = list(ws.iter_rows(values_only=True))

        if not rows:
            return EventsSchedule(source_path=path)

        # ------------------------------------------------------------------ #
        # Parse header row
        # ------------------------------------------------------------------ #
        header = [str(h).strip() if h is not None else "" for h in rows[0]]
        header_lower = [h.lower() for h in header]
        col_index: dict[str, int] = {}
        for col_name, attr in _REQUIRED_COLUMN_MAP.items():
            try:
                col_index[attr] = header.index(col_name)
            except ValueError:
                raise ValueError(
                    f"Required column '{col_name}' not found in sheet '{SHEET_NAME}'. "
                    f"Found columns: {header}"
                )

        website_headers = [
            "website",
            "event website",
            "website url",
            "url",
            "event url",
            "link",
        ]
        for col_name, attr in _OPTIONAL_COLUMN_MAP.items():
            if col_name.lower() == "website":
                for header_name in website_headers:
                    if header_name in header_lower:
                        col_index[attr] = header_lower.index(header_name)
                        break
            else:
                lower_name = col_name.lower()
                if lower_name in header_lower:
                    col_index[attr] = header_lower.index(lower_name)

        # ------------------------------------------------------------------ #
        # Parse data rows
        # ------------------------------------------------------------------ #
        schedule = EventsSchedule(source_path=path)

        for row_num, row in enumerate(rows[1:], start=2):
            # Skip completely blank rows
            if all(cell is None for cell in row):
                continue

            def get(attr: str) -> str:
                index = col_index.get(attr)
                if index is None or index >= len(row):
                    return ""
                val = row[index]
                if val is None:
                    return ""
                if isinstance(val, (datetime.datetime, datetime.date)):
                    return f"{val.day} {val.strftime('%b %Y')}"
                if isinstance(val, (int, float)):
                    # Keep numerics as plain numbers (no currency prefix).
                    if isinstance(val, float) and val.is_integer():
                        return str(int(val))
                    return str(val)
                return str(val).strip()

            entry = EventEntry(
                race_ref           = get("race_ref"),
                event_name         = get("event_name"),
                category           = get("category"),
                distance           = get("distance"),
                location           = get("location"),
                organiser          = get("organiser"),
                date_type          = get("date_type"),
                scheduled_dates    = get("scheduled_dates"),
                eligibility_window = get("eligibility_window"),
                entry_fee          = get("entry_fee"),
                scoring_basis      = get("scoring_basis"),
                notes              = get("notes"),
                status             = get("status") or STATUS_TBC,
                website            = get("website"),
            )
            schedule.events.append(entry)

        return schedule
    finally:
        wb.close()
