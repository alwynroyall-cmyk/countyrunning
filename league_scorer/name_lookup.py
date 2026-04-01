"""Load and update persisted reviewed name corrections."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Dict, Iterable

from openpyxl import Workbook, load_workbook

log = logging.getLogger(__name__)

_HEADERS = ("Raw Name", "Preferred Name")


def load_name_corrections(filepath: Path) -> Dict[str, str]:
    """Return lower-cased raw-name mappings to reviewed preferred names."""
    if not filepath.exists():
        return {}

    try:
        workbook = load_workbook(filepath)
    except Exception as exc:
        log.warning("Unable to read name corrections file '%s': %s", filepath, exc)
        return {}

    worksheet = workbook.active
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[1]]
    if not set(_HEADERS).issubset(headers):
        log.warning("Name corrections file '%s' is missing required headers", filepath)
        return {}

    header_map = {name: idx for idx, name in enumerate(headers)}
    mappings: Dict[str, str] = {}
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        raw = str(row[header_map["Raw Name"]]).strip() if row[header_map["Raw Name"]] is not None else ""
        preferred = str(row[header_map["Preferred Name"]]).strip() if row[header_map["Preferred Name"]] is not None else ""
        if not raw or not preferred:
            continue
        mappings[raw.lower()] = preferred
    return mappings


def read_name_lookup_state(filepath: Path) -> dict:
    return {"alias_to_preferred": load_name_corrections(filepath)}


def append_name_corrections(filepath: Path, selected: Iterable[dict]) -> dict:
    workbook, worksheet = _open_or_create_workbook(filepath)
    lookup_state = read_name_lookup_state(filepath)
    alias_to_preferred = lookup_state["alias_to_preferred"]

    written = 0
    skipped_existing = 0
    skipped_conflicts = 0

    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in worksheet[1]]
    header_map = {name: idx + 1 for idx, name in enumerate(headers)}

    for item in selected:
        raw_name = str(item.get("current_name", "")).strip()
        preferred_name = str(item.get("proposed_name", "")).strip()
        if not raw_name or not preferred_name:
            skipped_conflicts += 1
            continue

        existing = alias_to_preferred.get(raw_name.lower())
        if existing:
            if existing == preferred_name:
                skipped_existing += 1
            else:
                skipped_conflicts += 1
            continue

        next_row = worksheet.max_row + 1
        worksheet.cell(next_row, header_map["Raw Name"], raw_name)
        worksheet.cell(next_row, header_map["Preferred Name"], preferred_name)
        alias_to_preferred[raw_name.lower()] = preferred_name
        written += 1

    workbook.save(filepath)
    return {
        "written": written,
        "skipped_existing": skipped_existing,
        "skipped_conflicts": skipped_conflicts,
        "path": filepath,
    }


def _open_or_create_workbook(filepath: Path):
    if filepath.exists():
        workbook = load_workbook(filepath)
        worksheet = workbook.active
        return workbook, worksheet

    filepath.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Name Corrections"
    for index, header in enumerate(_HEADERS, start=1):
        worksheet.cell(1, index, header)
    return workbook, worksheet