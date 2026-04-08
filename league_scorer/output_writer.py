"""
Write output Excel files for publish/review packs.
"""

import logging
import re
from collections import defaultdict
from pathlib import Path
from typing import Dict, List

from .models import (
    RaceIssue,
    RunnerRaceEntry,
    RunnerSeasonRecord,
    TeamRaceResult,
    TeamSeasonRecord,
    UnrecognisedClub,
)

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

from .rules import get_max_races
_HEADER_FILL = PatternFill("solid", fgColor="4472C4")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_AGG_FILL    = PatternFill("solid", fgColor="D9E1F2")   # light blue tint for aggregate columns
_ALT_ROW_FILL = PatternFill("solid", fgColor="EEF3F8")

# Characters that trigger formula execution in spreadsheet clients
_FORMULA_PREFIXES = frozenset(("=", "+", "-", "@"))


def _sanitise_df_for_export(df: pd.DataFrame) -> pd.DataFrame:
    """Prefix formula-start characters in text cells to prevent spreadsheet injection."""
    df = df.copy()
    for col in df.columns:
        if df[col].dtype == object:
            df[col] = df[col].map(
                lambda v: ("'" + v) if isinstance(v, str) and v and v[0] in _FORMULA_PREFIXES else v
            )
    return df


# ──────────────────────────────────────────────────────────────── helpers ────

def _style_and_width(ws, df: pd.DataFrame, excel_time_columns: tuple[str, ...] = ()) -> None:
    """Bold blue header + approximate column widths + aggregate column shading."""
    # Style the header row
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center")

    # Build column-index sets once, then traverse data rows in a single pass.
    agg_col_indices = {
        col_idx
        for col_idx, col_name in enumerate(df.columns, start=1)
        if "Aggregate" in str(col_name)
    }
    issue_col_indices = {
        col_idx
        for col_idx, col_name in enumerate(df.columns, start=1)
        if str(col_name) in {"Warnings", "Other Issues"}
    }
    time_col_indices = {
        col_idx
        for col_idx, col_name in enumerate(df.columns, start=1)
        if str(col_name) in set(excel_time_columns)
    } if excel_time_columns else set()
    is_race_summary = ws.title == "Race Summary"
    _wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
        use_alt = is_race_summary and row_idx % 2 == 1
        for cell in row:
            col = cell.column
            if col in agg_col_indices:
                cell.fill = _AGG_FILL
            if is_race_summary:
                cell.alignment = _wrap
                if use_alt:
                    cell.fill = _ALT_ROW_FILL
            if col in issue_col_indices:
                cell.alignment = _wrap
            if col in time_col_indices and isinstance(cell.value, (int, float)):
                cell.number_format = "h:mm:ss.0"

    # Auto-width
    for col_idx, col_name in enumerate(df.columns, start=1):
        values = [str(col_name)] + [
            str(v) for v in df.iloc[:, col_idx - 1] if str(v) != ""
        ]
        max_width = 80 if str(col_name) in {"Warnings", "Other Issues"} else 50
        width = min(max((len(v) for v in values), default=8) + 2, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_df(df: pd.DataFrame, filepath: Path, sheet_name: str = "Data") -> None:
    try:
        df = _sanitise_df_for_export(df)
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name=sheet_name)
            _style_and_width(writer.sheets[sheet_name], df)
        log.debug("Written: %s", filepath.name)
    except Exception as exc:
        msg = (
            f"Failed to write output file '{filepath}'. "
            "Please close the workbook if it is open and try again. "
            f"Original error: {exc}"
        )
        log.error(msg)
        raise RuntimeError(msg) from exc


def _is_race_over_5k_name(name: str) -> bool:
    text = str(name).lower()

    km_match = re.search(r"(\d+(?:\.\d+)?)\s*k\b", text)
    if km_match:
        try:
            return float(km_match.group(1)) > 5.0
        except ValueError:
            pass

    mile_match = re.search(r"(\d+(?:\.\d+)?)\s*mile", text)
    if mile_match:
        try:
            miles = float(mile_match.group(1))
            return miles > 3.10686
        except ValueError:
            pass

    if "half marathon" in text or re.search(r"\bmarathon\b", text):
        return True

    return False


def _seconds_to_excel_time(seconds: float) -> float:
    return float(seconds) / 86400.0


# ─────────────────────────────────────────────────────── individual table ────

def write_results_workbook(
    highest_race: int,
    male_records: List[RunnerSeasonRecord],
    female_records: List[RunnerSeasonRecord],
    div1_teams: List[TeamSeasonRecord],
    div2_teams: List[TeamSeasonRecord],
    all_race_runners: Dict[int, List[RunnerRaceEntry]],
    race_files: Dict[int, Path],
    all_unrec_clubs: Dict[int, List[UnrecognisedClub]],
    race_issues: Dict[int, List[RaceIssue]],
    filepath: Path,
) -> None:
    """Write Race Summary, Div 1, Div 2, Male, Female, Category Review, and Race N sheets."""
    max_races = get_max_races()

    def _format_issue(issue: RaceIssue) -> str:
        if issue.source_row is not None:
            return f"Row {issue.source_row}: {issue.message}"
        return issue.message

    def _summarise_issue_bucket(messages: List[str]) -> str:
        if not messages:
            return ""
        return f"Count: {len(messages)}\n" + "\n".join(messages)

    # ── Individual table builder ─────────────────────────────────────────────
    def _ind_df(records: List[RunnerSeasonRecord]) -> pd.DataFrame:
        rows = []
        for rec in sorted(records, key=lambda r: (r.position, r.name)):
            row: dict = {
                "Position": rec.position,
                "Name":     rec.name,
                "Club":     rec.preferred_club,
                "Category": rec.category,
            }
            for n in range(1, max_races + 1):
                row[f"Race {n} Points"] = rec.race_points.get(n, "")
            row["Total Points"] = rec.total_points
            rows.append(row)
        return pd.DataFrame(rows)

    # ── Club table builder ───────────────────────────────────────────────────
    def _club_df(teams: List[TeamSeasonRecord]) -> pd.DataFrame:
        rows = []
        for team in sorted(teams, key=lambda t: (t.position, t.preferred_club)):
            row: dict = {
                "Position": team.position,
                "Club":     team.display_name,
            }
            for n in range(1, max_races + 1):
                rr = team.race_results.get(n)
                row[f"Race {n} Men Score"]   = rr.men_score   if rr else ""
                row[f"Race {n} Women Score"] = rr.women_score if rr else ""
                row[f"Race {n} Aggregate"]   = (
                    (rr.men_score or 0) + (rr.women_score or 0) if rr else ""
                )
                row[f"Race {n} Team Points"] = rr.team_points if rr else ""
            row["Total Points"] = team.total_points
            rows.append(row)
        return pd.DataFrame(rows)

    def _race_df(race_num: int, runners: List[RunnerRaceEntry], use_excel_time_format: bool) -> pd.DataFrame:
        rows = []
        sorted_runners = sorted(runners, key=lambda r: (r.time_seconds, r.name.lower()))
        gender_positions = {"M": 0, "F": 0}

        for overall_position, runner in enumerate(sorted_runners, start=1):
            gender_positions[runner.gender] = gender_positions.get(runner.gender, 0) + 1
            rows.append(
                {
                    "Overall Pos": overall_position,
                    "Gender Pos": gender_positions[runner.gender],
                    "Name": runner.name,
                    "Raw Club": runner.raw_club,
                    "Club": runner.preferred_club or "",
                    "Gender": runner.gender,
                    "Raw Category": runner.raw_category,
                    "Category": runner.normalised_category,
                    "Time": (
                        _seconds_to_excel_time(runner.time_seconds)
                        if use_excel_time_format and runner.time_seconds is not None and runner.time_seconds > 0
                        else runner.time_str
                    ),
                    "Wiltshire Eligible": "Yes" if runner.eligible else "No",
                    "Points": runner.points if runner.points > 0 else "",
                    "Team": runner.team_id,
                    "Warnings": "\n".join(runner.warnings),
                }
            )

        return pd.DataFrame(
            rows,
            columns=[
                "Overall Pos",
                "Gender Pos",
                "Name",
                "Raw Club",
                "Club",
                "Gender",
                "Raw Category",
                "Category",
                "Time",
                "Wiltshire Eligible",
                "Points",
                "Team",
                "Warnings",
            ],
        )

    def _race_summary_df() -> pd.DataFrame:
        rows = []
        for race_num in sorted(race_files):
            issues = race_issues.get(race_num, [])
            warnings = [_format_issue(issue) for issue in issues if issue.kind == "warning"]
            other_issues = [_format_issue(issue) for issue in issues if issue.kind != "warning"]
            runners = all_race_runners.get(race_num, [])
            unrec = all_unrec_clubs.get(race_num, [])
            rows.append(
                {
                    "Race": race_num,
                    "Source File": race_files[race_num].name,
                    "Status": "Processed" if race_num in all_race_runners else "Skipped",
                    "Runner Rows": len(runners),
                    "Unrecognised Clubs": len(unrec),
                    "Warnings": _summarise_issue_bucket(warnings),
                    "Other Issues": _summarise_issue_bucket(other_issues),
                }
            )
        return pd.DataFrame(
            rows,
            columns=[
                "Race",
                "Source File",
                "Status",
                "Runner Rows",
                "Unrecognised Clubs",
                "Warnings",
                "Other Issues",
            ],
        )

    def _is_unknown_raw_category(raw_category: str) -> bool:
        compact = "".join(ch for ch in str(raw_category).strip().lower() if ch.isalnum())
        return compact in {"", "unknown", "qry", "na", "nacat", "uncategorised", "uncategorized"}

    def _category_review_df() -> pd.DataFrame:
        grouped: Dict[tuple, List[RunnerRaceEntry]] = defaultdict(list)
        for race_num, runners in all_race_runners.items():
            for runner in runners:
                key = (
                    runner.name.strip().lower(),
                    (runner.preferred_club or runner.raw_club or "").strip().lower(),
                    runner.gender,
                )
                grouped[key].append(runner)

        rows = []
        for grouped_rows in grouped.values():
            unknown_rows = [r for r in grouped_rows if _is_unknown_raw_category(r.raw_category)]
            if not unknown_rows:
                continue

            sample = sorted(grouped_rows, key=lambda r: r.race_number)[0]
            known_rows = [r for r in grouped_rows if not _is_unknown_raw_category(r.raw_category)]

            unknown_races = ", ".join(str(r.race_number) for r in sorted(unknown_rows, key=lambda r: r.race_number))
            unknown_values = ", ".join(
                sorted({str(r.raw_category).strip() or "(blank)" for r in unknown_rows})
            )

            if known_rows:
                evidence_parts = []
                for race_row in sorted(known_rows, key=lambda r: r.race_number):
                    evidence_parts.append(
                        f"R{race_row.race_number}: raw '{race_row.raw_category}' -> {race_row.normalised_category}"
                    )
                suggested = sorted({r.normalised_category for r in known_rows if r.normalised_category})
                suggested_text = ", ".join(suggested)
                status = "Review Suggested"
                next_step = "Check race sheets listed in Evidence and confirm category for unknown races."
            else:
                evidence_parts = []
                suggested_text = ""
                status = "No Evidence"
                next_step = "No known category found in other races yet; leave unresolved until more results exist."

            rows.append(
                {
                    "Name": sample.name,
                    "Club": sample.preferred_club or sample.raw_club,
                    "Gender": sample.gender,
                    "Unknown Races": unknown_races,
                    "Unknown Source Categories": unknown_values,
                    "Suggested Category": suggested_text,
                    "Evidence": "\n".join(evidence_parts),
                    "Status": status,
                    "Next Step": next_step,
                }
            )

        return pd.DataFrame(
            rows,
            columns=[
                "Name",
                "Club",
                "Gender",
                "Unknown Races",
                "Unknown Source Categories",
                "Suggested Category",
                "Evidence",
                "Status",
                "Next Step",
            ],
        )

    try:
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            for df, sheet in [
                (_race_summary_df(),      "Race Summary"),
                (_club_df(div1_teams),    "Div 1"),
                (_club_df(div2_teams),    "Div 2"),
                (_ind_df(male_records),   "Male"),
                (_ind_df(female_records), "Female"),
                (_category_review_df(),   "Category Review"),
            ]:
                df.to_excel(writer, index=False, sheet_name=sheet)
                _style_and_width(writer.sheets[sheet], df)

            for race_num in sorted(all_race_runners):
                race_name = race_files.get(race_num).stem if race_num in race_files else f"Race {race_num}"
                is_long_race = _is_race_over_5k_name(race_name)
                race_df = _race_df(race_num, all_race_runners[race_num], is_long_race)
                sheet_name = f"Race {race_num}"
                race_df.to_excel(writer, index=False, sheet_name=sheet_name)
                _style_and_width(
                    writer.sheets[sheet_name],
                    race_df,
                    excel_time_columns=("Time",) if is_long_race else (),
                )
        log.debug("Written: %s", filepath.name)
    except Exception as exc:
        msg = (
            f"Failed to write results workbook '{filepath}'. "
            "Please close the workbook if it is open and try again. "
            f"Original error: {exc}"
        )
        log.error(msg)
        raise RuntimeError(msg) from exc


def write_unrecognised_clubs(
    unrec_clubs: List[UnrecognisedClub],
    filepath: Path,
) -> None:
    """Spec 13.2 Race # -- unused clubs.xlsx"""
    rows = [
        {
            "Raw Club Name": u.raw_club_name,
            "Occurrences":   u.occurrences,
            "Action":        "Excluded",
        }
        for u in sorted(unrec_clubs, key=lambda u: u.raw_club_name)
    ]
    df = pd.DataFrame(rows) if rows else pd.DataFrame(
        columns=["Raw Club Name", "Occurrences", "Action"]
    )
    _write_df(df, filepath, sheet_name="Unused Clubs")


def build_category_mismatch_todo_df(
    all_race_runners: Dict[int, List[RunnerRaceEntry]],
) -> pd.DataFrame:
    """Build a TODO list for cross-race category/club mismatches."""
    grouped: Dict[tuple, List[RunnerRaceEntry]] = defaultdict(list)
    for race_num in sorted(all_race_runners):
        for runner in all_race_runners[race_num]:
            if not runner.eligible or runner.preferred_club is None:
                continue
            key = (
                runner.name.strip().lower(),
                runner.preferred_club.strip().lower(),
                runner.gender,
            )
            grouped[key].append(runner)

    rows = []
    for grouped_rows in grouped.values():
        if len(grouped_rows) < 2:
            continue

        ordered = sorted(grouped_rows, key=lambda r: (r.race_number, r.source_row))
        categories_seen = [r.normalised_category for r in ordered if r.normalised_category]
        distinct_categories = list(dict.fromkeys(categories_seen))
        if len(set(distinct_categories)) <= 1:
            continue

        counts = defaultdict(int)
        first_seen = {}
        for idx, race_row in enumerate(ordered):
            cat = race_row.normalised_category
            if not cat:
                continue
            counts[cat] += 1
            first_seen.setdefault(cat, idx)

        suggested = ""
        if counts:
            suggested = sorted(
                counts,
                key=lambda cat: (-counts[cat], first_seen.get(cat, 10_000), cat),
            )[0]

        sample = ordered[0]
        race_categories = "; ".join(
            f"R{r.race_number}:{r.normalised_category or '(blank)'}" for r in ordered
        )
        source_values = "; ".join(
            f"R{r.race_number}:{(r.raw_category or '').strip() or '(blank)'}" for r in ordered
        )

        rows.append(
            {
                "Issue Type": "Category Mismatch",
                "Name": sample.name,
                "Club": sample.preferred_club,
                "Gender": sample.gender,
                "Races Completed": len(ordered),
                "Categories Seen": ", ".join(sorted(set(distinct_categories))),
                "Clubs Seen": sample.preferred_club,
                "Suggested Category": suggested,
                "Suggested Club": sample.preferred_club,
                "Race Category Sequence": race_categories,
                "Race Club Sequence": "; ".join(
                    f"R{r.race_number}:{r.preferred_club or r.raw_club or '(blank)'}" for r in ordered
                ),
                "Source Category Values": source_values,
                "Source Club Values": "; ".join(
                    f"R{r.race_number}:{(r.raw_club or '').strip() or '(blank)'}" for r in ordered
                ),
                "Next Step": "Review race rows and apply a consistent category across this runner's results.",
            }
        )

    fix_grouped: Dict[tuple, List[RunnerRaceEntry]] = defaultdict(list)
    for race_num in sorted(all_race_runners):
        for runner in all_race_runners[race_num]:
            if not runner.eligible or runner.preferred_club is None:
                continue
            if str(runner.normalised_category).strip().upper() != "FIX":
                continue
            key = (
                runner.name.strip().lower(),
                runner.preferred_club.strip().lower(),
                runner.gender,
            )
            fix_grouped[key].append(runner)

    for grouped_rows in fix_grouped.values():
        ordered = sorted(grouped_rows, key=lambda r: (r.race_number, r.source_row))
        sample = ordered[0]
        source_labels = sorted({(r.raw_category or "").strip() or "(blank)" for r in ordered})

        rows.append(
            {
                "Issue Type": "Category FIX",
                "Name": sample.name,
                "Club": sample.preferred_club,
                "Gender": sample.gender,
                "Races Completed": len(ordered),
                "Categories Seen": "FIX",
                "Clubs Seen": sample.preferred_club,
                "Suggested Category": "",
                "Suggested Club": sample.preferred_club,
                "Race Category Sequence": "; ".join(
                    f"R{r.race_number}:{r.normalised_category or '(blank)'}" for r in ordered
                ),
                "Race Club Sequence": "; ".join(
                    f"R{r.race_number}:{r.preferred_club or r.raw_club or '(blank)'}" for r in ordered
                ),
                "Source Category Values": ", ".join(source_labels),
                "Source Club Values": "; ".join(
                    f"R{r.race_number}:{(r.raw_club or '').strip() or '(blank)'}" for r in ordered
                ),
                "Next Step": "Manual correction required for FIX category rows (for example Top 3/Pacer labels).",
            }
        )

    by_name_gender: Dict[tuple, List[RunnerRaceEntry]] = defaultdict(list)
    for race_num in sorted(all_race_runners):
        for runner in all_race_runners[race_num]:
            key = (runner.name.strip().lower(), runner.gender)
            by_name_gender[key].append(runner)

    for grouped_rows in by_name_gender.values():
        if len(grouped_rows) < 2:
            continue

        ordered = sorted(grouped_rows, key=lambda r: (r.race_number, r.source_row))
        clubs_seen = []
        for race_row in ordered:
            club_value = (race_row.preferred_club or race_row.raw_club or "").strip()
            if club_value:
                clubs_seen.append(club_value)

        distinct_clubs = sorted(set(clubs_seen))
        eligible_clubs = sorted(
            {
                race_row.preferred_club
                for race_row in ordered
                if race_row.preferred_club
            }
        )

        if len(distinct_clubs) <= 1 or not eligible_clubs:
            continue

        preferred_counts = defaultdict(int)
        for race_row in ordered:
            if race_row.preferred_club:
                preferred_counts[race_row.preferred_club] += 1

        suggested_club = ""
        if preferred_counts:
            suggested_club = sorted(
                preferred_counts,
                key=lambda club: (-preferred_counts[club], club),
            )[0]

        sample = ordered[0]
        categories_seen = sorted(
            {
                race_row.normalised_category
                for race_row in ordered
                if race_row.normalised_category
            }
        )

        rows.append(
            {
                "Issue Type": "Club Mismatch",
                "Name": sample.name,
                "Club": suggested_club or " / ".join(eligible_clubs),
                "Gender": sample.gender,
                "Races Completed": len(ordered),
                "Categories Seen": ", ".join(categories_seen),
                "Clubs Seen": ", ".join(distinct_clubs),
                "Suggested Category": "",
                "Suggested Club": suggested_club,
                "Race Category Sequence": "; ".join(
                    f"R{r.race_number}:{r.normalised_category or '(blank)'}" for r in ordered
                ),
                "Race Club Sequence": "; ".join(
                    f"R{r.race_number}:{r.preferred_club or r.raw_club or '(blank)'}" for r in ordered
                ),
                "Source Category Values": "; ".join(
                    f"R{r.race_number}:{(r.raw_category or '').strip() or '(blank)'}" for r in ordered
                ),
                "Source Club Values": "; ".join(
                    f"R{r.race_number}:{(r.raw_club or '').strip() or '(blank)'}" for r in ordered
                ),
                "Next Step": "Review club history and set one consistent league club for this runner.",
            }
        )

    df = pd.DataFrame(
        rows,
        columns=[
            "Issue Type",
            "Name",
            "Club",
            "Gender",
            "Races Completed",
            "Categories Seen",
            "Clubs Seen",
            "Suggested Category",
            "Suggested Club",
            "Race Category Sequence",
            "Race Club Sequence",
            "Source Category Values",
            "Source Club Values",
            "Next Step",
        ],
    )
    if not df.empty:
        df = df.sort_values(by=["Issue Type", "Club", "Name"], ignore_index=True)
    return df


def write_category_mismatch_todo(
    all_race_runners: Dict[int, List[RunnerRaceEntry]],
    filepath: Path,
) -> None:
    """Write cross-race category mismatch TODO list for eligible-club runners."""
    df = build_category_mismatch_todo_df(all_race_runners)
    if df.empty:
        df = pd.DataFrame(
            columns=[
                "Issue Type",
                "Name",
                "Club",
                "Gender",
                "Races Completed",
                "Categories Seen",
                "Clubs Seen",
                "Suggested Category",
                "Suggested Club",
                "Race Category Sequence",
                "Race Club Sequence",
                "Source Category Values",
                "Source Club Values",
                "Next Step",
            ]
        )

    _write_df(df, filepath, sheet_name="Category Mismatch TODO")


def build_time_qry_todo_df(
    all_race_runners: Dict[int, List[RunnerRaceEntry]],
) -> pd.DataFrame:
    """Build a TODO list of rows with unresolved time values (QRY/invalid/blank)."""
    rows = []
    for race_num in sorted(all_race_runners):
        race_rows = sorted(
            all_race_runners[race_num],
            key=lambda r: (r.source_row if r.source_row is not None else 10_000_000, r.name.lower()),
        )
        for runner in race_rows:
            raw_time = str(runner.time_str or "").strip()
            is_qry = raw_time.upper() == "QRY"
            is_invalid_or_blank = (not raw_time) or (runner.time_seconds is None) or (runner.time_seconds <= 0)
            if not (is_qry or is_invalid_or_blank):
                continue

            rows.append(
                {
                    "Issue Type": "Time QRY",
                    "Race": runner.race_number,
                    "Source Row": runner.source_row or "",
                    "Name": runner.name,
                    "Raw Club": runner.raw_club,
                    "Club": runner.preferred_club or "",
                    "Gender": runner.gender,
                    "Category": runner.normalised_category,
                    "Current Time": raw_time or "(blank)",
                    "Status": "QRY" if is_qry else "Invalid",
                    "Next Step": "Use Runner History > Fix QRY Time to apply hh:mm:ss across input rows.",
                }
            )

    df = pd.DataFrame(
        rows,
        columns=[
            "Issue Type",
            "Race",
            "Source Row",
            "Name",
            "Raw Club",
            "Club",
            "Gender",
            "Category",
            "Current Time",
            "Status",
            "Next Step",
        ],
    )
    if not df.empty:
        df = df.sort_values(by=["Race", "Source Row", "Name"], ignore_index=True)
    return df


def write_time_qry_todo(
    all_race_runners: Dict[int, List[RunnerRaceEntry]],
    filepath: Path,
) -> None:
    """Write unresolved time TODO list for manual correction."""
    df = build_time_qry_todo_df(all_race_runners)
    if df.empty:
        df = pd.DataFrame(
            columns=[
                "Issue Type",
                "Race",
                "Source Row",
                "Name",
                "Raw Club",
                "Club",
                "Gender",
                "Category",
                "Current Time",
                "Status",
                "Next Step",
            ]
        )

    _write_df(df, filepath, sheet_name="Time QRY TODO")
