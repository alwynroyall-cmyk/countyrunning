"""Service-layer helpers for bulk manual workbook edits."""

from __future__ import annotations

import logging
from pathlib import Path

import openpyxl

from .structured_logging import log_event

log = logging.getLogger(__name__)


def apply_club_suggestions(updates_by_file: dict[Path, list[dict]]) -> tuple[int, list[dict], list[str]]:
    """Apply grouped club suggestions to input workbooks.

    Returns (applied_count, audit_changes, failed_messages).
    """
    applied = 0
    audit_changes: list[dict] = []
    failed: list[str] = []
    total_updates = sum(len(items) for items in updates_by_file.values())
    log_event(
        "manual_bulk_club_update_started",
        logger=log,
        file_count=len(updates_by_file),
        candidate_updates=total_updates,
    )

    for filepath, updates in updates_by_file.items():
        if not _is_raw_data_file(filepath):
            failed.append(f"{filepath.name}: edits are only allowed in inputs/raw_data")
            continue
        try:
            wb = openpyxl.load_workbook(filepath)
        except Exception as exc:
            failed.append(f"{filepath.name}: {exc}")
            log_event(
                "manual_bulk_club_update_open_failed",
                level="WARNING",
                logger=log,
                file_path=filepath,
                error=str(exc),
            )
            continue

        try:
            ws = wb.active
            file_applied = 0
            for update in updates:
                cell = ws.cell(row=update["row_idx"], column=update["club_col"])
                old_value = "" if cell.value is None else str(cell.value).strip()
                new_value = update["suggested_club"]
                if old_value == new_value:
                    continue

                cell.value = new_value
                applied += 1
                file_applied += 1
                audit_changes.append(
                    {
                        "runner": update["name"],
                        "field": "club",
                        "old_value": old_value,
                        "new_value": new_value,
                        "file_path": filepath,
                        "row_idx": update["row_idx"],
                    }
                )
            wb.save(filepath)
            log_event(
                "manual_bulk_club_update_file_saved",
                logger=log,
                file_path=filepath,
                applied=file_applied,
            )
        except Exception as exc:
            failed.append(f"{filepath.name}: {exc}")
            log_event(
                "manual_bulk_club_update_file_failed",
                level="ERROR",
                logger=log,
                file_path=filepath,
                error=str(exc),
            )
        finally:
            wb.close()

    log_event(
        "manual_bulk_club_update_completed",
        logger=log,
        applied=applied,
        audit_rows=len(audit_changes),
        failed_files=len(failed),
    )
    return applied, audit_changes, failed


def resolve_runner_field_across_files(
    race_files: dict[int, Path],
    *,
    selected_runner: str,
    field_type: str,
    target_value: str,
) -> tuple[int, int, list[dict], list[str]]:
    """Apply one runner field value across race input files.

    Returns (updated_rows, touched_files, audit_changes, failed_files).
    """
    runner_key = selected_runner.lower()
    updated_rows = 0
    touched_files = 0
    audit_changes: list[dict] = []
    failed_files: list[str] = []
    log_event(
        "manual_runner_field_resolve_started",
        logger=log,
        runner=selected_runner,
        field_type=field_type,
        target_value=target_value,
        file_count=len(race_files),
    )

    for _, path in race_files.items():
        if not _is_raw_data_file(path):
            failed_files.append(f"{path.name}: edits are only allowed in inputs/raw_data")
            continue
        try:
            wb = openpyxl.load_workbook(path)
        except Exception as exc:
            failed_files.append(f"{path.name}: {exc}")
            log_event(
                "manual_runner_field_resolve_open_failed",
                level="WARNING",
                logger=log,
                runner=selected_runner,
                file_path=path,
                error=str(exc),
            )
            continue

        try:
            ws = wb.active
            name_col, field_col = _find_columns(ws, field_type)
            if name_col is None or field_col is None:
                continue

            file_changed = False
            for row_idx in range(2, ws.max_row + 1):
                row_name = _row_name_value(ws, row_idx, name_col)
                if row_name.lower() != runner_key:
                    continue

                current = ws.cell(row=row_idx, column=field_col).value
                current_text = "" if current is None else str(current).strip()
                if current_text == target_value:
                    continue

                ws.cell(row=row_idx, column=field_col).value = target_value
                updated_rows += 1
                file_changed = True
                audit_changes.append(
                    {
                        "runner": selected_runner,
                        "field": field_type,
                        "old_value": current_text,
                        "new_value": target_value,
                        "file_path": path,
                        "row_idx": row_idx,
                    }
                )

            if file_changed:
                wb.save(path)
                touched_files += 1
                log_event(
                    "manual_runner_field_resolve_file_saved",
                    logger=log,
                    runner=selected_runner,
                    field_type=field_type,
                    file_path=path,
                )
        except Exception as exc:
            failed_files.append(f"{path.name}: {exc}")
            log_event(
                "manual_runner_field_resolve_file_failed",
                level="ERROR",
                logger=log,
                runner=selected_runner,
                field_type=field_type,
                file_path=path,
                error=str(exc),
            )
        finally:
            wb.close()

    log_event(
        "manual_runner_field_resolve_completed",
        logger=log,
        runner=selected_runner,
        field_type=field_type,
        updated_rows=updated_rows,
        touched_files=touched_files,
        failed_files=len(failed_files),
    )
    return updated_rows, touched_files, audit_changes, failed_files


def _find_columns(ws, field_type: str):
    headers = [
        str(c.value).strip().lower() if c.value is not None else ""
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    name_col = next(
        (i + 1 for i, h in enumerate(headers) if "name" in h and "first" not in h and "last" not in h),
        None,
    )
    if name_col is None:
        first_col = next((i + 1 for i, h in enumerate(headers) if "first" in h), None)
        last_col = next((i + 1 for i, h in enumerate(headers) if "last" in h), None)
        if first_col is not None and last_col is not None:
            name_col = (first_col, last_col)

    if field_type == "club":
        field_col = next((i + 1 for i, h in enumerate(headers) if "club" in h), None)
    elif field_type == "gender":
        field_col = next((i + 1 for i, h in enumerate(headers) if h == "gender" or h == "sex"), None)
        if field_col is None:
            field_col = next((i + 1 for i, h in enumerate(headers) if "gender" in h or "sex" in h), None)
    elif field_type == "time":
        field_col = next((i + 1 for i, h in enumerate(headers) if "chip" in h and "time" in h), None)
        if field_col is None:
            field_col = next((i + 1 for i, h in enumerate(headers) if h == "time"), None)
        if field_col is None:
            field_col = next((i + 1 for i, h in enumerate(headers) if "time" in h), None)
    else:
        field_col = next((i + 1 for i, h in enumerate(headers) if "category" in h or h == "cat"), None)

    return name_col, field_col


def _row_name_value(ws, row_idx: int, name_col) -> str:
    if isinstance(name_col, tuple):
        first_val = ws.cell(row=row_idx, column=name_col[0]).value
        last_val = ws.cell(row=row_idx, column=name_col[1]).value
        first = "" if first_val is None else str(first_val).strip()
        last = "" if last_val is None else str(last_val).strip()
        return f"{first} {last}".strip()
    val = ws.cell(row=row_idx, column=name_col).value
    return "" if val is None else str(val).strip()


def _is_raw_data_file(path: Path) -> bool:
    lower_parts = {part.lower() for part in Path(path).parts}
    return "raw_data" in lower_parts
