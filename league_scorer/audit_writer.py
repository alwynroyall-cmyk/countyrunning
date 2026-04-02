import logging
from pathlib import Path
from typing import Dict

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

_HEADER_FILL = PatternFill("solid", fgColor="3A4658")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL = PatternFill("solid", fgColor="EEF3F8")

_WRAP_COLUMNS = {
    "Issue Codes",
    "Summary",
    "Message",
    "Next Step",
    "Reason",
    "Notes",
    "Clubs Seen",
    "Sexes Seen",
    "Categories Seen",
    "Races Seen",
}


def write_audit_workbook(sheets: Dict[str, pd.DataFrame], filepath: Path) -> None:
    filepath.parent.mkdir(parents=True, exist_ok=True)
    try:
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            for sheet_name, df in sheets.items():
                df.to_excel(writer, index=False, sheet_name=sheet_name)
                _style_sheet(writer.sheets[sheet_name], df)
        log.info("Audit workbook written: %s", filepath)
    except Exception as exc:
        log.error("Failed to write audit workbook '%s': %s", filepath, exc)
        raise


def _style_sheet(ws, df: pd.DataFrame) -> None:
    for cell in ws[1]:
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=0):
        use_alt_fill = row_idx % 2 == 1
        for cell in row:
            col_name = str(df.columns[cell.column - 1]) if cell.column - 1 < len(df.columns) else ""
            cell.alignment = Alignment(
                horizontal="left",
                vertical="center",
                wrap_text=col_name in _WRAP_COLUMNS,
            )
            if use_alt_fill:
                cell.fill = _ALT_FILL

    for col_idx, col_name in enumerate(df.columns, start=1):
        values = [str(col_name)]
        values.extend(str(v) for v in df.iloc[:, col_idx - 1] if str(v) != "")
        max_width = 100 if str(col_name) in _WRAP_COLUMNS else 40
        width = min(max((len(v) for v in values), default=8) + 2, max_width)
        ws.column_dimensions[get_column_letter(col_idx)].width = width