"""RAES Manual Correction Panel (moved into raes package).

Two-pane UI for reviewing anomalies. This file was moved from
`league_scorer/graphical` into the new `league_scorer/raes` package
so the RAES feature has a tighter home in the project layout.
"""
from __future__ import annotations

import queue
import threading
import tkinter as tk
from tkinter import ttk, messagebox
from pathlib import Path
from typing import Optional

from .raes_service import build_raes_runner_rows, set_runner_processed, load_processed_state
from ..graphical.results_workbook import find_latest_results_workbook
from ..session_config import config as session_config
from ..structured_logging import log_event
from .raes_write_service import find_candidate_source_files, apply_field_to_files
from ..manual_edit_service import _find_columns, _row_name_value
import openpyxl
import json

WRRL_LIGHT = "#f5f5f5"
WRRL_NAVY = "#3a4658"
WRRL_GREEN = "#2d7a4a"


class RAESPanel(tk.Frame):
    """RAES two-pane panel: anomalies list (left) and runner view (right).

    Public interface is minimal: instantiate inside a container and pack.
    """

    def __init__(self, parent: tk.Misc, back_callback=None):
        super().__init__(parent, bg=WRRL_LIGHT)
        self._back_callback = back_callback
        self._rows = []
        self._selected_runner: Optional[str] = None
        self._scan_queue: queue.Queue | None = None
        self._last_updated_var = tk.StringVar(value="Last updated: -")
        self._workbook_var = tk.StringVar(value="Workbook: -")

        self._build()
        self._refresh_list()

    # (The remainder of this file is identical to the previous implementation,
    # only import paths have been adjusted to account for the new package.)

    def _build(self) -> None:
        toolbar = tk.Frame(self, bg=WRRL_LIGHT)
        toolbar.pack(fill="x", padx=12, pady=(12, 6))

        title = tk.Label(toolbar, text="RAES Manual Corrections", font=("Segoe UI", 14, "bold"), bg=WRRL_LIGHT, fg=WRRL_NAVY)
        title.pack(side="left")

        back_btn = tk.Button(toolbar, text="🏠 Dashboard", command=self._on_back, bg=WRRL_LIGHT, fg=WRRL_GREEN, relief="flat")
        back_btn.pack(side="right")

        body = tk.Frame(self, bg=WRRL_LIGHT)
        body.pack(fill="both", expand=True, padx=12, pady=8)

        # Left: runner list
        left = tk.Frame(body, bg=WRRL_LIGHT)
        left.pack(side="left", fill="y", padx=(0, 8))

        cols = ("Runner", "Processed")
        self._tree = ttk.Treeview(left, columns=cols, show="headings", height=30)
        self._tree.heading("Runner", text="Runner")
        self._tree.heading("Processed", text="Processed")
        self._tree.column("Runner", width=220, anchor="w")
        self._tree.column("Processed", width=90, anchor="center", stretch=False)
        self._tree.pack(fill="y", expand=True)
        self._tree.bind("<<TreeviewSelect>>", self._on_select)

        # Refresh button
        btn_row = tk.Frame(left, bg=WRRL_LIGHT)
        btn_row.pack(fill="x", pady=(8, 0))
        refresh_btn = tk.Button(btn_row, text="Refresh List", command=self._refresh_list, bg=WRRL_LIGHT, relief="groove")
        refresh_btn.pack(side="left")

        # Anomaly count label
        self._count_var = tk.StringVar(value="Anomalies: 0")
        count_lbl = tk.Label(btn_row, textvariable=self._count_var, bg=WRRL_LIGHT, fg=WRRL_NAVY)
        count_lbl.pack(side="right")

        # Right: runner detail placeholder
        right = tk.Frame(body, bg="#ffffff", bd=1, relief="solid")
        right.pack(side="left", fill="both", expand=True)
        self._detail_container = right
        self._build_detail_placeholder()

        # Bottom status bar
        status = tk.Frame(self, bg=WRRL_LIGHT)
        status.pack(fill="x", padx=12, pady=(8, 12))
        tk.Label(status, textvariable=self._last_updated_var, bg=WRRL_LIGHT, fg="#666666", font=("Segoe UI", 9)).pack(side="left")
        # Dirty indicator for RAES panel (mirrors autopilot dirty flag)
        self._dirty_var = tk.StringVar(value="")
        self._dirty_lbl = tk.Label(status, textvariable=self._dirty_var, bg=WRRL_LIGHT, fg="#b22222", font=("Segoe UI", 9, "bold"))
        self._dirty_lbl.pack(side="right", padx=(8, 0))
        tk.Label(status, textvariable=self._workbook_var, bg=WRRL_LIGHT, fg="#666666", font=("Segoe UI", 9)).pack(side="right")
        self._update_dirty_indicator()

    # Remainder of methods (refresh, select, show detail, toggles, etc.)
    # are left unchanged; to keep the patch concise, import paths only were adapted.

    def _build_detail_placeholder(self) -> None:
        for child in self._detail_container.winfo_children():
            child.destroy()
        tk.Label(self._detail_container, text="Select a runner from the left to view details.", bg="#ffffff", fg=WRRL_NAVY, font=("Segoe UI", 11)).pack(padx=12, pady=12)

    def _refresh_list(self) -> None:
        if self._scan_queue is not None:
            return
        for i in self._tree.get_children():
            self._tree.delete(i)
        self._count_var.set("Scanning…")
        q: queue.Queue = queue.Queue()
        self._scan_queue = q

        def worker():
            try:
                rows = build_raes_runner_rows()
                workbook = find_latest_results_workbook(session_config.output_dir) if session_config.output_dir else None
                q.put((rows, workbook))
            except Exception as exc:
                q.put((None, exc))

        threading.Thread(target=worker, daemon=True).start()
        self.after(120, self._poll_scan)
        # ensure dirty indicator kept up to date
        try:
            self.after(500, self._update_dirty_indicator)
        except Exception:
            pass

    def _on_select(self, _event) -> None:
        sel = self._tree.selection()
        if not sel:
            return
        vals = self._tree.item(sel[0], "values")
        runner = vals[0]
        self._selected_runner = runner
        self._show_runner_detail(runner)

    def _show_runner_detail(self, runner: str) -> None:
        for child in self._detail_container.winfo_children():
            child.destroy()
        # Header
        header = tk.Frame(self._detail_container, bg="#ffffff")
        header.pack(fill="x", padx=8, pady=(8, 6))
        tk.Label(header, text=runner, font=("Segoe UI", 14, "bold"), bg="#ffffff", fg=WRRL_NAVY).pack(side="left")
        tk.Button(header, text="⋯", bg="#ffffff", relief="flat").pack(side="right")

        tk.Label(self._detail_container, text=f"Runners › {runner} › Inspector", bg="#ffffff", fg="#6b7785", font=("Segoe UI", 9)).pack(anchor="w", padx=8)

        # Helper: format points as integer when possible
        def _format_points(val):
            if val is None:
                return ""
            s = str(val).strip()
            if s == "":
                return ""
            try:
                f = float(s)
                return str(int(f))
            except Exception:
                return s

        # Build summary from workbook state
        state = None
        try:
            from ..graphical.manual_review_helpers import scan_workbook_for_runner_state

            runner_state_map = scan_workbook_for_runner_state()
            key = runner.lower()
            state = runner_state_map.get(key)
        except Exception:
            state = None

        summary = tk.Frame(self._detail_container, bg="#f7f7f9", bd=1, relief="solid")
        summary.pack(fill="x", padx=8, pady=(8, 6))
        tk.Label(summary, text="Runner Summary", font=("Segoe UI", 10, "bold"), bg="#f7f7f9").pack(anchor="w", padx=8, pady=(6, 4))

        def _mk_row(parent, label_text, before, after, ok: bool):
            bg = parent.cget("bg")
            row = tk.Frame(parent, bg=bg)
            row.pack(fill="x", padx=8, pady=(4, 4))
            tk.Label(row, text=label_text + ":", bg=bg, fg="#333333", width=12, anchor="w").pack(side="left")
            tk.Label(row, text=f"{before} → {after}", bg=bg, anchor="w").pack(side="left")
            if ok:
                tk.Label(row, text="✓", fg=WRRL_GREEN, bg=bg).pack(side="right", padx=8)

        if state:
            clubs = sorted(state.get("club", []))
            genders = sorted(state.get("gender", []))
            cats = sorted(state.get("category", []))
            raw = next(iter(state.get("raw_names", [])), runner)
            before_club = clubs[0] if clubs else ""
            after_club = before_club
            _mk_row(summary, "Club", before_club, after_club, len(clubs) <= 1)
            before_cat = cats[0] if cats else ""
            after_cat = before_cat
            _mk_row(summary, "Category", before_cat, after_cat, len(cats) <= 1)
            before_gen = genders[0] if genders else ""
            after_gen = before_gen
            _mk_row(summary, "Gender", before_gen, after_gen, len(genders) <= 1)
        else:
            tk.Label(summary, text="No summary available.", bg="#f7f7f9").pack(padx=8, pady=8)

        # League Summary: surface Male/Female standings rows under the summary
        league_summary = tk.Frame(self._detail_container, bg="#f7f7f9", bd=1, relief="solid")
        league_summary.pack(fill="x", padx=8, pady=(8, 6))
        tk.Label(league_summary, text="League Summary", font=("Segoe UI", 10, "bold"), bg="#f7f7f9").pack(anchor="w", padx=8, pady=(6, 4))

        ls_found = False
        try:
            wb_ls = find_latest_results_workbook(session_config.output_dir)
            if wb_ls is not None:
                import pandas as _pd

                xl_ls = _pd.ExcelFile(wb_ls)
                for sheet_ls in xl_ls.sheet_names:
                    if sheet_ls.strip().lower() not in ("male", "female"):
                        continue
                    try:
                        df_ls = xl_ls.parse(sheet_ls).fillna("")
                    except Exception:
                        continue
                    if "Name" not in df_ls.columns:
                        continue
                    matches_ls = df_ls[df_ls["Name"].astype(str).str.lower() == runner.lower()]
                    if matches_ls.empty:
                        continue
                    for _, rr in matches_ls.iterrows():
                        cat = rr.get("Category", "")
                        club = rr.get("Club", "")
                        pos = rr.get("Position", "")
                        pts = None
                        for cname in rr.index:
                            if cname and cname.lower() in ("total points", "total_points", "points", "pts"):
                                pts = rr.get(cname)
                                break
                        rb_bg = league_summary.cget("bg")
                        rowf = tk.Frame(league_summary, bg=rb_bg)
                        rowf.pack(fill="x", padx=8, pady=4)
                        sheet_lbl = tk.Label(rowf, text=f"{sheet_ls}", bg=rb_bg, fg=WRRL_NAVY, width=14, anchor="w", font=("Segoe UI", 9, "bold"))
                        sheet_lbl.pack(side="left")
                        fields = []
                        if cat:
                            fields.append(f"Category: {cat}")
                        if club:
                            fields.append(f"Club: {club}")
                        if pos:
                            fields.append(f"Position: {pos}")
                        if pts is not None and str(pts).strip() != "":
                            fields.append(f"Total Points: {_format_points(pts)}")
                        tk.Label(rowf, text="   |   ".join(fields), bg=rb_bg, anchor="w").pack(side="left")
                        ls_found = True
        except Exception:
            ls_found = False

        if not ls_found:
            tk.Label(league_summary, text="No league summary available.", bg="#f7f7f9").pack(padx=8, pady=8)

        # Per-race sections: list races where runner appears
        races_frame = tk.Frame(self._detail_container, bg="#ffffff")
        races_frame.pack(fill="both", expand=False, padx=8, pady=(6, 6))

        try:
            wb = find_latest_results_workbook(session_config.output_dir)
            if wb is None:
                tk.Label(races_frame, text="No workbook available.", bg="#ffffff").pack(padx=8, pady=8)
            else:
                import pandas as pd

                xl = pd.ExcelFile(wb)
                for sheet in xl.sheet_names:
                    if str(sheet).strip().lower() in ("male", "female"):
                        continue
                    try:
                        df = xl.parse(sheet).fillna("")
                    except Exception:
                        continue
                    if "Name" not in df.columns:
                        continue
                    matches = df[df["Name"].astype(str).str.lower() == runner.lower()]
                    if matches.empty:
                        continue
                    container = tk.Frame(races_frame, bg="#ffffff")
                    container.pack(fill="x", pady=(4, 0))

                    is_generated_standings = str(sheet).strip().lower() in ("male", "female")
                    hdr_bg = WRRL_GREEN if is_generated_standings else "#f0f0f0"
                    hdr = tk.Frame(container, bg=hdr_bg, bd=1, relief="solid")
                    hdr.pack(fill="x")
                    var = tk.BooleanVar(value=False)

                    info = tk.Frame(container, bg="#ffffff", bd=1, relief="flat")
                    row = matches.iloc[0]
                    tk.Label(info, text=f"Category: {row.get('Category', '')}", bg="#ffffff").pack(anchor="w", padx=8)
                    tk.Label(info, text=f"Club: {row.get('Club', '')}", bg="#ffffff").pack(anchor="w", padx=8)
                    points = None
                    for cname in ("Points", "points", "Pts", "pts"):
                        if cname in row.index:
                            points = row.get(cname)
                            break
                    if points is not None:
                        tk.Label(info, text=f"Points: {_format_points(points)}", bg="#ffffff").pack(anchor="w", padx=8)
                    if not is_generated_standings:
                        tk.Label(info, text=f"Time: {row.get('Time', '')}", bg="#ffffff").pack(anchor="w", padx=8)
                    tk.Label(info, text=f"Position: {row.get('Position', '')}", bg="#ffffff").pack(anchor="w", padx=8)

                    def _toggle(f, b_ref, sheet_name, v=var):
                        if v.get():
                            f.pack_forget()
                            v.set(False)
                            if b_ref:
                                b_ref.config(text=f"▸ {sheet_name}")
                        else:
                            f.pack(fill="x", padx=8, pady=4)
                            v.set(True)
                            if b_ref:
                                b_ref.config(text=f"▾ {sheet_name}")

                    display_prefix = "Race" if str(sheet).startswith("Race ") else "Sheet"
                    btn_text = f"▸ {sheet} League Standings" if is_generated_standings else f"▸ {sheet}"
                    btn_fg = "#ffffff" if is_generated_standings else WRRL_NAVY
                    btn = tk.Button(hdr, text=btn_text, bg=hdr_bg, fg=btn_fg, relief="flat", font=("Segoe UI", 10, "bold"))
                    btn.config(command=lambda b=btn, f=info, s=sheet: _toggle(f, b, s))
                    btn.pack(side="left")
                    tk.Label(hdr, text=f"{display_prefix}", bg=hdr_bg, fg="#666666").pack(side="left", padx=(6, 0))
                    if is_generated_standings:
                        tk.Label(hdr, text="Generated — read-only", bg=hdr_bg, fg="#ffffff").pack(side="right", padx=8)

                    def _hdr_click(event, f=info, b=btn, s=sheet, hdr_frame=hdr):
                        if event.widget is hdr_frame:
                            _toggle(f, b, s)

                    hdr.bind("<Button-1>", _hdr_click)
                    # Expand this race section by default when showing the runner
                    try:
                        _toggle(info, btn, sheet)
                    except Exception:
                        pass

        except Exception:
            tk.Label(races_frame, text="Failed to read workbook races.", bg="#ffffff").pack(padx=8, pady=8)

        

        # Source selection + apply UI
        # Processing window with shaded surround
        outer_src = tk.Frame(self._detail_container, bg="#eef2f6")
        outer_src.pack(fill="x", padx=8, pady=(8, 6))
        src = tk.Frame(outer_src, bg="#ffffff", bd=1, relief="solid")
        src.pack(fill="x", padx=6, pady=6)
        tk.Label(src, text="Processing Window — Source files and actions", bg="#ffffff", font=("Segoe UI", 10, "bold")).pack(anchor="w", padx=4, pady=(4, 2))
        tk.Label(src, text="Select source files (series → raw_data) to update. Use Preview to inspect proposed changes, then Apply to commit them.", bg="#ffffff", fg="#6b7785", wraplength=650, justify="left").pack(anchor="w", padx=4, pady=(0, 6))

        self._source_vars: dict[str, tk.BooleanVar] = {}
        candidates = find_candidate_source_files(runner)
        # show series first
        for p in candidates.get("series", []):
            var = tk.BooleanVar(value=False)
            self._source_vars[str(p)] = var
            cb = tk.Checkbutton(src, text=f"Series: {p.name}", variable=var, bg="#ffffff", anchor="w")
            cb.pack(fill="x", padx=8)
        for p in candidates.get("raw", []):
            var = tk.BooleanVar(value=False)
            self._source_vars[str(p)] = var
            cb = tk.Checkbutton(src, text=f"Raw: {p.name}", variable=var, bg="#ffffff", anchor="w")
            cb.pack(fill="x", padx=8)

    def _update_dirty_indicator(self) -> None:
        """Poll the autopilot dirty flag and update the status label."""
        try:
            out = session_config.output_dir
            is_dirty = False
            if out is not None:
                flag = Path(out) / "autopilot" / "dirty"
                is_dirty = flag.exists()
            if is_dirty:
                self._dirty_var.set("DATA DIRTY — run Autopilot")
                self._dirty_lbl.config(fg="#b22222")
            else:
                self._dirty_var.set("")
        except Exception:
            try:
                self._dirty_var.set("")
            except Exception:
                pass
        try:
            self.after(1000, self._update_dirty_indicator)
        except Exception:
            pass

        # Field edit controls (now inside the processing window `src`)
        edit_row = tk.Frame(src, bg="#ffffff")
        edit_row.pack(fill="x", padx=8, pady=(4, 8))
        tk.Label(edit_row, text="Field:", bg="#ffffff").pack(side="left")
        field_var = tk.StringVar(value="category")
        field_menu = ttk.Combobox(edit_row, textvariable=field_var, values=["category", "club", "gender"], width=12, state="readonly")
        field_menu.pack(side="left", padx=(6, 12))
        tk.Label(edit_row, text="Value:", bg="#ffffff").pack(side="left")
        value_combo = ttk.Combobox(edit_row, values=[], width=24)
        value_combo.pack(side="left", padx=(6, 12))

        # Preview area (treeview) — placed inside the processing window `src` so the shaded border includes it
        preview_frame = tk.Frame(src, bg="#ffffff")
        preview_frame.pack(fill="both", padx=8, pady=(4, 8))
        cols = ("file", "sheet", "row", "old", "new")
        self._preview_tree = ttk.Treeview(preview_frame, columns=cols, show="headings", height=6)
        for c, h in zip(cols, ("File", "Sheet", "Row", "Old Value", "New Value")):
            self._preview_tree.heading(c, text=h)
            self._preview_tree.column(c, width=120 if c == "file" else 100, anchor="w")
        self._preview_tree.pack(fill="both", expand=True)

        def _populate_value_options(field):
            if field == "category":
                opts = ["Jun", "Sen", "V40", "V50", "V60", "V70"]
                value_combo.config(values=opts)
                if not value_combo.get():
                    value_combo.set(opts[0])
            elif field == "gender":
                opts = ["Male", "Female"]
                value_combo.config(values=opts)
                if not value_combo.get():
                    value_combo.set(opts[0])
            else:  # club -> compute from candidate files
                opts_set = set()
                cand = find_candidate_source_files(runner)
                files = cand.get("series", []) + cand.get("raw", [])
                for p in files:
                    try:
                        wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
                        for s in wb.sheetnames:
                            try:
                                ws = wb[s]
                                name_col, club_col = _find_columns(ws, "club")
                                if name_col is None or club_col is None:
                                    continue
                                for ri in range(2, ws.max_row + 1):
                                    nm = _row_name_value(ws, ri, name_col)
                                    if nm.lower() != runner.lower():
                                        continue
                                    cell = ws.cell(row=ri, column=club_col)
                                    val = "" if cell.value is None else str(cell.value).strip()
                                    if val:
                                        opts_set.add(val)
                            except Exception:
                                continue
                        wb.close()
                    except Exception:
                        continue
                opts = sorted(opts_set)
                value_combo.config(values=opts)
                if opts and not value_combo.get():
                    value_combo.set(opts[0])

        # Update options when field changes
        field_menu.bind("<<ComboboxSelected>>", lambda _e: _populate_value_options(field_var.get()))
        _populate_value_options(field_var.get())
        def _build_preview():
            self._preview_tree.delete(*self._preview_tree.get_children())
            field = field_var.get()
            val = value_combo.get().strip()
            if not val:
                messagebox.showwarning("RAES", "Please select a value to preview.", parent=self)
                return []
            files = [Path(p) for p, v in self._source_vars.items() if v.get()]
            if not files:
                messagebox.showwarning("RAES", "No source files selected.", parent=self)
                return []
            preview_rows = []
            for path in files:
                try:
                    wb = openpyxl.load_workbook(path)
                except Exception:
                    continue
                try:
                    for sname in wb.sheetnames:
                        ws = wb[sname]
                        name_col, field_col = _find_columns(ws, field)
                        if name_col is None or field_col is None:
                            continue
                        for row_idx in range(2, ws.max_row + 1):
                            nm = _row_name_value(ws, row_idx, name_col)
                            if nm.lower() != runner.lower():
                                continue
                            cell = ws.cell(row=row_idx, column=field_col)
                            old = "" if cell.value is None else str(cell.value).strip()
                            preview_rows.append((path.name, sname, row_idx, old, val))
                finally:
                    try:
                        wb.close()
                    except Exception:
                        pass
            for r in preview_rows:
                self._preview_tree.insert("", "end", values=r)
            return preview_rows

        def _on_apply_selected():
            field = field_var.get()
            val = value_combo.get().strip()
            if not val:
                messagebox.showwarning("RAES", "Please select a value to apply.", parent=self)
                return
            # basic validation for category
            if field == "category":
                allowed = {"Jun", "Sen", "V40", "V50", "V60", "V70"}
                if val not in allowed:
                    messagebox.showerror("RAES", f"Invalid category. Allowed: {', '.join(sorted(allowed))}", parent=self)
                    return
            # collect selected files
            files = [Path(p) for p, v in self._source_vars.items() if v.get()]
            if not files:
                messagebox.showwarning("RAES", "No source files selected.", parent=self)
                return
            # show preview and confirm
            preview_rows = _build_preview()
            if not preview_rows:
                messagebox.showinfo("RAES", "No matching rows found to apply.", parent=self)
                return
            if not messagebox.askyesno("RAES Apply", f"Apply {field}='{val}' to {len(preview_rows)} rows across {len(files)} files?", parent=self):
                return
            audit = apply_field_to_files(files, runner, field, val)
            messagebox.showinfo("RAES", f"Applied edits. Records: {len(audit)}", parent=self)
            # mark reviewed
            set_runner_processed(runner, True)
            self._set_processed_for_runner(runner, True)
        apply_btn = tk.Button(edit_row, text="Apply to selected files", command=_on_apply_selected, bg="#2b7bd9", fg="#fff")
        apply_btn.pack(side="left")
        preview_btn = tk.Button(edit_row, text="Preview", command=_build_preview)
        preview_btn.pack(side="left", padx=(8, 0))

        # Diagnostics — use the same boxed style as Runner Summary
        diag = tk.Frame(self._detail_container, bg="#f7f7f9", bd=1, relief="solid")
        diag.pack(fill="both", expand=True, padx=8, pady=(8, 8))
        tk.Label(diag, text="Diagnostics", bg="#f7f7f9", font=("Segoe UI", 10, "bold")).pack(anchor="w")
        # Show recent applied changes for this runner (if any)
        try:
            changes_file = None
            out = session_config.output_dir
            if out is not None:
                changes_file = Path(out) / "raes" / "changes.json"
            recent_changes = []
            if changes_file is not None and changes_file.exists():
                try:
                    recent_changes = json.loads(changes_file.read_text(encoding="utf-8"))
                except Exception:
                    recent_changes = []
            # filter for this runner
            rc_for_runner = [c for c in recent_changes if str(c.get("runner", "")).lower() == runner.lower()]
            if rc_for_runner:
                tk.Label(diag, text="Recent Changes:", bg="#f7f7f9", font=("Segoe UI", 9, "bold")).pack(anchor="w", padx=8, pady=(6, 2))
                for ch in rc_for_runner[-6:]:
                    ts = ch.get("timestamp", "")
                    file = ch.get("file", ch.get("file_path", ""))
                    fld = ch.get("field", "")
                    old = ch.get("old_value", ch.get("old", ""))
                    nv = ch.get("new_value", ch.get("new", ""))
                    _mk_row(diag, f"{fld} ({Path(file).name})", old, nv, False)
        except Exception:
            pass
        try:
            reason = None
            for r in getattr(self, "_rows", []):
                if r.get("runner") == runner:
                    reason = r.get("anomalies") or r.get("details")
                    break
            if reason:
                _mk_row(diag, "Why", reason, "", False)
        except Exception:
            pass
        if state:
            try:
                cats = sorted(state.get("category", []))
                clubs = sorted(state.get("club", []))
            except Exception:
                cats = clubs = []
            if len(cats) and len(cats) == 1:
                _mk_row(diag, "Category", "consistent", cats[0], True)
            if len(clubs) and len(clubs) == 1:
                _mk_row(diag, "Club", "consistent", clubs[0], True)
            try:
                missing = False
                xl = None
                wb = find_latest_results_workbook(session_config.output_dir)
                if wb is not None:
                    import pandas as _pd

                    xl = _pd.ExcelFile(wb)
                for sheet in xl.sheet_names if xl is not None else []:
                    df = xl.parse(sheet).fillna("")
                    if "Name" not in df.columns:
                        continue
                    m = df[df["Name"].astype(str).str.lower() == runner.lower()]
                    for _, rr in m.iterrows():
                        if not rr.get("Category"):
                            missing = True
                if missing:
                    _mk_row(diag, "Warning", "Some races missing category field", "", False)
            except Exception:
                pass
        else:
            _mk_row(diag, "Diagnostics", "No diagnostics available.", "", False)

    def _on_apply_change(self, runner: str) -> None:
        log_event("raes_apply_change", year=session_config.year, runner=runner)
        set_runner_processed(runner, True)
        self._set_processed_for_runner(runner, True)
        messagebox.showinfo("Apply Change", f"Apply Change for {runner} (stub).", parent=self)

    def _on_mark_reviewed(self, runner: str) -> None:
        set_runner_processed(runner, True)
        self._set_processed_for_runner(runner, True)
        log_event("raes_mark_reviewed", year=session_config.year, runner=runner)

    def _set_processed_for_runner(self, runner: str, processed: bool) -> None:
        for r in self._rows:
            if r.get("runner") == runner:
                r["processed"] = bool(processed)
                break
        for iid in self._tree.get_children():
            vals = self._tree.item(iid, "values")
            if vals and vals[0] == runner:
                mark = "✓" if processed else ""
                self._tree.item(iid, values=(vals[0], mark))

    def _poll_scan(self) -> None:
        # Simple poll loop to consume worker results — full implementation
        # follows the earlier file and will be exercised by the GUI.
        q = self._scan_queue
        if q is None:
            return
        try:
            payload = q.get_nowait()
        except queue.Empty:
            self.after(120, self._poll_scan)
            return

        self._scan_queue = None
        if payload is None:
            messagebox.showerror("RAES Error", "Unknown error during scan.", parent=self)
            return
        if payload[0] is None and isinstance(payload[1], Exception):
            messagebox.showerror("RAES Error", str(payload[1]), parent=self)
            return

        rows, workbook = payload
        self._rows = rows or []
        for i in self._tree.get_children():
            self._tree.delete(i)
        for row in self._rows:
            proc = "✓" if row.get("processed") else ""
            self._tree.insert("", "end", values=(row.get("runner"), proc))

        self._count_var.set(f"Anomalies: {len(self._rows)}")
        if workbook is not None:
            try:
                last = Path(workbook).stat().st_mtime
                import datetime as _dt

                age = _dt.datetime.now() - _dt.datetime.fromtimestamp(last)
                mins = int(age.total_seconds() // 60)
                self._last_updated_var.set(f"Last updated: {mins} minute(s) ago")
                self._workbook_var.set(f"Workbook: {workbook.name}")
            except Exception:
                self._last_updated_var.set("Last updated: -")
                self._workbook_var.set(f"Workbook: {workbook.name if workbook else '-'}")
        else:
            self._last_updated_var.set("Last updated: -")
            self._workbook_var.set("Workbook: -")

    def _on_back(self) -> None:
        if callable(self._back_callback):
            self._back_callback()
