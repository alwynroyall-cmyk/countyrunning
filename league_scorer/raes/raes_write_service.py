"""Write-service for RAES: locate candidate source files and apply field edits.

Provides series-first precedence and supports editing series workbooks
or raw_data input files. Changes are applied in-place using atomic saves
and an audit record is appended to outputs/raes/changes.json.
"""
from __future__ import annotations

from pathlib import Path
import json
import datetime
from typing import List, Dict
import re

import openpyxl

from ..session_config import config as session_config
from ..structured_logging import log_event
from ..manual_edit_service import _find_columns, _atomic_save
from ..name_lookup import load_name_corrections


def _ensure_raes_output_dir() -> Path | None:
    out = session_config.output_dir
    if out is None:
        return None
    d = Path(out) / "raes"
    d.mkdir(parents=True, exist_ok=True)
    return d


def find_candidate_source_files(runner: str, race_sheet: str | None = None) -> Dict[str, List[Path]]:
    """Return candidate files grouped by 'series' and 'raw'.

    - Series: any workbook under `session_config.series_dir` that contains
      a sheet named like `race_sheet` or which contains the runner.
    - Raw: any workbook under `session_config.raw_data_dir` that contains the runner.
    """
    runner_key = runner.lower()
    series_files: List[Path] = []
    raw_files: List[Path] = []

    # Load reviewed name corrections (alias -> preferred) so we match original/raw names
    alias_map = {}
    try:
        ctrl = session_config.control_dir
        if ctrl is not None:
            nmf = Path(ctrl) / "name_corrections.xlsx"
            alias_map = load_name_corrections(nmf)
    except Exception:
        alias_map = {}

    sdir = session_config.series_dir
    rdir = session_config.raw_data_dir

    def _list_xlsx(dirp: Path | None):
        if dirp is None or not dirp.exists():
            return []
        return [p for p in dirp.iterdir() if p.suffix.lower() in ('.xlsx', '.xlsm', '.xls')]

    # Scan series files first (series workbooks may contain multiple "series" files)
    for p in _list_xlsx(sdir):
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            names = [s for s in wb.sheetnames]
            matched = False
            if race_sheet and race_sheet in names:
                matched = True
            else:
                # scan sheets for runner (match displayed name or any reviewed raw-name alias)
                for s in names:
                    try:
                        ws = wb[s]
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            if not row:
                                continue
                            vals = [str(x).strip().lower() for x in row if x is not None]
                            # direct match
                            if runner_key in vals:
                                matched = True
                                break
                            # alias match
                            for v in vals:
                                pref = alias_map.get(v)
                                if pref and pref.strip().lower() == runner_key:
                                    matched = True
                                    break
                            if matched:
                                break
                        if matched:
                            break
                    except Exception:
                        continue
            wb.close()
            if matched:
                series_files.append(p)
        except Exception:
            continue
    # build set of race numbers present in series files to avoid showing
    # corresponding raw round copies later
    def _race_num_from_name(nm: str):
        if not nm:
            return None
        m = re.search(r"race\s*#?\s*(\d+)", nm, re.IGNORECASE)
        if m:
            try:
                return int(m.group(1))
            except Exception:
                return None
        return None

    series_race_nums = set()
    for p in series_files:
        rn = _race_num_from_name(p.stem)
        if rn is not None:
            series_race_nums.add(rn)

    # Raw data files
    for p in _list_xlsx(rdir):
        try:
            wb = openpyxl.load_workbook(p, read_only=True, data_only=True)
            found = False
            for s in wb.sheetnames:
                try:
                    ws = wb[s]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        if not row:
                            continue
                        vals = [str(x).strip().lower() for x in row if x is not None]
                        # match if any cell directly matches the preferred name
                        if runner_key in vals:
                            found = True
                            break
                        # or if any cell is a reviewed alias that maps to the preferred name
                        for v in vals:
                            pref = alias_map.get(v)
                            if pref and pref.strip().lower() == runner_key:
                                found = True
                                break
                        if found:
                            break
                    if found:
                        break
                except Exception:
                    continue
            wb.close()
            if found:
                # Exclude consolidated workbooks and series round workbooks from raw_data candidates
                stem = p.stem.lower()
                # also exclude round copies or raw files that match a series race number
                if "consolidated" in stem or "series" in stem:
                    continue
                # exclude if filename indicates a round (e.g. 'round')
                if "round" in stem:
                    continue
                # exclude if this raw file corresponds to a race number already present in series files
                rn = _race_num_from_name(p.stem)
                if rn is not None and rn in series_race_nums:
                    continue
                raw_files.append(p)
        except Exception:
            continue

    return {"series": series_files, "raw": raw_files}


def apply_field_to_files(files: List[Path], runner: str, field_type: str, target_value: str) -> List[Dict]:
    """Apply `target_value` for `field_type` to matching rows in each file.

    Returns a list of audit change records.
    """
    audit: List[Dict] = []
    runner_key = runner.lower()
    # load name alias mappings so we can match original/raw names when applying
    alias_map = {}
    try:
        ctrl = session_config.control_dir
        if ctrl is not None:
            nmf = Path(ctrl) / "name_corrections.xlsx"
            alias_map = load_name_corrections(nmf)
    except Exception:
        alias_map = {}
    for path in files:
        try:
            wb = openpyxl.load_workbook(path)
        except Exception as exc:
            audit.append({"file": str(path), "error": str(exc)})
            continue

        try:
            changed = False
            for sname in wb.sheetnames:
                ws = wb[sname]
                name_col, field_col = _find_columns(ws, field_type)
                if name_col is None or field_col is None:
                    continue
                # iterate rows
                for row_idx in range(2, ws.max_row + 1):
                    current_name = None
                    if isinstance(name_col, tuple):
                        first = ws.cell(row=row_idx, column=name_col[0]).value or ""
                        last = ws.cell(row=row_idx, column=name_col[1]).value or ""
                        current_name = f"{str(first).strip()} {str(last).strip()}".strip()
                    else:
                        current_name = ws.cell(row=row_idx, column=name_col).value or ""
                    current_name_l = str(current_name).strip().lower()
                    # match either the preferred name or a raw alias mapped to the preferred
                    if current_name_l != runner_key and alias_map.get(current_name_l, '').strip().lower() != runner_key:
                        continue
                    cell = ws.cell(row=row_idx, column=field_col)
                    old = "" if cell.value is None else str(cell.value).strip()
                    if old == target_value:
                        continue
                    cell.value = target_value
                    changed = True
                    audit.append({
                        "timestamp": datetime.datetime.utcnow().isoformat() + "Z",
                        "file": str(path),
                        "sheet": sname,
                        "runner": runner,
                        "field": field_type,
                        "old_value": old,
                        "new_value": target_value,
                    })
            if changed:
                _atomic_save(wb, path)
                log_event("raes_applied_field", year=session_config.year, file=str(path), runner=runner, field=field_type)
        except Exception as exc:
            audit.append({"file": str(path), "error": str(exc)})
        finally:
            try:
                wb.close()
            except Exception:
                pass

    # append audit to outputs/raes/changes.json
    outdir = _ensure_raes_output_dir()
    if outdir is not None:
        changes_file = outdir / "changes.json"
        existing = []
        try:
            if changes_file.exists():
                existing = json.loads(changes_file.read_text(encoding="utf-8"))
        except Exception:
            existing = []
        existing.extend(audit)
        try:
            changes_file.write_text(json.dumps(existing, indent=2), encoding="utf-8")
        except Exception:
            pass

    # Mark data as dirty so autopilot/UI know a run is required.
    try:
        out = session_config.output_dir
        if out is not None:
            flag = Path(out) / "autopilot" / "dirty"
            flag.parent.mkdir(parents=True, exist_ok=True)
            flag.write_text(datetime.datetime.utcnow().isoformat() + "Z\n", encoding="utf-8")
            log_event("data_dirty", year=session_config.year, file_changes=len(audit))
            # Also write a RAES-specific marker so the GUI can pick up edits immediately
            try:
                raes_flag = Path(out) / "raes" / "dirty"
                raes_flag.parent.mkdir(parents=True, exist_ok=True)
                raes_flag.write_text(datetime.datetime.utcnow().isoformat() + "Z\n", encoding="utf-8")
            except Exception:
                pass
    except Exception:
        # non-fatal if we cannot write the dirty flag
        pass

    return audit
