import pandas as pd
from openpyxl import load_workbook

from league_scorer.models import RunnerRaceEntry
from league_scorer.output_writer import (
    _is_race_over_5k_name,
    _sanitise_df_for_export,
    build_category_mismatch_todo_df,
    build_time_qry_todo_df,
    write_category_mismatch_todo,
    write_time_qry_todo,
    write_unrecognised_clubs,
)


_UNSET = object()


def _entry(
    *,
    name: str,
    club: str,
    gender: str,
    race_number: int,
    normalised_category: str,
    raw_category: str,
    raw_club: str | None = None,
    preferred_club: str | None | object = _UNSET,
    eligible: bool = True,
):
    resolved_preferred_club = club if preferred_club is _UNSET else preferred_club
    return RunnerRaceEntry(
        name=name,
        raw_club=raw_club if raw_club is not None else club,
        preferred_club=resolved_preferred_club,
        gender=gender,
        raw_category=raw_category,
        normalised_category=normalised_category,
        time_str="00:40:00",
        time_seconds=2400.0,
        race_number=race_number,
        eligible=eligible,
    )


def test_build_category_mismatch_todo_df_flags_eligible_mismatches_only():
    race_data = {
        1: [
            _entry(
                name="Alex Runner",
                club="Club A",
                gender="M",
                race_number=1,
                normalised_category="Sen",
                raw_category="Sen",
            ),
            _entry(
                name="Pat Stable",
                club="Club A",
                gender="F",
                race_number=1,
                normalised_category="V40",
                raw_category="Vet 40",
            ),
        ],
        2: [
            _entry(
                name="Alex Runner",
                club="Club A",
                gender="M",
                race_number=2,
                normalised_category="V40",
                raw_category="Vet 40",
            ),
            _entry(
                name="Pat Stable",
                club="Club A",
                gender="F",
                race_number=2,
                normalised_category="V40",
                raw_category="Vet 40",
            ),
            _entry(
                name="Casey Ineligible",
                club="Club B",
                gender="M",
                race_number=2,
                normalised_category="Sen",
                raw_category="Sen",
                eligible=False,
            ),
        ],
    }

    df = build_category_mismatch_todo_df(race_data)

    assert len(df) == 1
    row = df.iloc[0]
    assert row["Issue Type"] == "Category Mismatch"
    assert row["Name"] == "Alex Runner"
    assert row["Club"] == "Club A"
    assert row["Categories Seen"] == "Sen, V40"
    assert row["Suggested Category"] in {"Sen", "V40"}
    assert "R1:Sen" in row["Race Category Sequence"]
    assert "R2:V40" in row["Race Category Sequence"]


def test_build_category_mismatch_todo_df_includes_club_mismatch_with_one_eligible_club():
    race_data = {
        1: [
            _entry(
                name="Riley Switch",
                club="Club A",
                gender="F",
                race_number=1,
                normalised_category="Sen",
                raw_category="Sen",
                raw_club="Club A",
                preferred_club="Club A",
                eligible=True,
            ),
        ],
        2: [
            _entry(
                name="Riley Switch",
                club="Club A",
                gender="F",
                race_number=2,
                normalised_category="Sen",
                raw_category="Sen",
                raw_club="Guest Runners",
                preferred_club=None,
                eligible=False,
            ),
            _entry(
                name="Taylor NoEligible",
                club="Club X",
                gender="M",
                race_number=2,
                normalised_category="Sen",
                raw_category="Sen",
                raw_club="Guest X",
                preferred_club=None,
                eligible=False,
            ),
        ],
        3: [
            _entry(
                name="Taylor NoEligible",
                club="Club X",
                gender="M",
                race_number=3,
                normalised_category="Sen",
                raw_category="Sen",
                raw_club="Guest Y",
                preferred_club=None,
                eligible=False,
            ),
        ],
    }

    df = build_category_mismatch_todo_df(race_data)

    club_rows = df[df["Issue Type"] == "Club Mismatch"]
    assert len(club_rows) == 1
    row = club_rows.iloc[0]
    assert row["Name"] == "Riley Switch"
    assert "Club A" in row["Clubs Seen"]
    assert "Guest Runners" in row["Clubs Seen"]
    assert row["Suggested Club"] == "Club A"


def test_build_category_mismatch_todo_df_includes_fix_rows_for_manual_review():
    race_data = {
        3: [
            _entry(
                name="Westbury Marker",
                club="Club A",
                gender="M",
                race_number=3,
                normalised_category="FIX",
                raw_category="Top 3 Male",
                eligible=True,
            ),
            _entry(
                name="Westbury Marker",
                club="Club A",
                gender="M",
                race_number=3,
                normalised_category="FIX",
                raw_category="Pacer",
                eligible=True,
            ),
        ]
    }

    df = build_category_mismatch_todo_df(race_data)
    fix_rows = df[df["Issue Type"] == "Category FIX"]
    assert len(fix_rows) == 1
    row = fix_rows.iloc[0]
    assert row["Name"] == "Westbury Marker"
    assert "Top 3 Male" in row["Source Category Values"]
    assert "Pacer" in row["Source Category Values"]


def test_build_time_qry_todo_df_includes_qry_and_invalid_rows():
    qry_runner = _entry(
        name="Taylor Query",
        club="Club A",
        gender="F",
        race_number=2,
        normalised_category="Sen",
        raw_category="Sen",
    )
    qry_runner.time_str = "QRY"
    qry_runner.time_seconds = None
    qry_runner.source_row = 25

    invalid_runner = _entry(
        name="Jordan Zero",
        club="Club B",
        gender="M",
        race_number=1,
        normalised_category="V40",
        raw_category="V40",
    )
    invalid_runner.time_str = ""
    invalid_runner.time_seconds = 0.0
    invalid_runner.source_row = 17

    valid_runner = _entry(
        name="Casey Valid",
        club="Club C",
        gender="F",
        race_number=1,
        normalised_category="V45",
        raw_category="V45",
    )
    valid_runner.time_str = "00:43:10"
    valid_runner.time_seconds = 2590.0
    valid_runner.source_row = 9

    race_data = {
        1: [valid_runner, invalid_runner],
        2: [qry_runner],
    }

    df = build_time_qry_todo_df(race_data)

    assert len(df) == 2
    assert list(df["Name"]) == ["Jordan Zero", "Taylor Query"]
    assert list(df["Status"]) == ["Invalid", "QRY"]


def test_sanitise_df_for_export_prefixes_formula_values():
    df = _sanitise_df_for_export(
        pd.DataFrame({"Value": ["=1+1", "+2", "-3", "@cmd", "safe"]})
    )

    assert df.loc[0, "Value"] == "'=1+1"
    assert df.loc[1, "Value"] == "'+2"
    assert df.loc[2, "Value"] == "'-3"
    assert df.loc[3, "Value"] == "'@cmd"
    assert df.loc[4, "Value"] == "safe"


def test_write_unrecognised_clubs_creates_sheet(tmp_path):
    filepath = tmp_path / "unused_clubs.xlsx"
    write_unrecognised_clubs(
        [
            type("U", (), {"raw_club_name": "Guest Club", "occurrences": 2}),
        ],
        filepath,
    )

    wb = load_workbook(filepath, read_only=True, data_only=True)
    assert wb.sheetnames == ["Unused Clubs"]
    ws = wb.active
    assert [cell.value for cell in ws[1]] == ["Raw Club Name", "Occurrences", "Action"]
    assert [cell.value for cell in ws[2]] == ["Guest Club", 2, "Excluded"]


def test_write_category_mismatch_todo_writes_header_only_when_empty(tmp_path):
    filepath = tmp_path / "category_mismatch.xlsx"
    write_category_mismatch_todo({}, filepath)

    wb = load_workbook(filepath, read_only=True, data_only=True)
    assert wb.sheetnames == ["Category Mismatch TODO"]
    ws = wb.active
    assert [cell.value for cell in ws[1]] == [
        "Issue Type",
        "Name",
        "Club",
        "Gender",
        "Races Completed",
        "Categories Seen",
        "Clubs Seen",
        "Suggested Category",
        "Suggested Club",
        "Race Category Sequence",
        "Race Club Sequence",
        "Source Category Values",
        "Source Club Values",
        "Next Step",
    ]


def test_write_time_qry_todo_writes_header_only_when_empty(tmp_path):
    filepath = tmp_path / "time_qry.xlsx"
    write_time_qry_todo({}, filepath)

    wb = load_workbook(filepath, read_only=True, data_only=True)
    assert wb.sheetnames == ["Time QRY TODO"]
    ws = wb.active
    assert [cell.value for cell in ws[1]] == [
        "Issue Type",
        "Race",
        "Source Row",
        "Name",
        "Raw Club",
        "Club",
        "Gender",
        "Category",
        "Current Time",
        "Status",
        "Next Step",
    ]


def test_is_race_over_5k_name_scope_for_results_sheets():
    assert _is_race_over_5k_name("Race 2 - Corsham 10k 2025") is True
    assert _is_race_over_5k_name("Race 1 - Highworth 5 Mile 2025") is True
    assert _is_race_over_5k_name("Race 7 - Chippenham Half Marathon 2025") is True
    assert _is_race_over_5k_name("Race 3 - Westbury 5k") is False
