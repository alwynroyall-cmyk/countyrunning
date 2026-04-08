from pathlib import Path

from openpyxl import load_workbook

from league_scorer.name_lookup import append_name_corrections, load_name_corrections


def test_append_name_corrections_creates_lookup_workbook(tmp_path: Path):
    lookup_path = tmp_path / "name_corrections.xlsx"

    result = append_name_corrections(
        lookup_path,
        [
            {
                "current_name": "Liz Smith",
                "proposed_name": "Elizabeth Smith",
            }
        ],
    )

    assert result["written"] == 1
    assert result["skipped_existing"] == 0
    assert result["skipped_conflicts"] == 0
    assert lookup_path.exists()
    assert load_name_corrections(lookup_path) == {"liz smith": "Elizabeth Smith"}

    workbook = load_workbook(lookup_path)
    try:
        worksheet = workbook.active
        assert worksheet.title == "Name Corrections"
        assert worksheet.cell(1, 1).value == "Raw Name"
        assert worksheet.cell(1, 2).value == "Preferred Name"
    finally:
        workbook.close()


def test_append_name_corrections_skips_conflicting_existing_mapping(tmp_path: Path):
    lookup_path = tmp_path / "name_corrections.xlsx"
    append_name_corrections(
        lookup_path,
        [
            {
                "current_name": "Liz Smith",
                "proposed_name": "Elizabeth Smith",
            }
        ],
    )

    result = append_name_corrections(
        lookup_path,
        [
            {
                "current_name": "Liz Smith",
                "proposed_name": "Beth Smith",
            }
        ],
    )

    assert result["written"] == 0
    assert result["skipped_existing"] == 0
    assert result["skipped_conflicts"] == 1
    assert load_name_corrections(lookup_path) == {"liz smith": "Elizabeth Smith"}
