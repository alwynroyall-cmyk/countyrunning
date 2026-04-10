import openpyxl
from openpyxl import Workbook
from pathlib import Path

from league_scorer.issue_resolution_service import (
    _find_time_column,
    _normalise_time_value,
    _parse_race_num,
    _parse_row_num,
    quick_fix_prompt,
    supports_quick_fix,
)


def test_supports_quick_fix_and_prompt():
    assert supports_quick_fix("AUD-ROW-002") is True
    assert supports_quick_fix("AUD-ROW-001") is False

    title, prompt = quick_fix_prompt("AUD-ROW-002", "Alice")
    assert "Fix Invalid Time" in title
    assert "Alice" in prompt

    title, prompt = quick_fix_prompt("UNKNOWN", "Alice")
    assert title == "Quick Fix"


def test_normalise_time_value_formats_valid_times():
    assert _normalise_time_value("1:02:03") == "01:02:03"
    assert _normalise_time_value("00:00:30") == "00:00:30"
    assert _normalise_time_value("") is None
    assert _normalise_time_value("bad") is None


def test_parse_race_and_row_numbers():
    assert _parse_race_num("5") == 5
    assert _parse_race_num("Race 7") == 7
    assert _parse_race_num("") is None

    assert _parse_row_num("8") == 8
    assert _parse_row_num("8.0") == 8
    assert _parse_row_num("bad") is None


def test_find_time_column_prefers_chip_time():
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Chip Time", "Time"])
    assert _find_time_column(ws) == 2

    wb2 = Workbook()
    ws2 = wb2.active
    ws2.append(["Time", "Name"])
    assert _find_time_column(ws2) == 1

    wb3 = Workbook()
    ws3 = wb3.active
    ws3.append(["Start Time", "Name"])
    assert _find_time_column(ws3) == 1
