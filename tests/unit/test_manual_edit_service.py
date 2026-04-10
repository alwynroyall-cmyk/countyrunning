import os
from pathlib import Path

import openpyxl
from openpyxl import Workbook

from league_scorer.manual_edit_service import (
    _atomic_save,
    _is_raw_data_file,
    apply_club_suggestions,
    resolve_runner_field_across_files,
)


def test_atomic_save_replaces_file(tmp_path: Path) -> None:
    workbook = Workbook()
    ws = workbook.active
    ws.append(["Name", "Club"])
    filepath = tmp_path / "data.xlsx"

    _atomic_save(workbook, filepath)
    assert filepath.exists()

    workbook2 = Workbook()
    _atomic_save(workbook2, filepath)
    assert filepath.exists()


def test_is_raw_data_file_detects_raw_data_folder(tmp_path: Path) -> None:
    path = tmp_path / "inputs" / "raw_data" / "race.xlsx"
    assert _is_raw_data_file(path)
    assert not _is_raw_data_file(tmp_path / "outputs" / "race.xlsx")


def test_resolve_runner_field_across_files_updates_time(tmp_path: Path) -> None:
    raw_data = tmp_path / "inputs" / "raw_data"
    raw_data.mkdir(parents=True)
    filepath = raw_data / "Race #1.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Chip Time"])
    ws.append(["Alice", "00:40:00"])
    wb.save(filepath)

    updated_rows, touched_files, changes, failures = resolve_runner_field_across_files(
        {1: filepath},
        selected_runner="Alice",
        field_type="time",
        target_value="00:42:00",
    )

    assert updated_rows == 1
    assert touched_files == 1
    assert len(changes) == 1
    assert not failures

    reloaded = openpyxl.load_workbook(filepath)
    assert reloaded.active.cell(row=2, column=2).value == "00:42:00"


def test_apply_club_suggestions_updates_club_cells(tmp_path: Path) -> None:
    raw_data = tmp_path / "inputs" / "raw_data"
    raw_data.mkdir(parents=True)
    filepath = raw_data / "Race #1.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["Name", "Club"])
    ws.append(["Alice", "Old Club"])
    wb.save(filepath)

    applied, audit_changes, failures = apply_club_suggestions(
        {
            filepath: [
                {
                    "row_idx": 2,
                    "name": "Alice",
                    "suggested_club": "New Club",
                }
            ]
        }
    )

    assert applied == 1
    assert len(audit_changes) == 1
    assert not failures
    updated = openpyxl.load_workbook(filepath)
    assert updated.active.cell(row=2, column=2).value == "New Club"
