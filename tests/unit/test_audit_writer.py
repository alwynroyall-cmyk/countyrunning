from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from league_scorer.audit_writer import write_audit_workbook


def test_write_audit_workbook_creates_file_with_sheets(tmp_path: Path) -> None:
    workbook_path = tmp_path / "audit.xlsx"
    sheets = {
        "Actionable Issues": pd.DataFrame(
            [
                {"Type": "Row", "Issue Code": "AUD-ROW-001", "Message": "Missing gender"}
            ]
        ),
        "Race Audit Summary": pd.DataFrame(
            [{"Race": 1, "Status": "Open"}]
        ),
    }

    write_audit_workbook(sheets, workbook_path)

    assert workbook_path.exists()
    workbook = load_workbook(workbook_path)
    assert workbook.sheetnames == ["Actionable Issues", "Race Audit Summary"]

    header_cell = workbook["Actionable Issues"]["A1"]
    assert header_cell.value == "Type"
    assert header_cell.fill.patternType == "solid"
    assert header_cell.font.bold is True
    assert header_cell.fill.fgColor.rgb in {"003A4658", "FF3A4658", "3A4658"}


def test_write_audit_workbook_reuses_existing_directory(tmp_path: Path) -> None:
    nested_dir = tmp_path / "outputs" / "audit" / "workbooks"
    workbook_path = nested_dir / "season.xlsx"
    sheets = {"Sheet1": pd.DataFrame([{"A": 1}])}

    write_audit_workbook(sheets, workbook_path)

    assert workbook_path.exists()
    assert workbook_path.parent == nested_dir
